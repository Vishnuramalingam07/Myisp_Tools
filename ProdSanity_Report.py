"""
Azure DevOps Test Plan Execution Report Generator - Custom Format (OPTIMIZED)
"""

import requests
import json
import re
from datetime import datetime
import base64
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import shutil
import os
import sys
import io
import openpyxl

# Load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    print("⚠️  python-dotenv not installed. Install with: pip install python-dotenv")
    print("   Using default configuration values...")

# Import database utilities
try:
    from database_utils import save_test_data_to_db, test_database_connection
    DB_AVAILABLE = True
except ImportError:
    DB_AVAILABLE = False
    print("⚠️  Database utilities not available. Install psycopg2: pip install psycopg2-binary")

# Fix console encoding for Windows to support emojis
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace', line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace', line_buffering=True)

# ============================================================================
# HELPERS
# ============================================================================

def clean_display_name(name):
    """Remove trailing <guid\\email> fragment that Azure DevOps sometimes appends
    to display names, e.g. 'Hajara Ayyubkhan <e079...\\hajara@accenture.com>'."""
    if name and '<' in name:
        return name[:name.index('<')].strip()
    return name

# ============================================================================
# CONFIGURATION
# ============================================================================

ADO_CONFIG = {
    'organization': os.getenv('ADO_ORGANIZATION', 'accenturecio08'),
    'project': os.getenv('ADO_PROJECT', 'AutomationProcess_29697'),
    'plan_id': os.getenv('ADO_PLAN_ID', '4444223'),
    'suite_id': os.getenv('ADO_SUITE_ID', '4486314'),  # Prod Execution suite ID
    'prod_sanity_suite_id': os.getenv('ADO_PROD_SANITY_SUITE_ID', '4486508'),  # Prod Sanity Scenarios suite ID
    'prod_sanity_suite_name': os.getenv('ADO_PROD_SANITY_SUITE_NAME', 'Prod Sanity Scenarios'),
    'insprint_suite_id': os.getenv('ADO_INSPRINT_SUITE_ID', '4447760'),  # Insprint US Prod Status suite ID
    'insprint_suite_name': os.getenv('ADO_INSPRINT_SUITE_NAME', 'Insprint US Prod Status'),
    'target_suite_name': os.getenv('ADO_TARGET_SUITE_NAME', 'Prod Execution'),  # Target suite name to search for
    'pat_token': os.getenv('ADO_PAT'),  # ⚠️ REQUIRED: Set in .env file or environment variable
    'max_workers': int(os.getenv('ADO_MAX_WORKERS', '10')),  # Parallel API calls
}

# ============================================================================
# AZURE DEVOPS API CLIENT (OPTIMIZED)
# ============================================================================

class AzureDevOpsClient:
    def __init__(self, config):
        self.org = config['organization']
        self.project = config['project']    
        self.plan_id = config['plan_id']
        self.suite_id = config['suite_id']
        self.target_suite_name = config.get('target_suite_name', 'Prod Execution')
        self.pat = config['pat_token']
        self.max_workers = config.get('max_workers', 10)
        
        # Encode PAT for Basic Auth
        auth_bytes = f":{self.pat}".encode('ascii')
        encoded_pat = base64.b64encode(auth_bytes).decode('ascii')
        
        self.headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Basic {encoded_pat}',
            'Accept': 'application/json'
        }
        
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        
        # Cache for work items to avoid duplicate fetches
        self.work_item_cache = {}
        self.cache_lock = threading.Lock()
    
    def test_connection(self):
        """Test basic connection"""
        try:
            url = f"https://dev.azure.com/{self.org}/_apis/projects/{self.project}?api-version=7.0"
            print(f"\n🔍 Testing Connection...")
            
            response = self.session.get(url, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                print(f"   ✅ Connected to: {data.get('name', 'Unknown')}")
                return True
            else:
                print(f"   ❌ Failed: {response.status_code}")
                return False
                
        except Exception as e:
            print(f"   ❌ Error: {e}")
            return False
    
    def get_test_plan(self):
        """Fetch test plan details"""
        try:
            url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/testplan/plans/{self.plan_id}?api-version=7.0"
            
            print(f"\n📋 Fetching Test Plan {self.plan_id}...")
            
            response = self.session.get(url, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                print(f"   ✅ Plan: {data.get('name', 'N/A')}")
                return data
            else:
                url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/test/plans/{self.plan_id}?api-version=7.0"
                response = self.session.get(url, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    print(f"   ✅ Plan: {data.get('name', 'N/A')}")
                    return data
                    
                print(f"   ❌ Failed: {response.status_code}")
                return None
                
        except Exception as e:
            print(f"   ❌ Error: {e}")
            return None
    
    def get_all_suites_in_plan(self):
        """Get all suites in the plan with pagination support"""
        try:
            url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/testplan/plans/{self.plan_id}/suites?api-version=7.0"
            
            print(f"\n📦 Fetching All Suites...")
            
            all_suites = []
            continuation_token = None
            
            while True:
                request_url = url
                if continuation_token:
                    request_url = f"{url}&continuationToken={continuation_token}"
                
                response = self.session.get(request_url, timeout=30)
                
                if response.status_code != 200:
                    url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/test/plans/{self.plan_id}/suites?api-version=7.0"
                    response = self.session.get(url, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    suites = data.get('value', [])
                    all_suites.extend(suites)
                    
                    continuation_token = response.headers.get('x-ms-continuationtoken')
                    if not continuation_token:
                        break
                else:
                    break
            
            print(f"   ✅ Found {len(all_suites)} suites")
            return all_suites
                
        except Exception as e:
            print(f"   ❌ Error: {e}")
            return []
    
    def verify_suite_exists(self, suite_id):
        """Verify if a specific suite exists"""
        try:
            url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/test/plans/{self.plan_id}/suites/{suite_id}?api-version=7.0"
            
            print(f"\n🔍 Verifying Suite {suite_id}...")
            
            response = self.session.get(url, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                suite_name = data.get('name', 'Unknown')
                print(f"   ✅ Suite exists: {suite_name}")
                return data
            else:
                print(f"   ❌ Suite not accessible (Status: {response.status_code})")
                return None
                
        except Exception as e:
            print(f"   ❌ Error: {e}")
            return None
    
    def get_test_points_from_suite(self, suite_id):
        """Fetch test points from a suite"""
        try:
            url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/test/plans/{self.plan_id}/suites/{suite_id}/points?api-version=7.0"
            
            response = self.session.get(url, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                return data.get('value', [])
            
            return []
                
        except Exception as e:
            return []

    def get_tc_bug_map_from_plan_runs(self, suite_tc_ids=None):
        """Fetch bugs from associatedBugs on test results for every run in this plan.

        The classic test-points API does NOT expose run/result IDs, so we cannot
        derive them from test point responses.  The reliable path is:
          1. GET all test runs for the plan
          2. For each run: GET results?detailsToInclude=WorkItems
          3. Collect associatedBugs per test-case ID

        Args:
            suite_tc_ids: optional set/list of tc IDs to restrict results
                          (avoids storing bugs for unrelated test cases)

        Returns:
            dict {test_case_id (int) -> [bug_id (int), ...]}
        """
        tc_bug_map = defaultdict(list)
        filter_ids = set(int(i) for i in suite_tc_ids) if suite_tc_ids else None

        # Step A: get all runs for the plan, cap at most recent 200 to avoid
        # making thousands of sequential API calls for stale historical runs.
        MAX_RUNS = 200
        runs_url = (
            f"https://dev.azure.com/{self.org}/{self.project}"
            f"/_apis/test/runs?planId={self.plan_id}&includeRunDetails=true"
            f"&api-version=7.0"
        )
        try:
            resp = self.session.get(runs_url, timeout=30)
            if resp.status_code != 200:
                print(f"      ⚠️  Could not fetch runs for plan {self.plan_id}: {resp.status_code}")
                return {}
            all_runs = resp.json().get('value', [])
            # Sort by most recent (highest run ID) and cap to MAX_RUNS
            all_runs.sort(key=lambda r: r.get('id', 0), reverse=True)
            runs = all_runs[:MAX_RUNS]
            print(f"      ✓ {len(all_runs)} run(s) found for plan {self.plan_id} — processing most recent {len(runs)}")
        except Exception as e:
            print(f"      ⚠️  Error fetching runs: {e}")
            return {}

        # Step B: fetch results with WorkItem details in parallel
        def _fetch_run_results(run):
            run_id = run.get('id')
            if not run_id:
                return []
            try:
                results_url = (
                    f"https://dev.azure.com/{self.org}/{self.project}"
                    f"/_apis/test/runs/{run_id}/results"
                    f"?detailsToInclude=WorkItems&api-version=7.0"
                )
                r = self.session.get(results_url, timeout=30)
                if r.status_code != 200:
                    return []
                return r.json().get('value', [])
            except Exception:
                return []

        collected = []
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(_fetch_run_results, run): run for run in runs}
            done_count = 0
            for future in as_completed(futures):
                done_count += 1
                if done_count % 50 == 0:
                    print(f"      Progress: {done_count}/{len(runs)} runs processed")
                collected.extend(future.result())

        for result in collected:
            tc_ref = result.get('testCase', {})
            raw_tc_id = tc_ref.get('id')
            if not raw_tc_id:
                continue
            tc_id = int(raw_tc_id)
            if filter_ids and tc_id not in filter_ids:
                continue
            for bug_ref in result.get('associatedBugs', []):
                bug_id = bug_ref.get('id')
                if bug_id is not None and int(bug_id) not in tc_bug_map[tc_id]:
                    tc_bug_map[tc_id].append(int(bug_id))

        return dict(tc_bug_map)
    
    def get_work_items_batch(self, work_item_ids, expand_relations=False):
        """Fetch multiple work items in a single API call (OPTIMIZED)"""
        if not work_item_ids:
            return {}
        
        try:
            # Remove duplicates and None values
            unique_ids = list(set([id for id in work_item_ids if id]))
            
            if not unique_ids:
                return {}
            
            # Azure DevOps supports batch requests with up to 200 IDs
            batch_size = 200
            all_work_items = {}
            
            for i in range(0, len(unique_ids), batch_size):
                batch_ids = unique_ids[i:i+batch_size]
                ids_param = ','.join(map(str, batch_ids))
                expand = "&$expand=relations" if expand_relations else ""
                url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/wit/workitems?ids={ids_param}{expand}&api-version=7.0"
                
                response = self.session.get(url, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    work_items = data.get('value', [])
                    
                    for wi in work_items:
                        wi_id = wi.get('id')
                        all_work_items[wi_id] = wi
            
            return all_work_items
                
        except Exception as e:
            print(f"      ⚠️  Batch fetch error: {e}")
            return {}
    
    def get_child_suites_from_cache(self, parent_suite_id, all_suites):
        """Get child suites from cached all_suites list (OPTIMIZED)"""
        try:
            child_suites = []
            parent_suite_id_int = int(parent_suite_id)
            
            for suite in all_suites:
                parent = suite.get('parent', {}) or suite.get('parentSuite', {})
                parent_id = None
                
                if isinstance(parent, dict):
                    parent_id = parent.get('id')
                elif isinstance(parent, (int, str)):
                    try:
                        parent_id = int(parent)
                    except:
                        pass
                
                try:
                    parent_id_int = int(parent_id) if parent_id else None
                except:
                    parent_id_int = parent_id
                
                if parent_id_int == parent_suite_id_int:
                    child_suites.append(suite)
            
            return child_suites
                
        except Exception as e:
            return []
    
    def _build_suite_tree(self, root_suite_id, root_suite_name, all_suites):
        """Build complete suite tree with metadata"""
        suite_tree = []
        
        def traverse_suite(suite_id, suite_name, parent_lead=None, parent_module=None, test_type=None, depth=0):
            """
            Traverse suite hierarchy:
            - Depth 0: Root (Prod Execution)
            - Depth 1: Lead folders (Kavi, Pirtheebaa, etc.)
            - Depth 2: Module folders (SI OCP, SI DCTA, etc.)
            - Depth 3: Test Type folders (Automation, Manual)
            - Depth 4+: Deeper nesting
            """
            
            # Determine current level type
            current_lead = parent_lead
            current_module = parent_module
            current_test_type = test_type
            
            if depth == 1:
                # This is a Lead folder
                current_lead = suite_name
                current_module = None
                current_test_type = None
            elif depth == 2:
                # This is a Module folder
                current_module = suite_name
                current_test_type = None
            elif depth == 3:
                # This is Test Type folder (Automation or Manual)
                suite_name_lower = suite_name.lower()
                if 'automation' in suite_name_lower:
                    current_test_type = 'Automation'
                elif 'manual' in suite_name_lower:
                    current_test_type = 'Manual'    
                else:
                    # If not explicitly named, keep parent test type
                    current_test_type = test_type
            # depth >= 4: keep parent values
            
            suite_tree.append({
                'id': suite_id,
                'name': suite_name,
                'parent_lead': current_lead,
                'parent_module': current_module,
                'test_type': current_test_type,
                'depth': depth
            })
            
            # Get children
            children = self.get_child_suites_from_cache(suite_id, all_suites)
            
            for child in children:
                child_id = child.get('id')
                child_name = child.get('name', 'Unknown')
                
                traverse_suite(child_id, child_name, 
                             parent_lead=current_lead, 
                             parent_module=current_module,
                             test_type=current_test_type,
                             depth=depth+1)
        
        traverse_suite(root_suite_id, root_suite_name)
        return suite_tree
    
    def _collect_test_points_parallel(self, suite_tree):
        """Collect test points from all suites in parallel"""
        all_test_items = []
        
        def fetch_suite_tests(suite_info):
            suite_id = suite_info['id']
            suite_name = suite_info['name']
            parent_lead = suite_info['parent_lead']
            parent_module = suite_info['parent_module']
            test_type = suite_info['test_type']
            
            test_items = self.get_test_points_from_suite(suite_id)
            
            result = []
            for item in test_items:
                # Pass lead, module, and test type information
                result.append((item, suite_name, parent_lead, parent_module, test_type))
            
            return result
        
        # Use thread pool for parallel fetching
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(fetch_suite_tests, suite): suite for suite in suite_tree}
            
            completed = 0
            total = len(suite_tree)
            
            for future in as_completed(futures):
                completed += 1
                if completed % 10 == 0 or completed == total:
                    print(f"      Progress: {completed}/{total} suites processed", end='\r')
                
                try:
                    items = future.result()
                    all_test_items.extend(items)
                except Exception as e:
                    pass
        
        print()  # New line after progress
        return all_test_items

    def get_all_test_data_from_suite(self, suite_id, suite_name):
        """Get all test data from suite and children (OPTIMIZED)"""
        print(f"\n📊 Collecting Test Data from: {suite_name}...")
        print(f"   ⚡ Using optimized batch processing...")
        
        # Fetch all suites once
        print(f"\n   Step 1: Fetching suite hierarchy...")
        all_suites = self.get_all_suites_in_plan()
        
        # Build suite hierarchy
        print(f"   Step 2: Building suite tree...")
        suite_tree = self._build_suite_tree(suite_id, suite_name, all_suites)
        
        # Collect all test points in parallel
        print(f"   Step 3: Collecting test points from {len(suite_tree)} suites...")
        all_test_items = self._collect_test_points_parallel(suite_tree)
        
        if not all_test_items:
            print(f"   ⚠️  No test points found")
            return []
        
        print(f"   ✓ Found {len(all_test_items)} test items")

        # Extract unique work item IDs
        print(f"   Step 4: Extracting work item IDs...")
        work_item_ids = []
        for item_data in all_test_items:
            # Unpack the tuple with 5 elements now
            item, item_suite_name, parent_lead, parent_module, test_type = item_data

            test_case = item.get('testCase', {})
            work_item_id = test_case.get('id') or item.get('testCaseReference', {}).get('id')
            if work_item_id:
                try:
                    work_item_ids.append(int(work_item_id))
                except (TypeError, ValueError):
                    work_item_ids.append(work_item_id)

        # Step 4b: Fetch associated bugs from test run results.
        # The test-points API does NOT expose run/result IDs, so we query plan
        # runs directly and collect associatedBugs from each result.
        print(f"   Step 4b: Fetching bugs from test run results (associatedBugs)...")
        result_level_bugs = self.get_tc_bug_map_from_plan_runs(
            suite_tc_ids=set(work_item_ids)
        )
        rl_count = sum(len(v) for v in result_level_bugs.values())
        print(f"   ✓ Found {rl_count} result-level bug link(s) across {len(result_level_bugs)} test case(s)")
        if not result_level_bugs:
            print(f"   ⚠️  No result-level bugs found — check that test runs exist for plan {self.plan_id}")
        
        # Fetch all work items in batch (with relations to find linked User Stories)
        print(f"   Step 5: Fetching {len(set(work_item_ids))} unique work items in batch...")
        work_items_dict = self.get_work_items_batch(work_item_ids, expand_relations=True)
        print(f"   ✓ Retrieved {len(work_items_dict)} work items")

        # Step 5b: Collect all external relation IDs (requirements + bugs in one batch)
        print(f"   Step 5b: Finding linked requirements and bugs...")
        test_to_potential_req_ids = defaultdict(list)
        test_to_all_related_ids = defaultdict(list)
        all_external_ids = set()

        for wi_id, wi in work_items_dict.items():
            for rel in wi.get('relations', []):
                rel_type = rel.get('rel', '')
                rel_url = rel.get('url', '')
                try:
                    linked_id = int(rel_url.split('/')[-1])
                    if linked_id not in work_items_dict:
                        all_external_ids.add(linked_id)
                        # Store (id, rel_type) so Step 5c can filter by link type
                        test_to_all_related_ids[wi_id].append((linked_id, rel_type))
                        if any(t in rel_type for t in ['Hierarchy-Reverse', 'TestedBy', 'Tests']):
                            test_to_potential_req_ids[wi_id].append(linked_id)
                except Exception:
                    pass

        # Single batch fetch for all external related items
        external_items_dict = {}
        if all_external_ids:
            external_items_dict = self.get_work_items_batch(list(all_external_ids))

        # Build requirements dict from the batch
        req_types = {'User Story', 'Product Backlog Item', 'Requirement', 'Feature', 'Epic'}
        test_id_to_req_id = {}
        requirements_dict = {}
        for wi_id, related_ids in test_to_potential_req_ids.items():
            for related_id in related_ids:
                ext_wi = external_items_dict.get(related_id)
                if ext_wi:
                    wi_type = ext_wi.get('fields', {}).get('System.WorkItemType', '')
                    if wi_type in req_types:
                        requirements_dict[related_id] = ext_wi
                        if wi_id not in test_id_to_req_id:
                            test_id_to_req_id[wi_id] = related_id
                            break
        if requirements_dict:
            print(f"   ✓ Found {len(requirements_dict)} linked requirements")
        else:
            print(f"   ⚠️  No linked requirements found in relations")

        # Step 5c: Build linked bugs map  test_id -> [bug_ids]
        # Source 1: ANY work-item relation that points to a Bug (TestedBy, Related,
        #           Tests, Duplicate, ArtifactLink, etc.) — covers all ADO link styles
        # Source 2 (primary): bugs attached directly to the test result (associatedBugs)
        print(f"   Step 5c: Merging test-result bugs + work-item relation bugs...")
        test_id_to_bug_list = defaultdict(list)

        # -- Source 1: any relation whose target work item type is 'Bug' --
        for wi_id, related_id_rel_pairs in test_to_all_related_ids.items():
            for related_id, rel_type in related_id_rel_pairs:
                ext_wi = external_items_dict.get(related_id)
                if ext_wi:
                    wi_type = ext_wi.get('fields', {}).get('System.WorkItemType', '')
                    if wi_type == 'Bug' and related_id not in test_id_to_bug_list[wi_id]:
                        test_id_to_bug_list[wi_id].append(related_id)

        # -- Source 2: bugs from test results (takes priority / merged) --
        for tc_id, bug_ids in result_level_bugs.items():
            for bug_id in bug_ids:
                if bug_id not in test_id_to_bug_list[tc_id]:
                    test_id_to_bug_list[tc_id].insert(0, bug_id)  # result bugs listed first

        linked_bug_count = sum(len(v) for v in test_id_to_bug_list.values())
        print(f"   ✓ Total unique bug links: {linked_bug_count} across {len(test_id_to_bug_list)} test case(s)")

        # Step 5d: Fetch details for result-level bugs not yet in external_items_dict
        # (those bugs were added via associatedBugs and were never fetched as work items)
        missing_bug_ids = list({
            bid
            for bids in test_id_to_bug_list.values()
            for bid in bids
            if bid not in external_items_dict
        })
        if missing_bug_ids:
            print(f"   Step 5d: Fetching details for {len(missing_bug_ids)} result-level bug(s) not in relations...")
            missing_wi = self.get_work_items_batch(missing_bug_ids)
            external_items_dict.update(missing_wi)
            print(f"   ✓ Fetched {len(missing_wi)} additional bug work items")

        # Process test data
        print(f"   Step 6: Processing test data...")
        all_test_data = []

        for item_data in all_test_items:
            # Unpack the tuple with 5 elements now
            item, item_suite_name, parent_lead, parent_module, suite_test_type = item_data
            
            test_case = item.get('testCase', {})
            work_item_id = test_case.get('id') or item.get('testCaseReference', {}).get('id')
            # Normalize to int so lookups into dicts with int keys (work_items_dict,
            # test_id_to_bug_list, test_id_to_req_id) succeed regardless of API return type
            try:
                work_item_id = int(work_item_id) if work_item_id is not None else None
            except (TypeError, ValueError):
                pass

            # Determine test type based on suite folder structure
            if suite_test_type:
                # Use test type from suite hierarchy (Automation or Manual folder)
                test_type = suite_test_type
            else:
                # Fallback: Check work item tags/automation status
                work_item = work_items_dict.get(work_item_id) if work_item_id else None
                test_type = 'Manual'  # Default
                
                if work_item:
                    fields = work_item.get('fields', {})
                    tags = fields.get('System.Tags', '')
                    automation_status = fields.get('Microsoft.VSTS.TCM.AutomationStatus', '')
                    
                    if 'automation' in str(tags).lower() or 'automated' in str(automation_status).lower():
                        test_type = 'Automation'
            
            # Get outcome
            outcome = 'Not Run'
            if 'results' in item:
                outcome = item['results'].get('outcome', 'Not Run')
            elif 'lastResultOutcome' in item:
                outcome = item.get('lastResultOutcome', 'Not Run')
            elif 'outcome' in item:
                outcome = item.get('outcome', 'Not Run')
            # Normalize ADO "Active" / "Unspecified" / empty → "Not Run"
            if str(outcome).strip().lower() in ('active', 'unspecified', 'none', ''):
                outcome = 'Not Run'
            
            # Get assigned to
            assigned_to = 'Unassigned'
            if 'assignedTo' in item and item['assignedTo']:
                assigned_to = clean_display_name(item['assignedTo'].get('displayName', 'Unassigned'))
            
            # Determine final lead and module names
            final_lead = parent_lead if parent_lead else 'Unassigned'
            final_module = parent_module if parent_module else item_suite_name
            
            # Override with assigned user's first name if available (only if no parent_lead)
            if assigned_to and assigned_to != 'Unassigned' and not parent_lead:
                try:
                    final_lead = assigned_to.split()[0]
                except:
                    final_lead = assigned_to
            
            # Get extra fields from work item (Bug ID, Comments)
            work_item = work_items_dict.get(work_item_id) if work_item_id else None
            wi_fields = work_item.get('fields', {}) if work_item else {}
            # Prefer linked bug IDs discovered via ADO relations; fall back to custom fields
            linked_bug_ids = test_id_to_bug_list.get(work_item_id, [])
            if linked_bug_ids:
                bug_id = ', '.join(str(b) for b in sorted(linked_bug_ids))
            else:
                bug_id = (
                    wi_fields.get('Custom.BugID') or
                    wi_fields.get('Custom.BugId') or
                    wi_fields.get('Custom.RelatedBugId') or
                    wi_fields.get('Custom.DefectID') or
                    ''
                )
            # Build detailed bug list for State / Severity / Node Name columns
            bug_details = []
            for _bid in (linked_bug_ids if linked_bug_ids else []):
                _bwi = external_items_dict.get(_bid)
                if _bwi:
                    _bf = _bwi.get('fields', {})
                    _ap = _bf.get('System.AreaPath', '')
                    _nn = _ap.split('\\')[-1] if '\\' in _ap else _ap
                    bug_details.append({
                        'bug_id':   _bid,
                        'state':    _bf.get('System.State', ''),
                        'severity': _bf.get('Microsoft.VSTS.Common.Severity', ''),
                        'node_name': _nn,
                    })
            comments = (
                wi_fields.get('Custom.Comments') or
                wi_fields.get('Custom.Comment') or
                wi_fields.get('Custom.TestComments') or
                ''
            )
            # Use only the exact ADO field reference name shown in Test Plans Execute view
            text_verification = wi_fields.get('Custom.TextVerification') or ''
            text_verification1 = wi_fields.get('Custom.TextVerification1') or ''
            # Get linked requirement (User Story) title and ID from relations
            req_wi_id = test_id_to_req_id.get(work_item_id)
            req_wi = requirements_dict.get(req_wi_id) if req_wi_id else None
            if req_wi:
                req_wi_fields = req_wi.get('fields', {})
                us_title = req_wi_fields.get('System.Title', '')
                us_id = req_wi_id
            else:
                us_id = (
                    wi_fields.get('Custom.USID') or
                    wi_fields.get('Custom.UserStoryID') or
                    ''
                )
                us_title = (
                    wi_fields.get('Custom.UserStoryTitle') or
                    wi_fields.get('Custom.USTitle') or
                    ''
                )
            prod_sanity_status = wi_fields.get('Custom.ProdSanityStatus', '')
            automation_status = wi_fields.get('Microsoft.VSTS.TCM.AutomationStatus', '')

            all_test_data.append({
                'id': work_item_id or 'N/A',
                'name': test_case.get('name', 'N/A'),
                'suite': item_suite_name,
                'module': final_module,  # Use parent_module from hierarchy
                'lead': final_lead,      # Use parent_lead from hierarchy
                'assigned_to': assigned_to,  # Tester (assigned person)
                'type': test_type,       # Use test type from suite folder structure
                'outcome': outcome,
                'priority': item.get('priority', 2),
                'state': item.get('state', 'Active'),
                'bug_id': bug_id,
                'bug_details': bug_details,
                'comments': comments,
                'text_verification': text_verification,
                'text_verification1': text_verification1,
                'us_id': us_id,
                'us_title': us_title,
                'prod_sanity_status': prod_sanity_status,
                'automation_status': automation_status,
            })
        
        print(f"\n   ✅ Total test items collected: {len(all_test_data)}")
        return all_test_data
    
    def get_bugs_from_query(self, query_id):
        """Fetch bugs from ADO query - supports both flat and hierarchical queries"""
        try:
            url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/wit/wiql/{query_id}?api-version=7.0"
            
            print(f"\n🐛 Fetching Bugs from Query {query_id}...")
            
            response = self.session.get(url, timeout=30)
            
            if response.status_code == 200:
                data = response.json()

                # Handle both flat queries (workItems) and hierarchical queries (workItemRelations)
                if 'workItemRelations' in data:
                    # Hierarchical / "Work Items and Direct Links" query — collect all unique IDs
                    seen_ids = set()
                    for rel in data['workItemRelations']:
                        for node_key in ('source', 'target'):
                            node = rel.get(node_key)
                            if node and node.get('id'):
                                seen_ids.add(node['id'])
                    work_items = [{'id': i} for i in seen_ids]
                    print(f"   ✓ Hierarchical query: extracted {len(work_items)} unique work item(s)")
                else:
                    work_items = data.get('workItems', [])

                if not work_items:
                    print(f"   ⚠️  No bugs found in query")
                    return []
                
                # Extract work item IDs
                bug_ids = [wi['id'] for wi in work_items]
                print(f"   ✓ Found {len(bug_ids)} bugs in query")
                
                # Fetch bug details in batch
                print(f"   Step 2: Fetching bug details...")
                bugs_dict = self.get_work_items_batch(bug_ids)
                
                # Process bug data
                print(f"   Step 3: Processing bug data...")
                bug_data = []
                
                for bug_id, bug_wi in bugs_dict.items():
                    fields = bug_wi.get('fields', {})
                    
                    # Get bug details
                    # Try different possible field names for ExternalRef ID
                    external_ref = (
                        fields.get('Custom.ExternalRefID') or 
                        fields.get('Custom.ExternalRegID') or
                        fields.get('Custom.ExternalRef') or
                        fields.get('ExternalRefID') or
                        fields.get('ExternalRegID') or
                        'Unassigned'
                    )
                    
                    # Extract Node Name from Area Path (last segment after \)
                    area_path = fields.get('System.AreaPath', 'N/A')
                    if area_path != 'N/A' and '\\' in area_path:
                        node_name = area_path.split('\\')[-1]
                    else:
                        node_name = area_path
                    
                    bug_info = {
                        'id': bug_id,
                        'title': fields.get('System.Title', 'N/A'),
                        'state': fields.get('System.State', 'N/A'),
                        'severity': fields.get('Microsoft.VSTS.Common.Severity', 'N/A'),
                        'priority': fields.get('Microsoft.VSTS.Common.Priority', 'N/A'),
                        'assigned_to': clean_display_name(fields.get('System.AssignedTo', {}).get('displayName', 'Unassigned') if isinstance(fields.get('System.AssignedTo'), dict) else 'Unassigned'),
                        'created_date': fields.get('System.CreatedDate', 'N/A'),
                        'mpoc': external_ref,  # ExternalRef ID as MPOC
                        'area_path': area_path,
                        'tags': fields.get('System.Tags', ''),
                        'text_verification': fields.get('Custom.TextVerification', 'N/A'),
                        'defect_record': fields.get('Custom.DefectRecord', 'N/A'),
                        'node_name': node_name,
                        'eta': fields.get('Custom.ETA', 'N/A'),
                        'stage_found': fields.get('Custom.StageFound', 'N/A'),
                        'text_verification1': fields.get('Custom.TextVerification1', 'N/A')
                    }
                    
                    bug_data.append(bug_info)
                
                print(f"   ✅ Processed {len(bug_data)} bugs")
                return bug_data
            else:
                print(f"   ❌ Failed: {response.status_code}")
                return []
                
        except Exception as e:
            print(f"   ❌ Error fetching bugs: {e}")
            return []

    def get_us_bug_map_from_query(self, query_id):
        """Fetch bugs from a WIQL query and build a map of us_id -> list of bug info dicts.
        Handles both flat-list queries (workItems) and Work Items and Direct Links
        queries (workItemRelations). In the hierarchical case the US->Bug linkage
        is read directly from the query response.
        Each bug info dict contains: bug_id, state, severity."""
        try:
            url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/wit/wiql/{query_id}?api-version=7.0"
            print(f"\n🐛 Fetching US-linked Bugs from Query {query_id}...")
            response = self.session.get(url, timeout=30)
            if response.status_code != 200:
                print(f"   ❌ Failed: {response.status_code}")
                return {}
            data = response.json()

            us_to_bug_pairs = []   # [(us_id_int, bug_id_int)]
            all_ids = set()

            # ----------------------------------------------------------------
            # Case 1: "Work Items and Direct Links" query
            # Response shape: { workItemRelations: [{source:{id}, target:{id}, rel}, ...] }
            # source = User Story (or null for top-level), target = Bug
            # ----------------------------------------------------------------
            if 'workItemRelations' in data:
                relations = data['workItemRelations']
                for rel in relations:
                    source = rel.get('source')
                    target = rel.get('target')
                    if source and target:
                        us_to_bug_pairs.append((source['id'], target['id']))
                        all_ids.add(source['id'])
                        all_ids.add(target['id'])
                    elif target:
                        all_ids.add(target['id'])
                print(f"   ✓ Hierarchical query: {len(us_to_bug_pairs)} US→Bug links, {len(all_ids)} total items")

            # ----------------------------------------------------------------
            # Case 2: Flat-list query
            # Response shape: { workItems: [{id}, ...] }
            # Fall back to traversing each bug's ADO relations to find US.
            # ----------------------------------------------------------------
            elif 'workItems' in data:
                work_items = data.get('workItems', [])
                for wi in work_items:
                    all_ids.add(wi['id'])
                print(f"   ✓ Flat query: {len(all_ids)} bugs")
            else:
                print(f"   ⚠️  Unrecognised query response format")
                return {}

            if not all_ids:
                print(f"   ⚠️  No work items found in query")
                return {}

            # Batch-fetch all work items (no relations needed for hierarchical path)
            all_wi_dict = self.get_work_items_batch(list(all_ids))

            us_bug_map = defaultdict(list)

            if us_to_bug_pairs:
                # --- Hierarchical path: use explicit pairs from the query ---
                for us_id, bug_id in us_to_bug_pairs:
                    bug_wi = all_wi_dict.get(bug_id)
                    if bug_wi:
                        fields = bug_wi.get('fields', {})
                        _area_path = fields.get('System.AreaPath', 'N/A')
                        _node_name = _area_path.split('\\')[-1] if _area_path != 'N/A' and '\\' in _area_path else _area_path
                        bug_info = {
                            'bug_id': bug_id,
                            'state': fields.get('System.State', 'N/A'),
                            'severity': fields.get('Microsoft.VSTS.Common.Severity', 'N/A'),
                            'node_name': _node_name,
                        }
                        # Avoid duplicates
                        if not any(b['bug_id'] == bug_id for b in us_bug_map[us_id]):
                            us_bug_map[us_id].append(bug_info)
            else:
                # --- Flat path: traverse each bug's relations to find linked US ---
                us_types = {'User Story', 'Product Backlog Item', 'Requirement', 'Feature', 'Epic'}
                bug_id_set = set(all_ids)

                # Re-fetch with expanded relations
                bugs_with_relations = self.get_work_items_batch(list(all_ids), expand_relations=True)

                all_related_ids = set()
                for bug_wi in bugs_with_relations.values():
                    for rel in bug_wi.get('relations', []):
                        rel_url = rel.get('url', '')
                        try:
                            related_id = int(rel_url.split('/')[-1])
                            if related_id not in bug_id_set:
                                all_related_ids.add(related_id)
                        except Exception:
                            pass

                related_items = {}
                if all_related_ids:
                    related_items = self.get_work_items_batch(list(all_related_ids))

                for bug_id, bug_wi in bugs_with_relations.items():
                    fields = bug_wi.get('fields', {})
                    _area_path = fields.get('System.AreaPath', 'N/A')
                    _node_name = _area_path.split('\\')[-1] if _area_path != 'N/A' and '\\' in _area_path else _area_path
                    bug_info = {
                        'bug_id': bug_id,
                        'state': fields.get('System.State', 'N/A'),
                        'severity': fields.get('Microsoft.VSTS.Common.Severity', 'N/A'),
                        'node_name': _node_name,
                    }
                    for rel in bug_wi.get('relations', []):
                        rel_url = rel.get('url', '')
                        try:
                            related_id = int(rel_url.split('/')[-1])
                            wi = related_items.get(related_id)
                            if wi:
                                wi_type = wi.get('fields', {}).get('System.WorkItemType', '')
                                if wi_type in us_types:
                                    us_bug_map[related_id].append(bug_info)
                        except Exception:
                            pass

            print(f"   ✅ Built US-bug map covering {len(us_bug_map)} User Stories")
            return dict(us_bug_map)
        except Exception as e:
            print(f"   ❌ Error building US-bug map: {e}")
            return {}

    def get_defects_by_tag_and_date(self, tags, created_after_date):
        """
        Fetch defects (bugs) based on tags and creation date
        
        Args:
            tags: List of tags to filter by (e.g., ['Insprint_Regression'])
            created_after_date: Date string in format 'YYYY-MM-DD' (e.g., '2026-02-12')
        
        Returns:
            List of defect/bug information dictionaries
        """
        try:
            # Convert tags to a list if it's a string
            if isinstance(tags, str):
                tags = [tags]
            
            # Build WIQL query
            # Tags in Azure DevOps are stored as a semicolon-separated string
            # We need to check if any of the tags exists in the System.Tags field
            tag_conditions = " OR ".join([f"[System.Tags] CONTAINS '{tag}'" for tag in tags])
            
            wiql_query = f"""
            SELECT [System.Id], [System.Title], [System.State], [System.CreatedDate]
            FROM WorkItems
            WHERE [System.TeamProject] = @project
                AND [System.WorkItemType] = 'Bug'
                AND ({tag_conditions})
                AND [System.CreatedDate] >= '{created_after_date}'
            ORDER BY [System.CreatedDate] DESC
            """
            
            url = f"https://dev.azure.com/{self.org}/{self.project}/_apis/wit/wiql?api-version=7.0"
            
            print(f"\n🐛 Fetching Defects with tags {tags} created after {created_after_date}...")
            
            response = self.session.post(
                url,
                json={"query": wiql_query},
                timeout=30
            )
            
            if response.status_code == 200:
                data = response.json()
                work_items = data.get('workItems', [])
                
                if not work_items:
                    print(f"   ⚠️  No defects found matching criteria")
                    return []
                
                # Extract work item IDs
                defect_ids = [wi['id'] for wi in work_items]
                print(f"   ✓ Found {len(defect_ids)} defects matching criteria")
                
                # Fetch defect details in batch
                print(f"   Step 2: Fetching defect details...")
                defects_dict = self.get_work_items_batch(defect_ids)
                
                # Process defect data
                print(f"   Step 3: Processing defect data...")
                defect_data = []
                
                for defect_id, defect_wi in defects_dict.items():
                    fields = defect_wi.get('fields', {})
                    
                    # Get defect details
                    # Try different possible field names for ExternalRef ID
                    external_ref = (
                        fields.get('Custom.ExternalRefID') or 
                        fields.get('Custom.ExternalRegID') or
                        fields.get('Custom.ExternalRef') or
                        fields.get('ExternalRefID') or
                        fields.get('ExternalRegID') or
                        'Unassigned'
                    )
                    
                    # Extract Node Name from Area Path (last segment after \)
                    area_path = fields.get('System.AreaPath', 'N/A')
                    if area_path != 'N/A' and '\\' in area_path:
                        node_name = area_path.split('\\')[-1]
                    else:
                        node_name = area_path
                    
                    defect_info = {
                        'id': defect_id,
                        'title': fields.get('System.Title', 'N/A'),
                        'state': fields.get('System.State', 'N/A'),
                        'severity': fields.get('Microsoft.VSTS.Common.Severity', 'N/A'),
                        'priority': fields.get('Microsoft.VSTS.Common.Priority', 'N/A'),
                        'assigned_to': clean_display_name(fields.get('System.AssignedTo', {}).get('displayName', 'Unassigned') if isinstance(fields.get('System.AssignedTo'), dict) else 'Unassigned'),
                        'created_date': fields.get('System.CreatedDate', 'N/A'),
                        'mpoc': external_ref,  # ExternalRef ID as MPOC
                        'area_path': area_path,
                        'tags': fields.get('System.Tags', ''),
                        'text_verification': fields.get('Custom.TextVerification', 'N/A'),
                        'defect_record': fields.get('Custom.DefectRecord', 'N/A'),
                        'node_name': node_name,
                        'eta': fields.get('Custom.ETA', 'N/A'),
                        'stage_found': fields.get('Custom.StageFound', 'N/A'),
                        'text_verification1': fields.get('Custom.TextVerification1', 'N/A')
                    }
                    
                    defect_data.append(defect_info)
                
                print(f"   ✅ Processed {len(defect_data)} defects")
                return defect_data
            else:
                print(f"   ❌ Failed: {response.status_code}")
                try:
                    error_detail = response.json()
                    print(f"   ❌ Error details: {error_detail}")
                except:
                    print(f"   ❌ Response: {response.text[:200]}")
                return []
                
        except Exception as e:
            print(f"   ❌ Error fetching defects: {e}")
            import traceback
            traceback.print_exc()
            return []

# ============================================================================
# HTML REPORT GENERATOR - CUSTOM FORMAT
# ============================================================================

class CustomHTMLReportGenerator:
    def __init__(self, test_data, plan_info=None, suite_name=None, bug_data=None, insprint_data=None, insprint_defects=None, prod_sanity_data=None, prod_sanity_defects=None, us_bug_map=None):
        self.test_data = test_data
        self.prod_sanity_data = prod_sanity_data or []
        self.insprint_data = insprint_data or []
        self.plan_info = plan_info or {}
        self.suite_name = suite_name or 'Test Suite'
        self.timestamp = datetime.now().strftime("%B %d, %Y at %H:%M:%S")
        self.bug_data = bug_data or []
        self.insprint_defects = insprint_defects or []
        self.prod_sanity_defects = prod_sanity_defects or []
        self.us_bug_map = us_bug_map or {}
    
    def organize_data_by_lead_module(self):
        """Organize test data by Lead and Module"""
        organized = defaultdict(lambda: defaultdict(lambda: {
            'manual': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0},
            'automation': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0}
        }))
        
        for test in self.test_data:
            lead = test['lead']
            module = test['module']
            test_type = test['type'].lower()
            outcome = test['outcome']
            
            # Increment total count
            organized[lead][module][test_type]['total'] += 1
            
            # Map outcome to correct category (case-insensitive matching)
            outcome_lower = outcome.lower()
            
            if outcome_lower in ['passed', 'pass']:
                organized[lead][module][test_type]['passed'] += 1
            elif outcome_lower in ['failed', 'fail']:
                organized[lead][module][test_type]['failed'] += 1
            elif outcome_lower in ['blocked', 'block']:
                organized[lead][module][test_type]['blocked'] += 1
            elif outcome_lower in ['not applicable', 'na', 'n/a', 'notapplicable']:
                organized[lead][module][test_type]['na'] += 1
            elif outcome_lower in ['not run', 'notrun', 'active', 'none', '']:
                organized[lead][module][test_type]['not_run'] += 1
            else:
                # Any unrecognized outcome goes to 'not_run'
                organized[lead][module][test_type]['not_run'] += 1
        
        return organized
    
    def calculate_percentages(self, data):
        """Calculate Pass% and Execution% for individual Lead/Module rows"""
        passed = data['passed']
        failed = data['failed']
        blocked = data['blocked']
        total = data['total']
        na = data['na']
        
        # Pass % = (Pass / (Pass + Fail + Blocked)) * 100
        denominator_pass = passed + failed + blocked
        pass_percentage = (passed / denominator_pass * 100) if denominator_pass > 0 else 0
        
        # Execution % = (Pass + Fail) / (Total - NA) * 100
        denominator_exec = total - na
        execution_percentage = ((passed + failed) / denominator_exec * 100) if denominator_exec > 0 else 0
        
        return pass_percentage, execution_percentage
    
    def calculate_grand_total_percentages(self, data):
        """Calculate Pass% and Execution% for Grand Total row (different formula)"""
        passed = data['passed']
        failed = data['failed']
        blocked = data['blocked']
        total = data['total']
        na = data['na']
        
        # Pass % = (Pass / (Pass + Fail + Blocked)) * 100
        denominator_pass = passed + failed + blocked
        pass_percentage = (passed / denominator_pass * 100) if denominator_pass > 0 else 0
        
        # Execution % = (Pass + Fail + Blocked) / (Total - NA) * 100
        numerator_exec = passed + failed + blocked
        denominator_exec = total - na
        execution_percentage = (numerator_exec / denominator_exec * 100) if denominator_exec > 0 else 0
        
        return pass_percentage, execution_percentage
    
    def calculate_grand_totals(self, organized_data):
        """Calculate grand totals"""
        grand_totals = {
            'manual': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0},
            'automation': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0}
        }
        
        for lead_data in organized_data.values():
            for module_data in lead_data.values():
                for test_type in ['manual', 'automation']:
                    for key in grand_totals[test_type]:
                        grand_totals[test_type][key] += module_data[test_type][key]
        
        return grand_totals
    
    def process_bug_data_by_mpoc(self):
        """Process bug data and group by MPOC with severity breakdown - includes multiple bug states"""
        bug_summary = defaultdict(lambda: {
            '1 - Critical': 0,
            '2 - High': 0,
            '3 - Medium': 0,
            '4 - Low': 0
        })
        
        # Define allowed bug states (case-insensitive)
        allowed_states = {
            'new', 'active', 'blocked', 'ready to deploy', 'resolved', 
            'ba clarification', 're-open', 'blocked in pt', 'blocked in uat', 'deferred'
        }
        
        # Filter bugs by allowed states
        filtered_bugs = [bug for bug in self.bug_data if bug['state'].lower() in allowed_states]
        
        # Track unique MPOC names for case-insensitive grouping
        mpoc_mapping = {}  # Maps lowercase MPOC to original casing
        
        for bug in filtered_bugs:
            raw_mpoc = bug['mpoc'] if bug['mpoc'] and bug['mpoc'] != 'Unassigned' else 'Unassigned'
            
            # Normalize MPOC name to avoid case duplicates
            mpoc_lower = raw_mpoc.lower()
            
            # Use the first occurrence's casing as the canonical form
            if mpoc_lower not in mpoc_mapping:
                mpoc_mapping[mpoc_lower] = raw_mpoc
            
            mpoc = mpoc_mapping[mpoc_lower]
            severity = bug['severity']
            
            # Map severity to standardized format
            if severity in bug_summary[mpoc]:
                bug_summary[mpoc][severity] += 1
            else:
                # Handle any non-standard severity values
                bug_summary[mpoc]['3 - Medium'] += 1
        
        return dict(bug_summary)
    
    def organize_data_by_lead_module_insprint(self):
        """Organize insprint test data by Lead and Module"""
        organized = defaultdict(lambda: defaultdict(lambda: {
            'manual': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0},
            'automation': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0}
        }))
        
        for test in self.insprint_data:
            lead = test['lead']
            module = test['module']
            test_type = test['type'].lower()
            outcome = test['outcome']
            
            # Increment total count
            organized[lead][module][test_type]['total'] += 1
            
            # Map outcome to correct category (case-insensitive matching)
            outcome_lower = outcome.lower()
            
            if outcome_lower in ['passed', 'pass']:
                organized[lead][module][test_type]['passed'] += 1
            elif outcome_lower in ['failed', 'fail']:
                organized[lead][module][test_type]['failed'] += 1
            elif outcome_lower in ['blocked', 'block']:
                organized[lead][module][test_type]['blocked'] += 1
            elif outcome_lower in ['not applicable', 'na', 'n/a', 'notapplicable']:
                organized[lead][module][test_type]['na'] += 1
            elif outcome_lower in ['not run', 'notrun', 'active', 'none', '']:
                organized[lead][module][test_type]['not_run'] += 1
            else:
                # Any unrecognized outcome goes to 'not_run'
                organized[lead][module][test_type]['not_run'] += 1
        
        return organized

    def organize_data_by_lead_module_prod_sanity(self):
        """Organize Prod Sanity test data by Lead and Module"""
        organized = defaultdict(lambda: defaultdict(lambda: {
            'manual': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0},
            'automation': {'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0}
        }))
        for test in self.prod_sanity_data:
            lead = test['lead']
            module = test['module']
            test_type = test['type'].lower()
            outcome = test['outcome']
            organized[lead][module][test_type]['total'] += 1
            outcome_lower = outcome.lower()
            if outcome_lower in ['passed', 'pass']:
                organized[lead][module][test_type]['passed'] += 1
            elif outcome_lower in ['failed', 'fail']:
                organized[lead][module][test_type]['failed'] += 1
            elif outcome_lower in ['blocked', 'block']:
                organized[lead][module][test_type]['blocked'] += 1
            elif outcome_lower in ['not applicable', 'na', 'n/a', 'notapplicable']:
                organized[lead][module][test_type]['na'] += 1
            elif outcome_lower in ['not run', 'notrun', 'active', 'none', '']:
                organized[lead][module][test_type]['not_run'] += 1
            else:
                organized[lead][module][test_type]['not_run'] += 1
        return organized

    def generate_html(self):
        """Generate HTML report - Compact Design with Filters and Tabs"""
        # Build POD name -> PT Lead mapping dynamically from Excel (Test Lead POC sheet)
        POD_LEAD_MAP = {}
        try:
            _excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'POD details with M POC.xlsx')
            _wb = openpyxl.load_workbook(_excel_path, read_only=True, data_only=True)
            _ws = _wb['Test Lead POC']
            for _i, _row in enumerate(_ws.iter_rows(values_only=True)):
                if _i == 0:
                    continue  # skip header
                pod_name = str(_row[0]).strip() if _row[0] else ''
                lead = str(_row[3]).strip() if len(_row) > 3 and _row[3] else ''
                if pod_name and pod_name != 'None' and lead and lead != 'None':
                    POD_LEAD_MAP[pod_name.lower()] = lead
            _wb.close()
            print(f"   ✓ Loaded {len(POD_LEAD_MAP)} POD → Lead mappings from Excel")
        except Exception as _e:
            print(f"   ⚠️  Could not load POD_LEAD_MAP from Excel: {_e}")
        organized_data = self.organize_data_by_lead_module()
        grand_totals = self.calculate_grand_totals(organized_data)
        manual_gt = grand_totals['manual']
        auto_gt = grand_totals['automation']
        gt_pass_pct, gt_exec_pct = self.calculate_grand_total_percentages(manual_gt)
        auto_gt_pass_pct, auto_gt_exec_pct = self.calculate_grand_total_percentages(auto_gt)
        # Calculate leads-wise manual summary
        leads_summary = self.calculate_leads_summary(organized_data)
        # Calculate leads-wise automation summary
        automation_leads_summary = self.calculate_automation_leads_summary(organized_data)
        # Calculate insprint data for summary
        insprint_organized = self.organize_data_by_lead_module_insprint()
        insprint_grand_totals = self.calculate_grand_totals(insprint_organized)
        insprint_manual_gt = insprint_grand_totals['manual']
        insprint_gt_pass_pct, insprint_gt_exec_pct = self.calculate_grand_total_percentages(insprint_manual_gt)
        # Prod Sanity Scenarios data (Tab 1)
        ps_organized = self.organize_data_by_lead_module_prod_sanity()
        ps_grand_totals = self.calculate_grand_totals(ps_organized)
        ps_manual_gt = ps_grand_totals['manual']
        ps_auto_gt = ps_grand_totals['automation']
        ps_gt_pass_pct, ps_gt_exec_pct = self.calculate_grand_total_percentages(ps_manual_gt)
        ps_auto_gt_pass_pct, ps_auto_gt_exec_pct = self.calculate_grand_total_percentages(ps_auto_gt)
        ps_leads_summary = self.calculate_leads_summary(ps_organized)
        ps_automation_leads_summary = self.calculate_automation_leads_summary(ps_organized)
        plan_name = self.plan_info.get('name', f"Test Plan {ADO_CONFIG['plan_id']}")
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Test Execution Report - {self.suite_name}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            background: #f0f2f5;
            padding: 10px;
        }}
        .container {{
            max-width: 100%;
            margin: 0 auto;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 12px 15px;
            text-align: center;
        }}
        .header h1 {{ font-size: 18px; margin-bottom: 4px; }}
        .header p {{ font-size: 11px; opacity: 0.9; margin: 2px 0; }}
        
        .tabs {{
            display: flex;
            background: #e9ecef;
            border-bottom: 2px solid #475569;
        }}
        
        .tab {{
            padding: 10px 20px;
            font-size: 12px;
            font-weight: 600;
            cursor: pointer;
            border: none;
            background: transparent;
            color: #495057;
            transition: all 0.3s;
            border-bottom: 3px solid transparent;
        }}
        
        .tab:hover {{
            background: #dee2e6;
        }}
        
        .tab.active {{
            background: white;
            color: #667eea;
            border-bottom: 3px solid #667eea;
        }}
        
        .tab-content {{
            display: none;
        }}
        
        .tab-content.active {{
            display: block;
        }}
        
        .filter-section {{
            padding: 10px 15px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
            display: flex;
            gap: 15px;
            align-items: center;
            flex-wrap: wrap;
            position: relative;
            overflow: visible;
        }}
        
        .filter-group {{
            display: flex;
            align-items: center;
            gap: 8px;
            position: relative;
        }}
        
        .filter-group label {{
            font-size: 11px;
            font-weight: 600;
            color: #495057;
        }}
        
        .filter-group select {{
            padding: 5px 10px;
            font-size: 11px;
            border: 1px solid #dee2e6;
            border-radius: 4px;
            background: white;
            cursor: pointer;
            min-width: 150px;
        }}
        
        .filter-group select:focus {{
            outline: none;
            border-color: #667eea;
            box-shadow: 0 0 0 2px rgba(102, 126, 234, 0.1);
        }}
        
        .reset-btn {{
            padding: 5px 15px;
            font-size: 11px;
            background: #667eea;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-weight: 600;
        }}
        
        .reset-btn:hover {{
            background: #5568d3;
        }}
        
        .save-btn {{
            padding: 5px 15px;
            font-size: 11px;
            background: #28a745;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-weight: 600;
        }}
        
        .save-btn:hover {{
            background: #218838;
        }}
        
        .save-btn.saved {{
            background: #6c757d;
        }}
        
        .filter-info {{
            margin-left: auto;
            font-size: 11px;
            color: #6c757d;
            font-weight: 600;
        }}
        
        /* Custom Dropdown Styles */
        .custom-dropdown {{
            position: relative;
            display: inline-block;
            z-index: 100;
        }}
        
        .dropdown-toggle {{
            padding: 5px 10px;
            font-size: 11px;
            border: 1px solid #dee2e6;
            border-radius: 4px;
            background: white;
            cursor: pointer;
            min-width: 150px;
            text-align: left;
            font-weight: 600;
        }}
        
        .dropdown-toggle:hover {{
            background: #f8f9fa;
        }}
        
        .dropdown-menu {{
            display: none;
            position: absolute;
            top: 100%;
            left: 0;
            min-width: 200px;
            max-height: 350px;
            overflow-y: auto;
            overflow-x: hidden;
            background: white;
            border: 1px solid #dee2e6;
            border-radius: 4px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.2);
            z-index: 9999;
            margin-top: 2px;
        }}
        
        .dropdown-menu.show {{
            display: block;
        }}
        
        .dropdown-item {{
            padding: 6px 12px;
        }}
        
        .dropdown-item label {{
            display: flex;
            align-items: center;
            cursor: pointer;
            font-size: 11px;
            margin: 0;
            font-weight: normal;
        }}
        
        .dropdown-item label:hover {{
            background: #f8f9fa;
        }}
        
        .dropdown-item input[type="checkbox"] {{
            margin-right: 8px;
            cursor: pointer;
        }}
        
        .dropdown-divider {{
            height: 1px;
            background: #e9ecef;
            margin: 4px 0;
        }}
        
        .report-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 0;
            font-size: 10px;
        }}
        
        .report-table th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 6px 4px;
            text-align: center;
            font-size: 9px;
            font-weight: 600;
            border: 1px solid #dee2e6 !important;
            line-height: 1.2;
        }}
        
        .report-table th.main-header {{
            background: linear-gradient(135deg, #4299e1 0%, #667eea 100%);
            font-size: 10px;
            padding: 5px 3px;
        }}
        
        .report-table td {{
            padding: 4px 3px;
            text-align: center;
            border: 1px solid #dee2e6 !important;
            font-size: 10px;
            line-height: 1.3;
        }}
        
        .report-table tbody tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        
        .report-table tbody tr:hover {{
            background: #e9ecef;
        }}
        
        .report-table tbody tr.hidden {{
            display: none;
        }}
        .report-table tbody tr.bug-extra-row td {{
            background-color: inherit !important;
        }}
        
        .sno-col {{ width: 30px; font-weight: 600; }}
        .lead-col {{ width: 70px; font-weight: 600; font-size: 9px; }}
        .module-col {{ width: 100px; font-size: 9px; }}
        .title-col {{ max-width: 320px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; text-align: left; padding-left: 6px; cursor: default; }}
        .total-p1p2-col {{ width: 50px; font-weight: 700; background: #fd7e14 !important; color: white !important; }}
        .total-col {{ width: 40px; font-weight: 600; background: #e7f3ff !important; }}
        .pass-col {{ width: 35px; background: #d4edda !important; }}
        .fail-col {{ width: 35px; background: #f8d7da !important; }}
        .blocked-col {{ width: 35px; background: #fff3cd !important; }}
        .na-col {{ width: 35px; background: #e2e3e5 !important; }}
        .notrun-col {{ width: 40px; background: #cce5ff !important; }}
        .percentage-col {{ width: 45px; font-weight: 600; background: #fff9e6 !important; font-size: 9px; }}
        
        .grand-total-row {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            color: white !important;
            font-weight: 700 !important;
            font-size: 10px !important;
        }}

        /* ========================================================================
           LEADS WISE EXECUTION STATUS TABLE STYLES (TAB 3 & TAB 4)
           Compact Professional Design
           ======================================================================== */
        
        .leads-table {{
            width: auto;
            max-width: 700px;
            margin: 10px auto;
            border-collapse: collapse;
            font-size: 10px;
            background: white;
            box-shadow: 0 1px 3px rgba(0,0,0,0.08);
        }}
        
        /* Header Row - Compact Purple Gradient */
        .leads-table thead {{
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        
        .leads-table th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
            color: white !important;
            padding: 5px 6px;
            text-align: center;
            font-size: 9px;
            font-weight: 700;
            border: 1px solid #dee2e6 !important;
            line-height: 1.1;
            text-transform: uppercase;
            letter-spacing: 0.3px;
            white-space: nowrap;
        }}
        
        /* Data Cell Base Styles - Compact */
        .leads-table td {{
            padding: 4px 6px;
            text-align: center;
            border: 1px solid #dee2e6 !important;
            font-size: 10px;
            font-weight: 400;
            line-height: 1.2;
            color: #333 !important;
        }}
        
        /* Alternating Row Colors - Like Tab 1 */
        .leads-table tbody tr:nth-child(odd) td {{
            background: white !important;
        }}
        
        .leads-table tbody tr:nth-child(even) td {{
            background: #f8f9fa !important;
        }}
        
        /* Hover Effect - Matching Tab 1 */
        .leads-table tbody tr:not(.grand-total-row):hover td {{
            background: #e9ecef !important;
            cursor: pointer;
        }}
        
        /* Lead Name Column - Compact */
        .leads-table .lead-name-col {{
            width: 90px;
            font-weight: 700;
            text-align: left;
            padding-left: 8px;
            font-size: 10px;
            color: #1a1a1a !important;
            background: #e3f2fd !important;
        }}
        
        /* Pass Column - Compact */
        .leads-table .pass-col {{
            width: 40px;
            background: #d4edda !important;
            color: #155724 !important;
            font-weight: 600;
        }}
        
        /* Fail Column - Compact */
        .leads-table .fail-col {{
            width: 40px;
            background: #f8d7da !important;
            color: #721c24 !important;
            font-weight: 600;
        }}
        
        /* Blocked Column - Compact */
        .leads-table .blocked-col {{
            width: 40px;
            background: #fff3cd !important;
            color: #856404 !important;
            font-weight: 600;
        }}
        
        /* NA Column - Compact */
        .leads-table .na-col {{
            width: 40px;
            background: #e2e3e5 !important;
            color: #383d41 !important;
            font-weight: 600;
        }}
        
        /* Not Run Column - Compact */
        .leads-table .notrun-col {{
            width: 45px;
            background: #cce5ff !important;
            color: #004085 !important;
            font-weight: 600;
        }}
        
        /* Total Column - Compact */
        .leads-table .total-col {{
            width: 45px;
            font-weight: 700;
            background: #e7f3ff !important;
            color: #1a1a1a !important;
        }}
        
        /* Percentage Columns - Compact */
        .leads-table .percentage-col {{
            width: 50px;
            font-weight: 700;
            background: #fff9e6 !important;
            color: #1a1a1a !important;
            font-size: 10px;
        }}
        
        /* Grand Total Row - Compact Green Gradient */
        .leads-table .grand-total-row {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            font-weight: 700 !important;
            font-size: 10px !important;
        }}
        
        .leads-table .grand-total-row td {{
            border: 1px solid #dee2e6 !important;
            padding: 5px 6px;
            color: white !important;
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            font-weight: 700 !important;
            text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
        }}
        
        .leads-table .grand-total-row .lead-name-col {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            color: white !important;
            font-weight: 700 !important;
            text-align: left;
            padding-left: 8px;
            text-transform: uppercase;
        }}
        
        /* Override all status column colors for grand total */
        .leads-table .grand-total-row .pass-col,
        .leads-table .grand-total-row .fail-col,
        .leads-table .grand-total-row .blocked-col,
        .leads-table .grand-total-row .na-col,
        .leads-table .grand-total-row .notrun-col,
        .leads-table .grand-total-row .total-col,
        .leads-table .grand-total-row .percentage-col {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            color: white !important;
            font-weight: 700 !important;
        }}
        
        /* Ensure grand total ignores alternating colors */
        .leads-table tbody tr.grand-total-row:nth-child(even) td,
        .leads-table tbody tr.grand-total-row:nth-child(odd) td {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
        }}
        
        /* Grand total hover effect */
        .leads-table tbody tr.grand-total-row:hover td {{
            background: linear-gradient(90deg, #28a745 0%, #20c997 100%) !important;
            cursor: default;
        }}
        
        /* ========================================================================
           SUMMARY TABLE STYLES (TAB 1)
           Compact Professional Design
           ======================================================================== */
        
        .summary-table {{
            width: 100%;
            max-width: 900px;
            margin: 15px auto;
            border-collapse: collapse;
            font-size: 10px;
            box-shadow: 0 3px 10px rgba(0,0,0,0.12);
            border-radius: 6px;
            overflow: hidden;
        }}
        
        /* Summary Table Header */
        .summary-table thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        }}
        
        .summary-table th {{
            padding: 8px 10px;
            text-align: center;
            font-size: 9px;
            font-weight: 700;
            border: 1px solid #dee2e6 !important;
            color: white !important;
            text-transform: uppercase;
            letter-spacing: 0.6px;
            line-height: 1.2;
            text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
        }}
        
        /* Summary Table Data Cells */
        .summary-table td {{
            padding: 10px 8px;
            text-align: center;
            font-size: 11px;
            font-weight: 600;
            border: 1px solid #dee2e6 !important;
            color: white !important;
            text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
            transition: all 0.3s ease;
        }}
        
        .summary-table tbody tr:hover td {{
            transform: scale(1.01);
            box-shadow: 0 3px 6px rgba(0,0,0,0.15);
            cursor: pointer;
        }}
        
        /* Pass Column - Enhanced Green */
        .summary-table .summary-pass {{
            background: linear-gradient(135deg, #10b981 0%, #34d399 100%) !important;
            box-shadow: inset 0 2px 4px rgba(16, 185, 129, 0.3);
        }}
        
        .summary-table .summary-pass:hover {{
            background: linear-gradient(135deg, #059669 0%, #10b981 100%) !important;
        }}
        
        /* Fail Column - Enhanced Red */
        .summary-table .summary-fail {{
            background: linear-gradient(135deg, #ef4444 0%, #f87171 100%) !important;
            box-shadow: inset 0 2px 4px rgba(239, 68, 68, 0.3);
        }}
        
        .summary-table .summary-fail:hover {{
            background: linear-gradient(135deg, #dc2626 0%, #ef4444 100%) !important;
        }}
        
        /* Blocked Column - Enhanced Yellow/Orange */
        .summary-table .summary-blocked {{
            background: linear-gradient(135deg, #f59e0b 0%, #fbbf24 100%) !important;
            box-shadow: inset 0 2px 4px rgba(245, 158, 11, 0.3);
        }}
        
        .summary-table .summary-blocked:hover {{
            background: linear-gradient(135deg, #d97706 0%, #f59e0b 100%) !important;
        }}
        
        /* NA Column - Enhanced Gray */
        .summary-table .summary-na {{
            background: linear-gradient(135deg, #6b7280 0%, #9ca3af 100%) !important;
            box-shadow: inset 0 2px 4px rgba(107, 114, 128, 0.3);
        }}
        
        .summary-table .summary-na:hover {{
            background: linear-gradient(135deg, #4b5563 0%, #6b7280 100%) !important;
        }}
        
        /* Not Run Column - Enhanced Blue */
        .summary-table .summary-notrun {{
            background: linear-gradient(135deg, #3b82f6 0%, #60a5fa 100%) !important;
            box-shadow: inset 0 2px 4px rgba(59, 130, 246, 0.3);
        }}
        
        .summary-table .summary-notrun:hover {{
            background: linear-gradient(135deg, #2563eb 0%, #3b82f6 100%) !important;
        }}
        
        /* Total Column - Enhanced Cyan */
        .summary-table .summary-total {{
            background: linear-gradient(135deg, #06b6d4 0%, #22d3ee 100%) !important;
            box-shadow: inset 0 2px 4px rgba(6, 182, 212, 0.3);
            font-size: 12px;
            font-weight: 700;
        }}
        
        .summary-table .summary-total:hover {{
            background: linear-gradient(135deg, #0891b2 0%, #06b6d4 100%) !important;
        }}
        
        /* Pass% Column - Enhanced Lime Green */
        .summary-table .summary-pass-pct {{
            background: linear-gradient(135deg, #84cc16 0%, #a3e635 100%) !important;
            box-shadow: inset 0 2px 4px rgba(132, 204, 22, 0.3);
            font-size: 12px;
            font-weight: 700;
        }}
        
        .summary-table .summary-pass-pct:hover {{
            background: linear-gradient(135deg, #65a30d 0%, #84cc16 100%) !important;
        }}
        
        /* Exec% Column - Enhanced Emerald Green */
        .summary-table .summary-exec-pct {{
            background: linear-gradient(135deg, #10b981 0%, #34d399 100%) !important;
            box-shadow: inset 0 2px 4px rgba(16, 185, 129, 0.3);
            font-size: 12px;
            font-weight: 700;
        }}
        
        .summary-table .summary-exec-pct:hover {{
            background: linear-gradient(135deg, #059669 0%, #10b981 100%) !important;
        }}
        
        /* Responsive Design for Summary Table */
        @media (max-width: 1400px) {{
            .summary-table {{
                font-size: 9px;
                margin: 12px auto;
                max-width: 800px;
            }}
            
            .summary-table th {{
                font-size: 8px;
                padding: 7px 8px;
            }}
            
            .summary-table td {{
                font-size: 10px;
                padding: 8px 6px;
            }}
            
            .summary-table .summary-total,
            .summary-table .summary-pass-pct,
            .summary-table .summary-exec-pct {{
                font-size: 11px;
            }}
        }}
        
        @media (max-width: 1024px) {{
            .summary-table {{
                font-size: 8px;
                margin: 10px auto;
                max-width: 700px;
            }}
            
            .summary-table th {{
                font-size: 7px;
                padding: 6px 7px;
            }}
            
            .summary-table td {{
                font-size: 9px;
                padding: 7px 5px;
            }}
            
            .summary-table .summary-total,
            .summary-table .summary-pass-pct,
            .summary-table .summary-exec-pct {{
                font-size: 10px;
            }}
        }}
        
        @media (max-width: 768px) {{
            .summary-table {{
                font-size: 7px;
                margin: 8px;
                max-width: 100%;
            }}
            
            .summary-table th {{
                font-size: 6px;
                padding: 5px 6px;
            }}
            
            .summary-table td {{
                font-size: 8px;
                padding: 6px 4px;
            }}
            
            .summary-table .summary-total,
            .summary-table .summary-pass-pct,
            .summary-table .summary-exec-pct {{
                font-size: 9px;
            }}
        }}
        
        /* Print Styles for Summary Table */
        @media print {{
            .summary-table {{
                font-size: 9px;
                box-shadow: none;
                border-radius: 0;
            }}
            
            .summary-table th {{
                font-size: 8px;
                padding: 6px 8px;
                background: #667eea !important;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            .summary-table td {{
                padding: 7px 8px;
                font-size: 10px;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            .summary-table tbody tr:hover td {{
                transform: none;
                box-shadow: none;
            }}
        }}
        
        /* Bug Summary Table Styles - Compact Layout */
        .bug-summary-table {{
            width: auto;
            max-width: 750px;
            margin: 10px auto;
            border-collapse: collapse;
            box-shadow: 0 1px 4px rgba(0,0,0,0.1);
            background: white;
            border: 1px solid #dee2e6;
            font-size: 10px;
        }}
        
        .bug-summary-table thead {{
            background: linear-gradient(135deg, #dc2626 0%, #ef4444 100%);
            color: white;
        }}
        
        .bug-summary-table th {{
            padding: 6px 10px;
            text-align: center;
            font-size: 10px;
            font-weight: 700;
            border-right: 1px solid #dee2e6;
            border-bottom: 1px solid #dee2e6;
            line-height: 1.2;
            white-space: nowrap;
        }}
        
        .bug-summary-table th:last-child {{
            border-right: none;
        }}
        
        .bug-summary-table td {{
            padding: 5px 8px;
            text-align: center;
            font-size: 10px;
            border-bottom: 1px solid #dee2e6;
            border-right: 1px solid #dee2e6;
            line-height: 1.3;
        }}
        
        .bug-summary-table td:last-child {{
            border-right: none;
        }}
        
        .bug-summary-table tbody tr:hover {{
            background-color: #fef2f2;
        }}
        
        .bug-summary-table tbody tr:last-child td {{
            border-bottom: none;
        }}
        
        .bug-summary-table .bug-critical {{
            background: #fee2e2;
            color: #991b1b;
            font-weight: 600;
        }}
        
        .bug-summary-table .bug-high {{
            background: #fed7aa;
            color: #9a3412;
            font-weight: 600;
        }}
        
        .bug-summary-table .bug-medium {{
            background: #fef3c7;
            color: #92400e;
            font-weight: 600;
        }}
        
        .bug-summary-table .bug-low {{
            background: #dbeafe;
            color: #1e40af;
            font-weight: 600;
        }}
        
        .bug-summary-table .bug-total {{
            background: #f3f4f6;
            font-weight: 700;
            color: #1f2937;
        }}
        
        .bug-summary-table .grand-total-row {{
            background: #dc2626;
            color: white;
            font-weight: 700;
        }}
        
        .bug-summary-table .grand-total-row td {{
            border-bottom: none;
        }}
        
        /* Bug List Table Styles - Compact */
        .bug-list-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 10px 0;
            font-size: 10px;
            background: white;
            box-shadow: 0 1px 3px rgba(0,0,0,0.08);
        }}
        
        .bug-list-table thead {{
            background: linear-gradient(135deg, #dc2626 0%, #ef4444 100%);
            color: white;
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        
        .bug-list-table th {{
            padding: 6px 8px;
            text-align: center;
            font-size: 9px;
            font-weight: 700;
            border: 1px solid #dee2e6 !important;
            line-height: 1.2;
            white-space: nowrap;
        }}
        
        .bug-list-table td {{
            padding: 5px 8px;
            text-align: left;
            font-size: 10px;
            border: 1px solid #dee2e6 !important;
            line-height: 1.3;
        }}
        
        .bug-list-table tbody tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        
        .bug-list-table tbody tr:hover {{
            background: #fef2f2;
        }}
        
        .bug-list-table tbody tr.hidden {{
            display: none;
        }}
        
        .bug-list-table .bug-mpoc-col {{
            width: 110px;
            font-weight: 600;
            text-align: left;
        }}
        
        .bug-list-table .bug-id-col {{
            width: 70px;
            text-align: center;
            font-weight: 600;
        }}
        
        .bug-list-table .bug-title-col {{
            min-width: 200px;
            max-width: 350px;
        }}
        
        .bug-list-table .bug-state-col {{
            width: 100px;
            text-align: center;
            font-weight: 600;
        }}
        
        .bug-list-table .bug-defect-col {{
            width: 110px;
            text-align: center;
        }}
        
        .bug-list-table .bug-severity-col {{
            width: 90px;
            text-align: center;
            font-weight: 600;
        }}
        
        .bug-list-table .bug-node-col {{
            width: 130px;
        }}
        
        .bug-list-table .bug-stage-col {{
            width: 110px;
            text-align: center;
        }}

        .export-btn {{
            padding: 5px 12px;
            font-size: 11px;
            background: #1d6f42;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-weight: 600;
        }}
        .export-btn:hover {{ background: #166337; }}

        .export-all-btn {{
            padding: 5px 12px;
            font-size: 11px;
            background: #2563eb;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-weight: 600;
        }}
        .export-all-btn:hover {{ background: #1d4ed8; }}

        .refresh-btn {{
            padding: 5px 12px;
            font-size: 11px;
            background: #0891b2;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-weight: 600;
        }}
        .refresh-btn:hover {{ background: #0e7490; }}
        
    </style>
    <script src="https://cdn.jsdelivr.net/npm/exceljs@4.3.0/dist/exceljs.min.js"></script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 Prod Execution Report</h1>
            <p><strong>{plan_name}</strong> | <strong>Suite: {self.suite_name}</strong> | Generated: {self.timestamp}</p>
        </div>
        
        <!-- Tabs -->
        <div class="tabs">
            <button class="tab active" onclick="switchTab('overallProdSanity')">&#127775; Overall Production Sanity Status</button>
            <button class="tab" onclick="switchTab('prodSanity')">&#128202; Prod Sanity Scenarios</button>
            <button class="tab" onclick="switchTab('prodUSSanity')">&#128203; Prod US Sanity</button>
            <button class="tab" onclick="switchTab('insprintStatus')">&#128230; Insprint US Prod Status</button>
            <button class="tab" onclick="switchTab('readyForProdBug')">&#128027; Ready for Prod Bug</button>
            <button class="tab" onclick="switchTab('prodSanityDefects')">&#128308; Prod Sanity Defects</button>
            <button class="tab" onclick="switchTab('overallStatus')">&#127775; Leadwise Prod Sanity Status</button>
        </div>

        <!-- Tab 0: Overall Production Sanity Status -->
<div id="overallProdSanityTab" class="tab-content active">
"""
        # ── Pivot helper ─────────────────────────────────────────────────────
        pivot_statuses = [
            'Working fine',
            'Working fine with active bug',
            'In Progress',
            'Failed',
            'Blocked',
            'Cannot be validated',
            'Can be validated on Monday',
            'Taken care by BPMS, IMS test team',
            'Story not deployed to prod as part of this release',
            'Toggle off/Code reversal validated',
            'Yet to start',
        ]

        def _build_pivot(records, title_key='name', lead_key='lead', status_key='prod_sanity_status'):
            """Return (leads_sorted, pivot_dict) where pivot_dict[lead][status] = count."""
            from collections import defaultdict as _dd
            pivot = _dd(lambda: _dd(int))
            for r in records:
                lead   = r.get(lead_key) or 'Unassigned'
                status = r.get(status_key) or ''
                if not status:
                    status = 'Yet to start'
                pivot[lead][status] += 1
            leads_sorted = sorted(pivot.keys())
            return leads_sorted, {l: dict(v) for l, v in pivot.items()}

        def _render_pivot_table(html_acc, table_title, leads, pivot, statuses, container_id=None):
            """Append a styled pivot HTML table to html_acc and return it."""
            # Determine which status columns actually have data
            used_statuses = [s for s in statuses if any(pivot.get(l, {}).get(s, 0) for l in leads)]
            if not used_statuses:
                used_statuses = statuses

            # Grand total row
            grand = {s: sum(pivot.get(l, {}).get(s, 0) for l in leads) for s in used_statuses}
            grand_total = sum(grand.values())
            _div_id_attr = f' id="{container_id}"' if container_id else ''

            html_acc += f"""
            <div{_div_id_attr} style="margin:20px 0;">
                <h3 style="font-size:13px;font-weight:700;color:#4a4a8a;margin-bottom:8px;padding-left:4px;border-left:4px solid #667eea;">{table_title}</h3>
                <div class="table-wrapper" style="overflow-x:auto;">
                    <table style="border-collapse:collapse;font-size:10px;width:auto;min-width:400px;">
                        <thead>
                            <tr>
                                <th style="background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);color:white;padding:6px 10px;border:1px solid #dee2e6;text-align:left;white-space:nowrap;">Test Lead</th>
"""
            for s in used_statuses:
                html_acc += f'                                <th style="background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);color:white;padding:6px 8px;border:1px solid #dee2e6;text-align:center;white-space:nowrap;font-size:9px;">{s}</th>\n'
            html_acc += '                                <th style="background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);color:white;padding:6px 10px;border:1px solid #dee2e6;text-align:center;white-space:nowrap;">Total</th>\n'
            html_acc += '                            </tr>\n                        </thead>\n                        <tbody>\n'

            for idx, lead in enumerate(leads):
                row_total = sum(pivot.get(lead, {}).get(s, 0) for s in used_statuses)
                bg = '#f8f9fa' if idx % 2 == 0 else 'white'
                html_acc += f'                            <tr style="background:{bg};">\n'
                html_acc += f'                                <td style="padding:5px 10px;border:1px solid #dee2e6;font-weight:600;text-align:left;white-space:nowrap;background:#e3f2fd;">{lead}</td>\n'
                for s in used_statuses:
                    cnt = pivot.get(lead, {}).get(s, 0)
                    cell_bg = '#f0fff0' if cnt > 0 else ''
                    style_extra = f'background:{cell_bg};' if cell_bg else ''
                    html_acc += f'                                <td style="padding:5px 8px;border:1px solid #dee2e6;text-align:center;{style_extra}">{cnt if cnt else "—"}</td>\n'
                html_acc += f'                                <td style="padding:5px 10px;border:1px solid #dee2e6;text-align:center;font-weight:700;background:#e7f3ff;">{row_total}</td>\n'
                html_acc += '                            </tr>\n'

            # Grand total row
            html_acc += '                            <tr style="background:linear-gradient(90deg,#28a745 0%,#20c997 100%);">\n'
            html_acc += '                                <td style="padding:5px 10px;border:1px solid #dee2e6;font-weight:700;color:white;text-align:left;">Grand Total</td>\n'
            for s in used_statuses:
                html_acc += f'                                <td style="padding:5px 8px;border:1px solid #dee2e6;font-weight:700;color:white;text-align:center;">{grand[s] if grand[s] else "—"}</td>\n'
            html_acc += f'                                <td style="padding:5px 10px;border:1px solid #dee2e6;font-weight:700;color:white;text-align:center;">{grand_total}</td>\n'
            html_acc += '                            </tr>\n'
            html_acc += '                        </tbody>\n                    </table>\n                </div>\n            </div>\n'
            return html_acc

        # ── Build pivots from the three data sources ──────────────────────────
        # Pivot 1: Prod Sanity Scenarios — one row per test scenario
        ps_leads, ps_pivot = _build_pivot(self.prod_sanity_data)

        # Helper: resolve the dedup key for a US story record.
        # Priority: t.get('us_id') first, then number extracted from module field, then name.
        # This MUST match the logic used in _resolve_us_key (used by the table rows) so
        # the pivot count is identical to the table row count.
        def _pivot_us_key(t):
            raw_tv = t.get('module') or ''
            tv_match = re.match(r'^(\d+)\s*[:]\s*(.+)$', raw_tv, re.DOTALL) \
                    or re.match(r'^(\d+)\s*[-\u2013\u2014]\s*(.+)$', raw_tv, re.DOTALL)
            tv_us_id = tv_match.group(1).strip() if tv_match else ''
            return str(t.get('us_id') or tv_us_id or t['name'])

        # Pivot 2: Prod US Sanity — one row per unique US
        # Uses us_id-first key to match the table's _resolve_us_key deduplication.
        from collections import OrderedDict as _OD2
        _us_seen = _OD2()
        for t in self.test_data:
            _key = _pivot_us_key(t)
            if _key not in _us_seen:
                _us_seen[_key] = {
                    'lead': t['lead'],
                    'prod_sanity_status': t.get('prod_sanity_status') or '',
                }
            elif not _us_seen[_key]['prod_sanity_status'] and t.get('prod_sanity_status'):
                _us_seen[_key]['prod_sanity_status'] = t['prod_sanity_status']
        us_leads, us_pivot = _build_pivot(list(_us_seen.values()))

        # Pivot 3: Insprint US Prod Status — one row per unique US in insprint suite
        # Uses module-first key to match ins_story_map deduplication.
        _ins_seen = _OD2()
        for t in self.insprint_data:
            raw_mod = t.get('module') or ''
            mod_m2 = re.match(r'^(\d+)\s*[:]\s*(.+)$', raw_mod, re.DOTALL) or \
                     re.match(r'^(\d+)\s*[-\u2013\u2014]\s*(.+)$', raw_mod, re.DOTALL)
            ins_key = mod_m2.group(1).strip() if mod_m2 else (str(t.get('us_id') or '') or t['name'])
            if ins_key not in _ins_seen:
                _ins_seen[ins_key] = {
                    'lead': t.get('lead', ''),
                    'prod_sanity_status': t.get('prod_sanity_status') or '',
                }
            elif not _ins_seen[ins_key]['prod_sanity_status'] and t.get('prod_sanity_status'):
                _ins_seen[ins_key]['prod_sanity_status'] = t['prod_sanity_status']
        ins_leads, ins_pivot = _build_pivot(list(_ins_seen.values()))

        overall_html = """
    <div class="filter-section">
        <button class="export-btn" onclick="exportCurrentTabToExcel()" title="Export current tab to Excel (.xlsx)">&#128229; Export Tab (.xlsx)</button>
        <button class="export-all-btn" onclick="exportAllTabsToExcel()" title="Export all 6 tabs to Excel workbook (.xlsx)">&#128194; Export All Tabs (.xlsx)</button>
    </div>
    <div id="overallStatusContent" style="padding:20px;">"""
        overall_html  = _render_pivot_table(overall_html, 'Prod Sanity Scenarios — Status by Test Lead',  ps_leads,  ps_pivot,  pivot_statuses, 'pivotProdSanity')
        overall_html  = _render_pivot_table(overall_html, 'Prod US Sanity — Status by Test Lead',          us_leads,  us_pivot,  pivot_statuses, 'pivotProdUSSanity')
        overall_html  = _render_pivot_table(overall_html, 'Insprint US Prod Status — Status by Test Lead', ins_leads, ins_pivot, pivot_statuses, 'pivotInsprintUS')

        # Pivot 4: Ready for Prod Bug — Bug State by Test Lead
        from collections import defaultdict as _dd4
        _bug_pivot = _dd4(lambda: _dd4(int))
        _bug_states_seen = set()
        for _bug in self.bug_data:
            _lead = POD_LEAD_MAP.get(_bug.get('node_name', ''), 'Unassigned') or 'Unassigned'
            _state = _bug.get('state') or 'Unknown'
            _bug_pivot[_lead][_state] += 1
            _bug_states_seen.add(_state)
        _bug_leads_sorted = sorted(_bug_pivot.keys())
        _bug_states_sorted = sorted(_bug_states_seen)
        _bug_pivot_dict = {l: dict(v) for l, v in _bug_pivot.items()}
        overall_html = _render_pivot_table(overall_html, 'Ready for Prod Bug — State by Test Lead', _bug_leads_sorted, _bug_pivot_dict, _bug_states_sorted, 'pivotReadyForProdBug')
        overall_html += '</div>\n'

        # ── Module-wise Executive Pivot — Overall Production Sanity Status tab ──
        from collections import defaultdict as _mdd
        _esc2 = lambda v: str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        _exec_statuses = ['Passed', 'Failed', 'Blocked', 'Not Applicable', 'Not Run']
        _snorm = {
            'passed': 'Passed', 'pass': 'Passed',
            'failed': 'Failed', 'fail': 'Failed',
            'blocked': 'Blocked', 'block': 'Blocked',
            'not applicable': 'Not Applicable', 'na': 'Not Applicable',
            'n/a': 'Not Applicable', 'notapplicable': 'Not Applicable',
            'not run': 'Not Run', 'notrun': 'Not Run',
            'active': 'Not Run', 'none': 'Not Run', '': 'Not Run',
        }
        def _no(raw): return _snorm.get((raw or '').strip().lower(), 'Not Run')
        def _agg_oc(ocs):
            fl = [_no(o) for o in ocs if o]
            non_na = [f for f in fl if f != 'Not Applicable']
            if not non_na: return 'Not Applicable' if fl else 'Not Run'
            if 'Failed' in non_na: return 'Failed'
            if 'Blocked' in non_na: return 'Blocked'
            if all(f == 'Passed' for f in non_na): return 'Passed'
            return 'Not Run'
        _cst = {
            'Passed':         ('#d1fae5', '#065f46', '#059669'),
            'Failed':         ('#fee2e2', '#7f1d1d', '#dc2626'),
            'Blocked':        ('#ffedd5', '#7c2d12', '#ea580c'),
            'Not Applicable': ('#f1f5f9', '#475569', '#64748b'),
            'Not Run':        ('#f0f9ff', '#0c4a6e', '#0284c7'),
        }

        def _exec_pivot_card(title, pivot_d, mods, cols, accent, icon):
            used = [c for c in cols if any(pivot_d.get(m, {}).get(c, 0) for m in mods)] or list(cols)
            grand = {c: sum(pivot_d.get(m, {}).get(c, 0) for m in mods) for c in used}
            gtotal = sum(grand.values())
            badges = ''
            for c in used:
                bg, fg, _ = _cst.get(c, ('#f1f5f9', '#475569', '#64748b'))
                badges += (f'<span style="background:{bg};color:{fg};padding:2px 9px;border-radius:10px;'
                           f'font-size:9px;font-weight:700;white-space:nowrap;">{_esc2(c)}: {grand[c]}</span> ')
            hdrs = ''
            for c in used:
                bg, _, hdr = _cst.get(c, ('#f1f5f9', '#475569', '#64748b'))
                hdrs += (f'<th style="padding:8px 10px;text-align:center;font-weight:700;color:{hdr};'
                         f'border-bottom:2px solid {hdr}40;white-space:nowrap;background:{bg};font-size:10px;">'
                         f'{_esc2(c)}</th>')
            rows = ''
            for i, mod in enumerate(mods):
                rt = sum(pivot_d.get(mod, {}).get(c, 0) for c in used)
                br = '#fafafa' if i % 2 == 0 else '#ffffff'
                cells = ''
                for c in used:
                    cnt = pivot_d.get(mod, {}).get(c, 0)
                    cb, cf, _ = _cst.get(c, ('#f1f5f9', '#475569', '#64748b'))
                    cs = f'background:{cb};color:{cf};font-weight:700;' if cnt > 0 else 'color:#cbd5e1;'
                    cells += (f'<td style="padding:7px 10px;text-align:center;'
                              f'border-bottom:1px solid #94a3b8;{cs}">'
                              f'{cnt if cnt > 0 else "&#8212;"}</td>')
                rows += (f'<tr style="background:{br};">'
                         f'<td style="padding:7px 14px;color:#1e293b;font-weight:500;'
                         f'border-bottom:1px solid #94a3b8;max-width:200px;overflow:hidden;'
                         f'text-overflow:ellipsis;white-space:nowrap;" title="{_esc2(mod)}">{_esc2(mod)}</td>'
                         f'{cells}'
                         f'<td style="padding:7px 10px;text-align:center;font-weight:800;color:#0f172a;'
                         f'border-bottom:1px solid #94a3b8;background:#f8fafc;">{rt}</td></tr>')
            gts = ''.join(
                f'<td style="padding:9px 10px;text-align:center;font-weight:800;'
                f'color:{_cst.get(c, (None,None,"#64748b"))[2]};">'
                f'{grand[c] if grand[c] else "&#8212;"}</td>'
                for c in used)
            return (
                f'<div style="background:white;border-radius:12px;box-shadow:0 4px 20px rgba(0,0,0,0.08);'
                f'margin-bottom:20px;overflow:hidden;border-top:4px solid {accent};">'
                f'<div style="background:linear-gradient(135deg,{accent}18 0%,transparent 100%);'
                f'padding:14px 20px;border-bottom:1px solid {accent}22;display:flex;align-items:flex-start;gap:10px;">'
                f'<div style="font-size:22px;line-height:1;">{icon}</div>'
                f'<div style="flex:1;">'
                f'<div style="font-size:12px;font-weight:800;color:{accent};letter-spacing:0.2px;">{_esc2(title)}</div>'
                f'<div style="font-size:10px;color:#64748b;margin-top:2px;">'
                f'Module-wise breakdown &middot; <strong style="color:{accent};">{gtotal}</strong> total items</div>'
                f'</div>'
                f'<div style="display:flex;gap:6px;flex-wrap:wrap;max-width:55%;justify-content:flex-end;">{badges}</div>'
                f'</div>'
                f'<div style="overflow-x:auto;">'
                f'<table style="width:100%;border-collapse:collapse;font-size:10.5px;">'
                f'<thead><tr style="background:#f8fafc;">'
                f'<th style="padding:8px 14px;text-align:left;font-weight:700;color:#334155;'
                f'border-bottom:2px solid #475569;white-space:nowrap;min-width:130px;font-size:10px;">Module</th>'
                f'{hdrs}'
                f'<th style="padding:8px 10px;text-align:center;font-weight:700;color:#334155;'
                f'border-bottom:2px solid #94a3b8;background:#f1f5f9;white-space:nowrap;font-size:10px;">Total</th>'
                f'</tr></thead>'
                f'<tbody>{rows}'
                f'<tr style="background:linear-gradient(90deg,{accent}18,{accent}08);">'
                f'<td style="padding:9px 14px;font-weight:800;color:{accent};font-size:11px;'
                f'border-top:2px solid {accent}30;">Grand Total</td>'
                f'{gts}'
                f'<td style="padding:9px 10px;text-align:center;font-weight:800;color:{accent};'
                f'font-size:11px;border-top:2px solid {accent}30;">{gtotal}</td>'
                f'</tr></tbody></table></div></div>'
            )

        # 1. Prod Sanity Scenarios — by text_verification × outcome
        _ps_mp = _mdd(lambda: _mdd(int))
        for _t in self.prod_sanity_data:
            _ps_mp[(_t.get('text_verification') or 'Unassigned').strip() or 'Unassigned'][_no(_t.get('outcome', ''))] += 1
        _ps_mods = sorted(_ps_mp.keys())
        _ps_mpd = {m: dict(v) for m, v in _ps_mp.items()}

        # 2. Prod US Sanity — deduplicated stories by text_verification × aggregated outcome
        _us_ms = _OD2()
        for _t in self.test_data:
            _k = _pivot_us_key(_t)
            _tv = (_t.get('text_verification') or 'Unassigned').strip() or 'Unassigned'
            if _k not in _us_ms:
                _us_ms[_k] = {'mod': _tv, 'ocs': [_t.get('outcome') or 'Not Run']}
            else:
                _us_ms[_k]['ocs'].append(_t.get('outcome') or 'Not Run')
        _us_mp = _mdd(lambda: _mdd(int))
        for _s in _us_ms.values():
            _us_mp[_s['mod']][_agg_oc(_s['ocs'])] += 1
        _us_mods = sorted(_us_mp.keys())
        _us_mpd = {m: dict(v) for m, v in _us_mp.items()}

        # 3. Insprint US Prod Status — deduplicated stories by text_verification × aggregated outcome
        _ins_ms2 = _OD2()
        for _t in self.insprint_data:
            _rm = _t.get('module') or ''
            _mx = re.match(r'^(\d+)\s*[:]\s*(.+)$', _rm, re.DOTALL) or \
                  re.match(r'^(\d+)\s*[-\u2013\u2014]\s*(.+)$', _rm, re.DOTALL)
            _ik = _mx.group(1).strip() if _mx else (str(_t.get('us_id') or '') or _t['name'])
            _tv2 = (_t.get('text_verification') or 'Unassigned').strip() or 'Unassigned'
            if _ik not in _ins_ms2:
                _ins_ms2[_ik] = {'mod': _tv2, 'ocs': [_t.get('outcome') or 'Not Run']}
            else:
                _ins_ms2[_ik]['ocs'].append(_t.get('outcome') or 'Not Run')
        _ins_mp = _mdd(lambda: _mdd(int))
        for _s in _ins_ms2.values():
            _ins_mp[_s['mod']][_agg_oc(_s['ocs'])] += 1
        _ins_mods = sorted(_ins_mp.keys())
        _ins_mpd = {m: dict(v) for m, v in _ins_mp.items()}

        # 4. Ready for Prod Bugs — State → Count summary
        from collections import defaultdict as _dd5
        _bug_state_counts = _dd5(int)
        for _bg in self.bug_data:
            _bug_state_counts[_bg.get('state') or 'Unknown'] += 1
        # Keep a stable display order: open-ish states first, then resolved/closed
        _bug_state_order = [
            'New', 'Active', 'Blocked', 'Blocked in PT', 'Blocked in UAT',
            'Re-Open', 'BA Clarification', 'Deferred',
            'In Progress', 'Ready to Deploy', 'Resolved',
            'Fixed and Ready to Deploy', 'Under Testing', 'Closed',
        ]
        _bug_states_ordered = (
            [s for s in _bug_state_order if s in _bug_state_counts] +
            sorted(s for s in _bug_state_counts if s not in _bug_state_order)
        )
        # Colour map per state (bg, text)
        _bug_state_colors = {
            'New':                      ('#fef3c7', '#78350f'),
            'Active':                   ('#fee2e2', '#7f1d1d'),
            'Blocked':                  ('#ffedd5', '#7c2d12'),
            'Blocked in PT':            ('#ffedd5', '#7c2d12'),
            'Blocked in UAT':           ('#ffedd5', '#7c2d12'),
            'Re-Open':                  ('#fce7f3', '#831843'),
            'BA Clarification':         ('#fdf4ff', '#6b21a8'),
            'Deferred':                 ('#f1f5f9', '#475569'),
            'In Progress':              ('#fef9c3', '#713f12'),
            'Ready to Deploy':          ('#dcfce7', '#14532d'),
            'Resolved':                 ('#d1fae5', '#065f46'),
            'Fixed and Ready to Deploy':('#d1fae5', '#065f46'),
            'Under Testing':            ('#dbeafe', '#1e3a8a'),
            'Closed':                   ('#eff6ff', '#1d4ed8'),
        }

        # KPI totals
        _ps_t = len(self.prod_sanity_data)
        _ps_p = sum(1 for t in self.prod_sanity_data if _no(t.get('outcome', '')) == 'Passed')
        _ps_f = sum(1 for t in self.prod_sanity_data if _no(t.get('outcome', '')) == 'Failed')
        _us_t = len(_us_ms)
        _us_p = sum(1 for s in _us_ms.values() if _agg_oc(s['ocs']) == 'Passed')
        _us_f = sum(1 for s in _us_ms.values() if _agg_oc(s['ocs']) == 'Failed')
        _ins_t = len(_ins_ms2)
        _ins_p = sum(1 for s in _ins_ms2.values() if _agg_oc(s['ocs']) == 'Passed')
        _ins_f = sum(1 for s in _ins_ms2.values() if _agg_oc(s['ocs']) == 'Failed')
        _bg_t  = len(self.bug_data)
        _bg_act = sum(1 for b in self.bug_data if (b.get('state') or '').lower() == 'active')
        def _pct2(n, d): return f'{round(100 * n / d)}%' if d else '&#8212;'

        def _build_bug_state_card(state_counts, states_ordered, state_colors, total):
            """Render a compact State → Count card for Ready for Prod Bugs."""
            accent = '#dc2626'
            rows_html = ''
            for st in states_ordered:
                cnt = state_counts.get(st, 0)
                bg, fg = state_colors.get(st, ('#f1f5f9', '#334155'))
                pct = round(100 * cnt / total) if total else 0
                bar_w = max(pct, 2) if cnt else 0
                rows_html += (
                    f'<tr>'
                    f'<td style="padding:7px 14px;font-weight:600;color:#1e293b;'
                    f'border-bottom:1px solid #94a3b8;white-space:nowrap;">'
                    f'<span style="display:inline-block;width:10px;height:10px;'
                    f'border-radius:50%;background:{bg};border:2px solid {fg};'
                    f'margin-right:7px;vertical-align:middle;"></span>{_esc2(st)}</td>'
                    f'<td style="padding:7px 18px;text-align:center;font-weight:800;'
                    f'font-size:13px;color:{fg};border-bottom:1px solid #94a3b8;">'
                    f'<span style="background:{bg};color:{fg};padding:2px 10px;'
                    f'border-radius:12px;font-size:11px;">{cnt}</span></td>'
                    f'<td style="padding:7px 14px;border-bottom:1px solid #94a3b8;min-width:120px;">'
                    f'<div style="background:#f1f5f9;border-radius:4px;height:8px;overflow:hidden;">'
                    f'<div style="width:{bar_w}%;background:{fg};height:100%;border-radius:4px;'
                    f'transition:width 0.4s;"></div></div>'
                    f'<span style="font-size:9px;color:#94a3b8;margin-top:2px;display:block;">{pct}%</span>'
                    f'</td></tr>'
                )
            return (
                f'<div style="background:white;border-radius:12px;'
                f'box-shadow:0 4px 20px rgba(0,0,0,0.08);margin-bottom:20px;'
                f'overflow:hidden;border-top:4px solid {accent};">'
                f'<div style="background:linear-gradient(135deg,{accent}18 0%,transparent 100%);'
                f'padding:14px 20px;border-bottom:1px solid {accent}22;'
                f'display:flex;align-items:center;gap:10px;">'
                f'<div style="font-size:22px;line-height:1;">&#128027;</div>'
                f'<div style="flex:1;">'
                f'<div style="font-size:12px;font-weight:800;color:{accent};letter-spacing:0.2px;">'
                f'Ready for Prod Bugs \u2014 State Summary</div>'
                f'<div style="font-size:10px;color:#64748b;margin-top:2px;">'
                f'Bug count by state &middot; <strong style="color:{accent};">{total}</strong> total bugs</div>'
                f'</div>'
                f'<span style="background:{accent};color:white;padding:4px 14px;'
                f'border-radius:20px;font-size:13px;font-weight:800;">{total} total</span>'
                f'</div>'
                f'<div style="padding:4px 0 8px;">'
                f'<table style="width:100%;border-collapse:collapse;font-size:11px;">'
                f'<thead><tr style="background:#f8fafc;">'
                f'<th style="padding:8px 14px;text-align:left;font-weight:700;color:#334155;'
                f'border-bottom:2px solid #475569;font-size:10px;">State</th>'
                f'<th style="padding:8px 18px;text-align:center;font-weight:700;color:#334155;'
                f'border-bottom:2px solid #475569;font-size:10px;white-space:nowrap;">Count</th>'
                f'<th style="padding:8px 14px;font-weight:700;color:#334155;'
                f'border-bottom:2px solid #475569;font-size:10px;min-width:150px;">Share</th>'
                f'</tr></thead>'
                f'<tbody>{rows_html}</tbody>'
                f'<tfoot><tr style="background:linear-gradient(90deg,{accent}18,{accent}08);">'
                f'<td style="padding:9px 14px;font-weight:800;color:{accent};font-size:11px;'
                f'border-top:2px solid {accent}30;">Grand Total</td>'
                f'<td style="padding:9px 18px;text-align:center;font-weight:800;'
                f'color:{accent};font-size:13px;border-top:2px solid {accent}30;">{total}</td>'
                f'<td style="border-top:2px solid {accent}30;"></td>'
                f'</tr></tfoot>'
                f'</table></div></div>'
            )

        def _kpi_card(label, total, s1l, s1v, s1bg, s1fg, s2l, s2v, s2bg, s2fg, accent):
            return (
                f'<div style="background:white;border-radius:10px;padding:16px 18px;'
                f'box-shadow:0 2px 14px rgba(0,0,0,0.07);border-left:4px solid {accent};'
                f'position:relative;overflow:hidden;">'
                f'<div style="position:absolute;right:-8px;top:-8px;width:55px;height:55px;'
                f'background:{accent};opacity:0.07;border-radius:50%;"></div>'
                f'<div style="font-size:9px;color:{accent};font-weight:800;text-transform:uppercase;'
                f'letter-spacing:1.2px;margin-bottom:4px;">{label}</div>'
                f'<div style="font-size:28px;font-weight:900;color:#0f172a;margin-bottom:8px;line-height:1;">{total}</div>'
                f'<div style="display:flex;gap:6px;flex-wrap:wrap;">'
                f'<span style="background:{s1bg};color:{s1fg};padding:2px 8px;border-radius:10px;'
                f'font-size:9px;font-weight:700;">{s1l}: {s1v}</span>'
                f'<span style="background:{s2bg};color:{s2fg};padding:2px 8px;border-radius:10px;'
                f'font-size:9px;font-weight:700;">{s2l}: {s2v}</span>'
                f'</div></div>'
            )

        kpi_html = (
            _kpi_card('Prod Sanity Scenarios', _ps_t,
                      '&#10003; Passed', f'{_ps_p} ({_pct2(_ps_p, _ps_t)})', '#d1fae5', '#065f46',
                      '&#10007; Failed', str(_ps_f), '#fee2e2', '#7f1d1d', '#059669') +
            _kpi_card('Prod US Sanity Stories', _us_t,
                      '&#10003; Passed', f'{_us_p} ({_pct2(_us_p, _us_t)})', '#dbeafe', '#1e3a8a',
                      '&#10007; Failed', str(_us_f), '#fee2e2', '#7f1d1d', '#2563eb') +
            _kpi_card('Insprint US Prod Stories', _ins_t,
                      '&#10003; Passed', f'{_ins_p} ({_pct2(_ins_p, _ins_t)})', '#ede9fe', '#3b0764',
                      '&#10007; Failed', str(_ins_f), '#fee2e2', '#7f1d1d', '#7c3aed') +
            _kpi_card('Ready for Prod Bugs', _bg_t,
                      '&#9889; Active', str(_bg_act), '#fef3c7', '#78350f',
                      '&#8226; Resolved/Closed', str(_bg_t - _bg_act), '#f1f5f9', '#475569', '#dc2626')
        )

        new_overall_html = (
            '<div style="background:#f0f4f8;min-height:100%;padding:0;">'
            + '<div style="background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 45%,#1d4ed8 100%);'
            + 'padding:22px 28px 20px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;">'
            + '<div>'
            + '<div style="font-size:9px;color:#93c5fd;letter-spacing:2.5px;text-transform:uppercase;font-weight:700;margin-bottom:5px;">Executive Dashboard</div>'
            + '<h2 style="color:white;font-size:20px;font-weight:900;margin:0;letter-spacing:0.2px;">&#127775; Overall Production Sanity Status</h2>'
            + f'<div style="color:#bfdbfe;font-size:10px;margin-top:4px;">Module-wise execution analysis &middot; Generated: {self.timestamp}</div>'
            + '</div>'
            + '<div style="display:flex;gap:10px;flex-wrap:wrap;">'
            + '<button onclick="exportCurrentTabToExcel()" style="background:rgba(255,255,255,0.14);border:1px solid rgba(255,255,255,0.28);color:white;padding:7px 14px;border-radius:7px;font-size:10px;cursor:pointer;font-weight:600;letter-spacing:0.3px;">&#128229; Export Tab</button>'
            + '<button onclick="exportAllTabsToExcel()" style="background:rgba(255,255,255,0.14);border:1px solid rgba(255,255,255,0.28);color:white;padding:7px 14px;border-radius:7px;font-size:10px;cursor:pointer;font-weight:600;letter-spacing:0.3px;">&#128194; Export All</button>'
            + '</div></div>'
            + f'<div style="padding:18px 24px 8px;display:grid;grid-template-columns:repeat(4,1fr);gap:14px;">{kpi_html}</div>'
            + '<div style="padding:8px 24px 28px;">'
            + _exec_pivot_card('Prod Sanity Scenarios \u2014 Module \u00d7 Execution Status',
                               _ps_mpd, _ps_mods, _exec_statuses, '#059669', '&#128202;')
            + _exec_pivot_card('Prod US Sanity \u2014 Module \u00d7 Execution Status',
                               _us_mpd, _us_mods, _exec_statuses, '#2563eb', '&#128203;')
            + _exec_pivot_card('Insprint US Prod Status \u2014 Module \u00d7 Execution Status',
                               _ins_mpd, _ins_mods, _exec_statuses, '#7c3aed', '&#128230;')
            + _build_bug_state_card(_bug_state_counts, _bug_states_ordered, _bug_state_colors, _bg_t)
            + '</div></div>'
        )

        html += new_overall_html
        html += '</div>\n\n        <!-- Tab 1: Prod Sanity Scenarios -->\n<div id="prodSanityTab" class="tab-content">'

        # Build unique filter values from prod_sanity_data
        ps_unique_leads = sorted(set(t['lead'] for t in self.prod_sanity_data))
        ps_unique_modules = sorted(set(t.get('text_verification') or '' for t in self.prod_sanity_data if t.get('text_verification')))
        ps_fixed_statuses = [
            'Working fine',
            'Working fine with active bug',
            'In Progress',
            'Failed',
            'Blocked',
            'Cannot be validated',
            'Can be validated on Monday',
            'Taken care by BPMS, IMS test team',
            'Story not deployed to prod as part of this release',
            'Toggle off/Code reversal validated',
            'Yet to start',
        ]

        # Filter bar
        html += """
    <div class="filter-section">
        <div class="filter-group">
            <label for="psLeadFilter">🔍 Filter by Lead:</label>
            <select id="psLeadFilter" onchange="applyPSFilters()">
                <option value="all">-- All Leads --</option>
"""
        for lead in ps_unique_leads:
            html += f'                <option value="{lead}">{lead}</option>\n'
        ps_unique_testers = sorted(set(t.get('assigned_to') or 'Unassigned' for t in self.prod_sanity_data))
        html += """
            </select>
        </div>
        <div class="filter-group">
            <label for="psTesterFilter">🔍 Filter by Tester:</label>
            <select id="psTesterFilter" onchange="applyPSFilters()">
                <option value="all">-- All Testers --</option>
"""
        for tester in ps_unique_testers:
            html += f'                <option value="{tester}">{tester}</option>\n'
        html += """
            </select>
        </div>
        <div class="filter-group">
            <label for="psModuleFilter">🔍 Filter by Module:</label>
            <select id="psModuleFilter" onchange="applyPSFilters()">
                <option value="all">-- All Modules --</option>
"""
        for module in ps_unique_modules:
            html += f'                <option value="{module}">{module}</option>\n'
        ps_unique_outcomes = sorted(set(t.get('outcome', 'Not Run') or 'Not Run' for t in self.prod_sanity_data))
        html += """
            </select>
        </div>
        <div class="filter-group">
            <label for="psOutcomeFilter">🔍 Filter by Execution Status (TFS):</label>
            <select id="psOutcomeFilter" onchange="applyPSFilters()">
                <option value="all">-- All Execution Status (TFS) --</option>
"""
        for outcome_opt in ps_unique_outcomes:
            html += f'                <option value="{outcome_opt}">{outcome_opt}</option>\n'
        html += """
            </select>
        </div>
        <div class="filter-group">
            <label for="psStatusFilter">🔍 Filter by Prod Sanity Status:</label>
            <select id="psStatusFilter" onchange="applyPSFilters()">
                <option value="all">-- All Statuses --</option>
"""
        for status in ps_fixed_statuses:
            html += f'                <option value="{status}">{status}</option>\n'
        html += """
            </select>
        </div>
        <button class="reset-btn" onclick="resetPSFilters()">&#8635; Reset Filters</button>
        <button class="save-btn" onclick="saveTabData('prodSanityTab', this)">&#128190; Save</button>
        <button class="export-btn" onclick="exportCurrentTabToExcel()" title="Export current tab to Excel (.xlsx)">&#128229; Export Tab (.xlsx)</button>
        <button class="export-all-btn" onclick="exportAllTabsToExcel()" title="Export all 6 tabs to Excel workbook (.xlsx)">&#128194; Export All Tabs (.xlsx)</button>
        <div class="filter-info"><span id="psVisibleCount">Showing all rows</span></div>
    </div>

    <div class="table-wrapper">
        <table class="report-table" style="font-size:11px;">
            <thead>
                <tr>
                    <th rowspan="2" style="width:30px;">S.No</th>
                    <th rowspan="2" style="min-width:110px;">Module</th>
                    <th rowspan="2" style="min-width:220px;max-width:320px;">Scenarios</th>
                    <th rowspan="2" style="width:110px;">Tester</th>
                    <th rowspan="2" style="width:100px;">Test Lead</th>
                    <th rowspan="2" style="width:100px;">Execution Status (TFS)</th>
                    <th rowspan="2" style="min-width:130px;">Reason for Not Run/NA</th>
                    <th rowspan="2" style="width:100px;">Prod Sanity Status</th>
                    <th colspan="4" style="background:linear-gradient(135deg,#e63946 0%,#c1121f 100%);">Bug Details</th>
                    <th rowspan="2" style="min-width:150px;">Comments</th>
                </tr>
                <tr>
                    <th style="width:90px;">Bug ID</th>
                    <th style="width:80px;">State</th>
                    <th style="width:90px;">Severity</th>
                    <th style="width:110px;">Node Name</th>
                </tr>
            </thead>
            <tbody id="psScenariosBody">
"""

        prod_sanity_status_options = ps_fixed_statuses

        for sno, test in enumerate(self.prod_sanity_data, 1):
            # Escape HTML in text fields
            def esc(v): return str(v).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
            module_display = esc(test.get('text_verification') or '')
            current_ps_status = test.get('prod_sanity_status') or 'Yet to start'
            options_html = ''.join(
                f'<option value="{esc(s)}"{" selected" if s == current_ps_status else ""}>{esc(s)}</option>'
                for s in prod_sanity_status_options
            )
            raw_bug_id = test.get('bug_id', '') or ''
            bug_details_list = test.get('bug_details', []) or []
            comments_val = esc(test.get('comments','') or '')
            ps_tester_val = esc(test.get('assigned_to') or 'Unassigned')
            outcome_val = test.get('outcome', 'Not Run') or 'Not Run'
            _outcome_color_map = {
                'passed': '#28a745', 'pass': '#28a745',
                'failed': '#dc3545', 'fail': '#dc3545',
                'blocked': '#fd7e14',
                'not applicable': '#adb5bd', 'na': '#adb5bd', 'n/a': '#adb5bd',
                'not run': '#6c757d', 'notrun': '#6c757d',
            }
            _oc = _outcome_color_map.get(outcome_val.lower(), '#6c757d')
            outcome_badge = f'<span style="background:{_oc};color:white;padding:2px 5px;border-radius:3px;font-size:9px;font-weight:600;">{esc(outcome_val)}</span>'
            # Bug detail cells — single coordinated block for aligned rendering
            _ado_base = f"https://dev.azure.com/{ADO_CONFIG['organization']}/{ADO_CONFIG['project']}/_workitems/edit"
            def _bug_detail_cells(bugs, base):
                """Returns (first_4_cells, [extra_4_cells, ...]). Each is 4 <td> strings.
                0 bugs: (4 dash tds, []). 1+ bugs: (first tds, [rest...])."""
                dash = '<td style="text-align:center;color:#adb5bd;font-size:10px;padding:4px 6px;vertical-align:middle;">—</td>'
                if not bugs:
                    return (dash * 4, [])
                def _lnk(b):
                    bid = str(b.get('bug_id',''))
                    return (f'<a href="{base}/{bid}" target="_blank" style="color:#dc3545;'
                            f'font-weight:600;font-size:10px;text-decoration:none;" '
                            f'title="Open Bug {bid} in ADO">#{bid}</a>')
                def _four(b):
                    nn = esc(b.get('node_name','') or '')
                    return (
                        f'<td style="text-align:center;font-size:10px;vertical-align:middle;padding:4px 6px;">{_lnk(b)}</td>'
                        f'<td style="text-align:center;font-size:10px;vertical-align:middle;padding:4px 6px;">{esc(b.get("state",""))}</td>'
                        f'<td style="text-align:center;font-size:10px;vertical-align:middle;padding:4px 6px;">{esc(b.get("severity",""))}</td>'
                        f'<td style="text-align:left;padding:4px 6px;font-size:10px;vertical-align:middle;">{nn}</td>'
                    )
                return (_four(bugs[0]), [_four(b) for b in bugs[1:]])
            _first_bug_cells, _extra_bug_list = _bug_detail_cells(bug_details_list, _ado_base)
            _bug_n = max(1, len(bug_details_list))
            _rs = f' rowspan="{_bug_n}"' if _bug_n > 1 else ''
            _extra_trs = ''.join(
                f'                <tr class="bug-extra-row" '
                f'data-lead="{esc(test["lead"])}" data-module="{module_display}" '
                f'data-status="{esc(current_ps_status)}" data-tester="{ps_tester_val}" '
                f'data-outcome="{esc(outcome_val)}">{xc}</tr>\n'
                for xc in _extra_bug_list
            )
            name_attr = esc(test['name']).replace('"', '&quot;')
            html += f"""
                <tr data-lead="{esc(test['lead'])}" data-module="{module_display}" data-status="{esc(current_ps_status)}" data-tester="{ps_tester_val}" data-outcome="{esc(outcome_val)}">
                    <td{_rs} style="text-align:center;font-weight:600;">{sno}</td>
                    <td{_rs} style="text-align:left;padding-left:6px;">{module_display}</td>
                    <td{_rs} class="title-col" title="{name_attr}">{esc(test['name'])}</td>
                    <td{_rs}>{esc(test.get('assigned_to',''))}</td>
                    <td{_rs} style="font-weight:600;">{esc(test['lead'])}</td>
                    <td{_rs} style="text-align:center;">{outcome_badge}</td>
                    <td{_rs}>{esc(test.get('text_verification1') or '')}</td>
                    <td{_rs}><select style="font-size:10px;width:100%;border:1px solid #ccc;border-radius:3px;padding:2px;">{options_html}</select></td>
                    {_first_bug_cells}
                    <td{_rs}><input type="text" value="{comments_val}" style="font-size:10px;width:100%;border:1px solid #ccc;border-radius:3px;padding:2px;"/></td>
                </tr>
{_extra_trs}"""

        html += """
            </tbody>
        </table>
    </div>
</div>

        <!-- Tab 2: Prod US Sanity -->
        <div id="prodUSSanityTab" class="tab-content">
            <div class="filter-section">
                <div class="filter-group">
                    <label for="usLeadFilter">🔍 Filter by Lead:</label>
                    <select id="usLeadFilter" onchange="applyUSFilters()">
                        <option value="all">-- All Leads --</option>
"""
        # Add unique lead options for Tab 2 (all test types - shows all stories)
        us_unique_leads = sorted(set(t['lead'] for t in self.test_data))
        for lead in us_unique_leads:
            html += f'                        <option value="{lead}">{lead}</option>\n'

        html += """                    </select>
                </div>

                <div class="filter-group">
                    <label for="usTesterFilter">🔍 Filter by Tester:</label>
                    <select id="usTesterFilter" onchange="applyUSFilters()">
                        <option value="all">-- All Testers --</option>
"""
        us_unique_testers = sorted(set(t.get('assigned_to','') or 'Unassigned' for t in self.test_data))
        for tester in us_unique_testers:
            html += f'                        <option value="{tester}">{tester}</option>\n'

        html += """                    </select>
                </div>

                <div class="filter-group">
                    <label for="usModuleFilter">🔍 Filter by Module:</label>
                    <select id="usModuleFilter" onchange="applyUSFilters()">
                        <option value="all">-- All Modules --</option>
"""
        us_unique_modules = sorted(set(t.get('text_verification') or '' for t in self.test_data if t.get('text_verification')))
        for mod in us_unique_modules:
            html += f'                        <option value="{mod}">{mod}</option>\n'

        html += """                    </select>
                </div>

                <div class="filter-group">
                    <label for="usOutcomeFilter">🔍 Filter by Execution Status (TFS):</label>
                    <select id="usOutcomeFilter" onchange="applyUSFilters()">
                        <option value="all">-- All Execution Status (TFS) --</option>
                        <option value="Passed">Passed</option>
                        <option value="Failed">Failed</option>
                        <option value="Blocked">Blocked</option>
                        <option value="Not Applicable">Not Applicable</option>
                        <option value="Not Run">Not Run</option>
                    </select>
                </div>

                <div class="filter-group">
                    <label for="usStatusFilter">🔍 Filter by Prod Sanity Status:</label>
                    <select id="usStatusFilter" onchange="applyUSFilters()">
                        <option value="all">-- All Statuses --</option>
"""
        for s in prod_sanity_status_options:
            html += f'                        <option value="{s}">{s}</option>\n'

        html += """                    </select>
                </div>

                <button class="reset-btn" onclick="resetUSFilters()">↻ Reset Filters</button>
                <button class="save-btn" onclick="saveTabData('prodUSSanityTab', this)">&#128190; Save</button>
                <button class="export-btn" onclick="exportCurrentTabToExcel()" title="Export current tab to Excel (.xlsx)">&#128229; Export Tab (.xlsx)</button>
                <button class="export-all-btn" onclick="exportAllTabsToExcel()" title="Export all 6 tabs to Excel workbook (.xlsx)">&#128194; Export All Tabs (.xlsx)</button>

                <div class="filter-info">
                    <span id="usVisibleCount">Showing all rows</span>
                </div>
            </div>

            <div class="table-wrapper">
                <table class="report-table" style="font-size:11px;">
                    <thead>
                        <tr>
                            <th rowspan="2" style="width:30px;">S.No</th>
                            <th rowspan="2" style="min-width:80px;">PT Lead</th>
                            <th rowspan="2" style="min-width:110px;">Module</th>
                            <th rowspan="2" style="min-width:220px;max-width:320px;">Title</th>
                            <th rowspan="2" style="width:120px;">US ID</th>
                            <th rowspan="2" style="width:110px;">Tester</th>
                            <th rowspan="2" style="width:110px;">Execution Status (TFS)</th>
                            <th rowspan="2" style="width:110px;">Reason for Not Run/NA</th>
                            <th rowspan="2" style="width:110px;">Prod Sanity Status</th>
                            <th colspan="4" style="background:linear-gradient(135deg,#e63946 0%,#c1121f 100%);">Bug Details</th>
                            <th rowspan="2" style="min-width:150px;">Comments</th>
                        </tr>
                        <tr>
                            <th style="width:90px;">Bug ID</th>
                            <th style="width:80px;">State</th>
                            <th style="width:90px;">Severity</th>
                            <th style="width:110px;">Node Name</th>
                        </tr>
                    </thead>
                    <tbody id="usSanityTableBody">
"""
        # Add Tab 2 data rows - one row per unique User Story
        # -------------------------------------------------------
        # Group ALL tests (manual + automation) by their resolved US ID
        # so every story in the Prod Execution suite appears here.
        # -------------------------------------------------------
        # _resolve_us_key: same priority as _pivot_us_key (us_id first, then module regex,
        # then name) — keeps the table row count consistent with the Overall pivot.
        def _resolve_us_key(t):
            """Return (us_id_str, us_title_str, lookup_int) for a test record.
            Parses the module field which contains values like:
            '4176533 : [GEN AI - QDI] Instances to be displayed...'
            """
            raw_tv = t.get('module') or ''
            # Supports separators: " : ", " - ", " – ", " — "
            tv_match = re.match(r'^(\d+)\s*[:]\s*(.+)$', raw_tv, re.DOTALL) \
                    or re.match(r'^(\d+)\s*[-\u2013\u2014]\s*(.+)$', raw_tv, re.DOTALL)
            tv_us_id    = tv_match.group(1).strip()    if tv_match else ''
            tv_us_title = tv_match.group(2).strip()    if tv_match else ''
            us_id_str   = str(t.get('us_id') or tv_us_id or t['name'])
            us_title_str = t.get('us_title') or tv_us_title or t['name']
            try:
                lookup_int = int(t.get('us_id') or tv_us_id or 0)
            except (ValueError, TypeError):
                lookup_int = 0
            return us_id_str, us_title_str, lookup_int

        # Build ordered mapping: us_id_str -> representative story dict
        # Include ALL test types so every story in the suite is represented.
        from collections import OrderedDict
        us_story_map = OrderedDict()
        for t in self.test_data:
            us_id_str, us_title_str, lookup_int = _resolve_us_key(t)
            key = us_id_str if us_id_str else t['name']
            if key not in us_story_map:
                us_story_map[key] = {
                    'us_id_str':   us_id_str,
                    'us_title_str': us_title_str,
                    'lookup_int':  lookup_int,
                    'lead':        t['lead'],
                    'module':      t['module'],
                    'text_verification': t.get('text_verification') or '',
                    'text_verification1': t.get('text_verification1') or '',
                    'assigned_to': t.get('assigned_to') or 'Unassigned',
                    'prod_sanity_status': t.get('prod_sanity_status') or '',
                    'comments':    t.get('comments') or '',
                    'outcomes':    [],
                }
            if not us_story_map[key]['comments'] and t.get('comments'):
                us_story_map[key]['comments'] = t['comments']
            us_story_map[key]['outcomes'].append(t.get('outcome') or 'Not Run')

        _us_outcome_color_map = {
            'passed':   '#28a745',
            'failed':   '#dc3545',
            'blocked':  '#fd7e14',
            'not applicable': '#adb5bd',
            'not run':  '#6c757d',
        }

        def _aggregate_us_outcome(outcomes):
            """Aggregate test-case outcomes for a story row.
            Rules (in priority order):
              1. Any Failed          → Failed
              2. Any Blocked         → Blocked
              3. All Passed          → Passed
              4. Otherwise           → Not Run
            """
            flags = [str(o).strip().lower() for o in outcomes if o]
            if any(f in ('failed', 'fail') for f in flags):
                return 'Failed'
            if any(f in ('blocked', 'block') for f in flags):
                return 'Blocked'
            if flags and all(f in ('passed', 'pass') for f in flags):
                return 'Passed'
            return 'Not Run'

        for us_sno, (key, story) in enumerate(us_story_map.items(), 1):
            def esc(v): return str(v).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

            us_tester         = esc(story['assigned_to'])
            us_module_val     = esc(story['text_verification'])
            us_reason_val     = esc(story.get('text_verification1') or '')
            us_id_display     = esc(story['us_id_str'])
            us_title_display  = esc(story['us_title_str'])
            us_comments_val   = esc(story['comments'])

            # Aggregate outcome across all test cases for this story
            raw_outcomes = story.get('outcomes', [])
            worst_outcome = _aggregate_us_outcome(raw_outcomes)
            _oc = _us_outcome_color_map.get(worst_outcome.lower(), '#6c757d')
            us_outcome_badge = f'<span style="background:{_oc};color:white;padding:2px 5px;border-radius:3px;font-size:9px;font-weight:600;">{esc(worst_outcome)}</span>'

            us_current_status = story['prod_sanity_status'] or 'Yet to start'
            us_options_html = ''.join(
                f'<option value="{esc(s)}"{" selected" if s == us_current_status else ""}>{esc(s)}</option>'
                for s in prod_sanity_status_options
            )

            # Bug detail cells — single coordinated block for aligned rendering
            linked_bugs = self.us_bug_map.get(story['lookup_int'], []) if story['lookup_int'] else []
            _us_ado_base = f"https://dev.azure.com/{ADO_CONFIG['organization']}/{ADO_CONFIG['project']}/_workitems/edit"
            def _bug_detail_cells(bugs, base):
                """Returns (first_4_cells, [extra_4_cells, ...]). Each is 4 <td> strings.
                0 bugs: (4 dash tds, []). 1+ bugs: (first tds, [rest...])."""
                dash = '<td style="text-align:center;color:#adb5bd;font-size:10px;padding:4px 6px;vertical-align:middle;">—</td>'
                if not bugs:
                    return (dash * 4, [])
                def _lnk(b):
                    bid = str(b.get('bug_id',''))
                    return (f'<a href="{base}/{bid}" target="_blank" style="color:#dc3545;'
                            f'font-weight:600;font-size:10px;text-decoration:none;" '
                            f'title="Open Bug {bid} in ADO">#{bid}</a>')
                def _four(b):
                    nn = esc(b.get('node_name','') or '')
                    return (
                        f'<td style="text-align:center;font-size:10px;vertical-align:middle;padding:4px 6px;">{_lnk(b)}</td>'
                        f'<td style="text-align:center;font-size:10px;vertical-align:middle;padding:4px 6px;">{esc(b.get("state",""))}</td>'
                        f'<td style="text-align:center;font-size:10px;vertical-align:middle;padding:4px 6px;">{esc(b.get("severity",""))}</td>'
                        f'<td style="text-align:left;padding:4px 6px;font-size:10px;vertical-align:middle;">{nn}</td>'
                    )
                return (_four(bugs[0]), [_four(b) for b in bugs[1:]])
            _first_us_bug_cells, _extra_us_bug_list = _bug_detail_cells(linked_bugs, _us_ado_base)
            _us_bug_n = max(1, len(linked_bugs))
            _us_rs = f' rowspan="{_us_bug_n}"' if _us_bug_n > 1 else ''
            _us_extra_trs = ''.join(
                f'                        <tr class="bug-extra-row" '
                f'data-lead="{esc(story["lead"])}" data-tester="{us_tester}" '
                f'data-module="{esc(story["text_verification"])}" '
                f'data-status="{esc(story["prod_sanity_status"])}" data-outcome="{esc(worst_outcome)}">{xc}</tr>\n'
                for xc in _extra_us_bug_list
            )

            us_title_attr = us_title_display.replace('"', '&quot;')
            html += f"""
                        <tr data-lead="{esc(story['lead'])}" data-tester="{us_tester}" data-module="{esc(story['text_verification'])}" data-status="{esc(story['prod_sanity_status'])}" data-outcome="{esc(worst_outcome)}">
                            <td{_us_rs} style="text-align:center;font-weight:600;">{us_sno}</td>
                            <td{_us_rs} style="font-weight:600;">{esc(story['lead'])}</td>
                            <td{_us_rs} style="text-align:left;padding-left:6px;">{us_module_val}</td>
                            <td{_us_rs} class="title-col" title="{us_title_attr}">{us_title_display}</td>
                            <td{_us_rs} style="text-align:center;">{us_id_display}</td>
                            <td{_us_rs}>{us_tester}</td>
                            <td{_us_rs} style="text-align:center;">{us_outcome_badge}</td>
                            <td{_us_rs}>{us_reason_val}</td>
                            <td{_us_rs}><select style="font-size:10px;width:100%;border:1px solid #ccc;border-radius:3px;padding:2px;">{us_options_html}</select></td>
                            {_first_us_bug_cells}
                            <td{_us_rs}><input type="text" value="{us_comments_val}" style="font-size:10px;width:100%;border:1px solid #ccc;border-radius:3px;padding:2px;"/></td>
                        </tr>
{_us_extra_trs}"""

        html += """
                    </tbody>
                </table>
            </div>
        </div>

        <!-- Tab 3: Insprint US Prod Status -->
        <div id="insprintStatusTab" class="tab-content">
            <div class="filter-section">
                <div class="filter-group">
                    <label for="leadFilterInsprint">🔍 Filter by Lead:</label>
                    <select id="leadFilterInsprint" onchange="applyFiltersInsprint()">
                        <option value="all">-- All Leads --</option>
"""
        # Use ALL insprint_data (not filtered by type/automation_status)
        insprint_all = self.insprint_data

        insprint_fixed_statuses = [
            'Working fine',
            'Working fine with active bug',
            'In Progress',
            'Failed',
            'Blocked',
            'Cannot be validated',
            'Can be validated on Monday',
            'Taken care by BPMS',
            'IMS test team',
            'Story not deployed to prod as part of this release',
            'Toggle off/Code reversal validated',
            'Yet to start',
        ]

        insprint_unique_leads = sorted(set(t.get('lead', 'Unassigned') for t in insprint_all))
        for lead in insprint_unique_leads:
            html += f'                <option value="{esc(lead)}">{esc(lead)}</option>\n'

        insprint_unique_testers = sorted(set(t.get('assigned_to') or 'Unassigned' for t in insprint_all))
        html += """                    </select>
                </div>
                <div class="filter-group">
                    <label for="insprintTesterFilter">&#128269; Filter by Tester:</label>
                    <select id="insprintTesterFilter" onchange="applyFiltersInsprint()">
                        <option value="all">-- All Testers --</option>
"""
        for tester in insprint_unique_testers:
            html += f'                <option value="{esc(tester)}">{esc(tester)}</option>\n'
        html += """                    </select>
                </div>
                <div class="filter-group">
                    <label for="insprintModuleFilter">&#128269; Filter by Module:</label>
                    <select id="insprintModuleFilter" onchange="applyFiltersInsprint()">
                        <option value="all">-- All Modules --</option>
"""
        insprint_unique_modules = sorted(set(t.get('text_verification') or '' for t in insprint_all if t.get('text_verification')))
        for mod in insprint_unique_modules:
            html += f'                <option value="{esc(mod)}">{esc(mod)}</option>\n'
        html += """                    </select>
                </div>
                <div class="filter-group">
                    <label for="insprintOutcomeFilter">&#128269; Filter by Execution Status (TFS):</label>
                    <select id="insprintOutcomeFilter" onchange="applyFiltersInsprint()">
                        <option value="all">-- All Execution Status (TFS) --</option>
                        <option value="Passed">Passed</option>
                        <option value="Failed">Failed</option>
                        <option value="Blocked">Blocked</option>
                        <option value="Not Applicable">Not Applicable</option>
                        <option value="Not Run">Not Run</option>
                    </select>
                </div>
                <div class="filter-group">
                    <label for="insprintStatusFilter">&#128269; Filter by Prod Sanity Status:</label>
                    <select id="insprintStatusFilter" onchange="applyFiltersInsprint()">
                        <option value="all">-- All Statuses --</option>
"""
        for status in insprint_fixed_statuses:
            html += f'                <option value="{status}">{status}</option>\n'
        html += """                    </select>
                </div>
                <button class="reset-btn" onclick="resetFiltersInsprint()">&#8635; Reset Filters</button>
                <button class="save-btn" onclick="saveTabData('insprintStatusTab', this)">&#128190; Save</button>
                <button class="export-btn" onclick="exportCurrentTabToExcel()" title="Export current tab to Excel (.xlsx)">&#128229; Export Tab (.xlsx)</button>
                <button class="export-all-btn" onclick="exportAllTabsToExcel()" title="Export all 6 tabs to Excel workbook (.xlsx)">&#128194; Export All Tabs (.xlsx)</button>
                <div class="filter-info"><span id="visibleCountInsprint">Showing all rows</span></div>
            </div>

            <div class="table-wrapper">
                <table class="report-table" style="font-size:11px;">
                    <thead>
                        <tr>
                            <th rowspan="2" class="sno-col">S.No</th>
                            <th rowspan="2" class="lead-col">PT Lead</th>
                            <th rowspan="2" style="min-width:110px;">Module</th>
                            <th rowspan="2" style="min-width:220px;max-width:320px;">Title</th>
                            <th rowspan="2" style="width:120px;">US ID</th>
                            <th rowspan="2" style="width:110px;">Tester</th>
                            <th rowspan="2" style="width:100px;">Execution Status (TFS)</th>
                            <th rowspan="2" style="min-width:130px;">Reason for Not Run/NA</th>
                            <th rowspan="2" style="width:100px;">Prod Sanity Status</th>
                            <th colspan="4" style="background:linear-gradient(135deg,#e63946 0%,#c1121f 100%);">Bug Details</th>
                            <th rowspan="2" style="min-width:150px;">Comments</th>
                        </tr>
                        <tr>
                            <th style="width:90px;">Bug ID</th>
                            <th style="width:80px;">State</th>
                            <th style="width:90px;">Severity</th>
                            <th style="width:110px;">Node Name</th>
                        </tr>
                    </thead>
                    <tbody id="reportTableBodyInsprint">
"""
        # Build per-US story map from all insprint data
        from collections import OrderedDict as _OD

        def _ins_aggregate_outcome(outcomes):
            """Aggregate test-case outcomes for a US story row.
            Rules:
              1. All Passed (ignoring NA) → Passed
              2. Any Failed → Failed
              3. Any Blocked (no Failed) → Blocked
              4. Failed + Blocked mix → Failed
              5. All NA → Not Applicable
              6. Some NA → ignore, apply rules to rest
            """
            flags = [str(o).strip().lower() for o in outcomes if o]
            non_na = [f for f in flags if f not in ('not applicable', 'na', 'n/a', 'notapplicable')]
            if not non_na:
                if flags:
                    return 'Not Applicable'
                return 'Not Run'
            if any(f in ('failed', 'fail') for f in non_na):
                return 'Failed'
            if any(f in ('blocked', 'block') for f in non_na):
                return 'Blocked'
            if all(f in ('passed', 'pass') for f in non_na):
                return 'Passed'
            return 'Not Run'

        _ins_outcome_color_map = {
            'passed':         '#28a745',
            'failed':         '#dc3545',
            'blocked':        '#fd7e14',
            'not applicable': '#adb5bd',
            'not run':        '#6c757d',
        }

        ins_story_map = _OD()
        for t in insprint_all:
            raw_mod = t.get('module') or ''
            mod_match = re.match(r'^(\d+)\s*[:]\s*(.+)$', raw_mod, re.DOTALL) \
                     or re.match(r'^(\d+)\s*[-\u2013\u2014]\s*(.+)$', raw_mod, re.DOTALL)
            ins_us_id_str    = mod_match.group(1).strip() if mod_match else str(t.get('us_id') or '')
            ins_us_title_str = mod_match.group(2).strip() if mod_match else (t.get('us_title') or t['name'])
            try:
                ins_lookup_int = int(ins_us_id_str or 0)
            except (ValueError, TypeError):
                ins_lookup_int = 0
            key = ins_us_id_str if ins_us_id_str else t['name']
            if key not in ins_story_map:
                ins_story_map[key] = {
                    'us_id_str':    ins_us_id_str,
                    'us_title_str': ins_us_title_str,
                    'lookup_int':   ins_lookup_int,
                    'lead':         t.get('lead', ''),
                    'text_verification': t.get('text_verification') or '',
                    'text_verification1': t.get('text_verification1') or '',
                    'assigned_to':  t.get('assigned_to') or 'Unassigned',
                    'prod_sanity_status': t.get('prod_sanity_status') or '',
                    'comments':     t.get('comments') or '',
                    'outcomes':     [],
                }
            if not ins_story_map[key]['comments'] and t.get('comments'):
                ins_story_map[key]['comments'] = t['comments']
            ins_story_map[key]['outcomes'].append(t.get('outcome') or 'Not Run')

        _ins_ado_base = f"https://dev.azure.com/{ADO_CONFIG['organization']}/{ADO_CONFIG['project']}/_workitems/edit"

        for ins_sno, (key, story) in enumerate(ins_story_map.items(), 1):
            current_ins_status = story['prod_sanity_status'] or 'Yet to start'
            ins_tester    = esc(story['assigned_to'])
            ins_us_id     = esc(story['us_id_str'])
            ins_us_title  = esc(story['us_title_str'])
            ins_us_title_attr = ins_us_title.replace('"', '&quot;')
            ins_module_val = esc(story['text_verification'])
            ins_reason_val = esc(story.get('text_verification1') or '')
            ins_comments_val = esc(story['comments'])

            # Aggregate outcome
            worst_outcome = _ins_aggregate_outcome(story['outcomes'])
            _oc = _ins_outcome_color_map.get(worst_outcome.lower(), '#6c757d')
            ins_outcome_badge = f'<span style="background:{_oc};color:white;padding:2px 5px;border-radius:3px;font-size:9px;font-weight:600;">{esc(worst_outcome)}</span>'

            # Prod Sanity Status dropdown
            status_opts = ''
            for s in insprint_fixed_statuses:
                sel = ' selected' if s == current_ins_status else ''
                status_opts += f'<option value="{esc(s)}"{sel}>{esc(s)}</option>'

            # Bug data from us_bug_map — single coordinated block for aligned rendering
            linked_bugs = self.us_bug_map.get(story['lookup_int'], []) if story['lookup_int'] else []
            def _bug_detail_cells(bugs, base):
                """Returns (first_4_cells, [extra_4_cells, ...]). Each is 4 <td> strings.
                0 bugs: (4 dash tds, []). 1+ bugs: (first tds, [rest...])."""
                dash = '<td style="text-align:center;color:#adb5bd;font-size:10px;padding:4px 6px;vertical-align:middle;">—</td>'
                if not bugs:
                    return (dash * 4, [])
                def _lnk(b):
                    bid = str(b.get('bug_id',''))
                    return (f'<a href="{base}/{bid}" target="_blank" style="color:#dc3545;'
                            f'font-weight:600;font-size:10px;text-decoration:none;" '
                            f'title="Open Bug {bid} in ADO">#{bid}</a>')
                def _four(b):
                    nn = esc(b.get('node_name','') or '')
                    return (
                        f'<td style="text-align:center;font-size:10px;vertical-align:middle;padding:4px 6px;">{_lnk(b)}</td>'
                        f'<td style="text-align:center;font-size:10px;vertical-align:middle;padding:4px 6px;">{esc(b.get("state",""))}</td>'
                        f'<td style="text-align:center;font-size:10px;vertical-align:middle;padding:4px 6px;">{esc(b.get("severity",""))}</td>'
                        f'<td style="text-align:left;padding:4px 6px;font-size:10px;vertical-align:middle;">{nn}</td>'
                    )
                return (_four(bugs[0]), [_four(b) for b in bugs[1:]])
            _first_ins_bug_cells, _extra_ins_bug_list = _bug_detail_cells(linked_bugs, _ins_ado_base)
            _ins_bug_n = max(1, len(linked_bugs))
            _ins_rs = f' rowspan="{_ins_bug_n}"' if _ins_bug_n > 1 else ''
            _ins_extra_trs = ''.join(
                f'                        <tr class="bug-extra-row" '
                f'data-lead="{esc(story["lead"])}" data-tester="{ins_tester}" '
                f'data-module="{esc(story["text_verification"])}" '
                f'data-outcome="{esc(worst_outcome)}" data-status="{esc(current_ins_status)}">{xc}</tr>\n'
                for xc in _extra_ins_bug_list
            )

            html += f"""
                        <tr data-lead="{esc(story['lead'])}" data-tester="{ins_tester}" data-module="{esc(story['text_verification'])}" data-outcome="{esc(worst_outcome)}" data-status="{esc(current_ins_status)}">
                            <td{_ins_rs} class="sno-col">{ins_sno}</td>
                            <td{_ins_rs}>{esc(story['lead'])}</td>
                            <td{_ins_rs} style="text-align:left;padding-left:6px;">{ins_module_val}</td>
                            <td{_ins_rs} class="title-col" title="{ins_us_title_attr}">{ins_us_title}</td>
                            <td{_ins_rs} style="text-align:center;">{ins_us_id}</td>
                            <td{_ins_rs}>{ins_tester}</td>
                            <td{_ins_rs} style="text-align:center;">{ins_outcome_badge}</td>
                            <td{_ins_rs}>{ins_reason_val}</td>
                            <td{_ins_rs}><select class="inline-select" style="font-size:10px;width:100%;border:1px solid #ccc;border-radius:3px;padding:2px;" onchange="this.closest('tr').setAttribute('data-status',this.value)">
                                <option value="">-- Select --</option>
                                {status_opts}
                            </select></td>
                            {_first_ins_bug_cells}
                            <td{_ins_rs}><input type="text" class="inline-input" value="{ins_comments_val}" style="font-size:10px;width:100%;border:1px solid #ccc;border-radius:3px;padding:2px;"/></td>
                        </tr>
{_ins_extra_trs}"""

        if not ins_story_map:
            html += '                    <tr><td colspan="14" style="text-align:center;color:#6b7280;padding:20px;">No data found in Insprint US Prod Status suite.</td></tr>\n'

        html += """                    </tbody>
                </table>
            </div>
        </div>

        <!-- Tab 4: Ready for Prod Bug -->
        <div id="readyForProdBugTab" class="tab-content">
            <div style="padding: 20px;">
                <h2 style="text-align: center; color: #dc2626; margin-bottom: 15px; font-size: 16px;">🐛 Overall Ready for Prod bugs</h2>
                
                <!-- Bug List Filters -->
                <div class="filter-section">
                    <div class="filter-group">
                        <label for="bugMpocFilter">🔍 Filter by MPOC:</label>
                        <select id="bugMpocFilter" onchange="updateStateOptions(); applyBugFilters();">
                            <option value="all">-- All MPOCs --</option>
"""
        
        # Add unique MPOC options with case-insensitive deduplication and include Unassigned
        mpoc_dict = {}
        has_unassigned = False
        
        for bug in self.bug_data:
            mpoc = bug.get('mpoc', '')
            if mpoc in ['', 'N/A']:
                has_unassigned = True
            elif mpoc == 'Unassigned':
                has_unassigned = True
            else:
                # Use lowercase as key to prevent case-sensitive duplicates
                mpoc_lower = mpoc.lower()
                if mpoc_lower not in mpoc_dict:
                    mpoc_dict[mpoc_lower] = mpoc
        
        # Get unique MPOCs sorted
        unique_mpocs = sorted(mpoc_dict.values())
        
        # Add Unassigned first if exists
        if has_unassigned:
            html += f"""                            <option value="Unassigned">Unassigned</option>\n"""
        
        # Add other MPOCs
        for mpoc in unique_mpocs:
            html += f"""                            <option value="{mpoc}">{mpoc}</option>\n"""
        
        # Lead filter (unique PT Leads from POD_LEAD_MAP lookup on bug rows)
        bug_unique_leads = sorted(set(
            POD_LEAD_MAP.get(bug.get('node_name', '').lower(), '') for bug in self.bug_data
            if POD_LEAD_MAP.get(bug.get('node_name', '').lower(), '')
        ))
        html += """                        </select>
                    </div>
                    <div class="filter-group">
                        <label for="bugLeadFilter">&#128269; Filter by Lead:</label>
                        <select id="bugLeadFilter" onchange="applyBugFilters()">
                            <option value="all">-- All Leads --</option>
"""
        for lead in bug_unique_leads:
            html += f'                            <option value="{lead}">{lead}</option>\n'
        html += """                        </select>
                    </div>
                    <div class="filter-group">
                        <label for="bugPsStatusFilter">&#128269; Filter by Prod Sanity Status:</label>
                        <select id="bugPsStatusFilter" onchange="applyBugFilters()">
                            <option value="all">-- All Statuses --</option>
"""
        for s in [
            'Working fine', 'Working fine with active bug', 'In Progress',
            'Failed', 'Blocked', 'Cannot be validated', 'Can be validated on Monday',
            'Taken care by BPMS', 'IMS test team',
            'Story not deployed to prod as part of this release',
            'Toggle off/Code reversal validated', 'Yet to start',
        ]:
            html += f'                            <option value="{s}">{s}</option>\n'
        html += """                        </select>
                    </div>
                    <div class="filter-group">
                        <label>&#128269; Filter by State:</label>
                        <div class="custom-dropdown">
                            <button type="button" class="dropdown-toggle" id="stateDropdownToggle" onclick="toggleStateDropdown()">
                                All States ▼
                            </button>
                            <div class="dropdown-menu" id="stateDropdownMenu">
                                <div class="dropdown-item">
                                    <label>
                                        <input type="checkbox" value="all" checked onchange="toggleAllStates(this)"> All States
                                    </label>
                                </div>
                                <div class="dropdown-divider"></div>
"""
        
        # Add unique state options with checkboxes
        unique_states = sorted(set([bug['state'] for bug in self.bug_data if bug.get('state')]))
        for state in unique_states:
            html += f"""                                <div class="dropdown-item">
                                    <label>
                                        <input type="checkbox" class="state-checkbox" value="{state}" checked onchange="updateStateFilter()"> {state}
                                    </label>
                                </div>\n"""
        
        html += """                            </div>
                        </div>
                    </div>
                    
                    <button class="reset-btn" onclick="resetBugFilters()">↻ Reset Filters</button>
                    <button class="save-btn" onclick="saveTabData('readyForProdBugTab', this)">&#128190; Save</button>
                    <button class="export-btn" onclick="exportCurrentTabToExcel()" title="Export current tab to Excel (.xlsx)">&#128229; Export Tab (.xlsx)</button>
                    <button class="export-all-btn" onclick="exportAllTabsToExcel()" title="Export all 6 tabs to Excel workbook (.xlsx)">&#128194; Export All Tabs (.xlsx)</button>
                    
                    <div class="filter-info">
                        <span id="bugVisibleCount">Showing all bugs</span>
                    </div>
                </div>
                
                <!-- Bug List Table -->
                <div class="table-wrapper">
                    <table class="bug-list-table">
                        <thead>
                            <tr>
                                <th class="bug-mpoc-col">ExternalRef ID</th>
                                <th class="bug-id-col">ID</th>
                                <th class="bug-title-col">Title</th>
                                <th class="bug-state-col">State</th>
                                <th class="bug-defect-col">Defect Record</th>
                                <th class="bug-severity-col">Severity</th>
                                <th class="bug-node-col">Node Name</th>
                                <th class="bug-stage-col">StageFound</th>
                                <th>PT Lead</th>
                                <th>Tester</th>
                                <th>Prod Sanity Status</th>
                            </tr>
                        </thead>
                        <tbody id="bugListTableBody">
"""
        bug_ps_statuses = [
            'Working fine',
            'Working fine with active bug',
            'In Progress',
            'Failed',
            'Blocked',
            'Cannot be validated',
            'Can be validated on Monday',
            'Taken care by BPMS',
            'IMS test team',
            'Story not deployed to prod as part of this release',
            'Toggle off/Code reversal validated',
            'Yet to start',
        ]
        # Add bug rows
        for bug in self.bug_data:
            bug_id = bug.get('id', 'N/A')
            mpoc = bug.get('mpoc', 'Unassigned')
            title = bug.get('title', 'N/A')
            state = bug.get('state', 'N/A')
            defect_record = bug.get('defect_record', 'N/A')
            severity = bug.get('severity', 'N/A')
            node_name = bug.get('node_name', 'N/A')
            stage_found = bug.get('stage_found', 'N/A')
            pt_lead = POD_LEAD_MAP.get(node_name.lower(), '')
            bug_ps_opts = ''.join(
                f'<option value="{s}">{s}</option>' for s in bug_ps_statuses
            )
            html += f"""
                            <tr data-mpoc="{mpoc}" data-state="{state}" data-lead="{pt_lead}" data-ps-status="">
                                <td class="bug-mpoc-col">{mpoc}</td>
                                <td class="bug-id-col">{bug_id}</td>
                                <td class="bug-title-col">{title}</td>
                                <td class="bug-state-col">{state}</td>
                                <td class="bug-defect-col">{defect_record}</td>
                                <td class="bug-severity-col">{severity}</td>
                                <td class="bug-node-col">{node_name}</td>
                                <td class="bug-stage-col">{stage_found}</td>
                                <td>{pt_lead}</td>
                                <td><input type="text" class="inline-input" placeholder="Enter tester..." style="font-size:10px;width:100%;border:1px solid #ccc;border-radius:3px;padding:2px;"/></td>
                                <td><select class="inline-select" style="min-width:160px">
                                    <option value="">-- Select --</option>
                                    {bug_ps_opts}
                                </select></td>
                            </tr>
"""
        
        html += """                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- Tab 5: Prod Sanity Defects -->
        <div id="prodSanityDefectsTab" class="tab-content">
            <div style="padding: 20px;">
                <h2 style="text-align: center; color: #7c3aed; margin-bottom: 15px; font-size: 16px;">&#128308; Prod Sanity Defects</h2>
"""
        # ── Defect Summary calculations ────────────────────────────────────────
        _sev_order  = ['1 - Critical', '2 - High', '3 - Medium', '4 - Low']
        _sev_labels = {'1 - Critical': 'Critical', '2 - High': 'High', '3 - Medium': 'Medium', '4 - Low': 'Low'}
        _sev_colors = {'1 - Critical': '#8b0000', '2 - High': '#e02020', '3 - Medium': '#e6a817', '4 - Low': '#f5f0c0'}
        _sev_text   = {'1 - Critical': 'white',    '2 - High': 'white',  '3 - Medium': '#333',    '4 - Low': '#333'}

        def _norm_sev(s):
            s = str(s or '')
            for k in _sev_order:
                if k.lower() in s.lower(): return k
            return 'Other'

        _total_count = len(self.prod_sanity_defects)
        _sev_totals  = {k: 0 for k in _sev_order}
        for _d in self.prod_sanity_defects:
            _sk = _norm_sev(_d.get('severity', ''))
            if _sk in _sev_totals:
                _sev_totals[_sk] += 1

        # State breakdown rows — same 4 states as the screenshot
        _breakdown_states = ['Active', 'Fixed and Ready to Deploy', 'Under Testing', 'Closed']
        # Build counts: {state: {sev: count}}
        _bd = {s: {k: 0 for k in _sev_order} for s in _breakdown_states}
        _bd_total = {s: 0 for s in _breakdown_states}
        _total_raised_by_sev = {k: 0 for k in _sev_order}
        _total_raised = 0
        for _d in self.prod_sanity_defects:
            _st  = str(_d.get('state', '') or '').strip()
            _sk  = _norm_sev(_d.get('severity', ''))
            _total_raised += 1
            if _sk in _total_raised_by_sev:
                _total_raised_by_sev[_sk] += 1
            for _bs in _breakdown_states:
                if _st.lower() == _bs.lower():
                    _bd_total[_bs] += 1
                    if _sk in _bd[_bs]:
                        _bd[_bs][_sk] += 1

        # ── Defect query link ──────────────────────────────────────────────────
        _defect_query_url = 'https://dev.azure.com/accenturecio08/AutomationProcess_29697/_queries/query-edit/b6ab53a6-0320-4395-b854-4821b7d47889/'

        # ── Render summary section ─────────────────────────────────────────────
        html += f"""
                <!-- Defect Summary Section -->
                <div style="margin-bottom:24px;">
                    <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
                        <span style="font-size:22px;">&#128027;</span>
                        <h3 style="margin:0;font-size:16px;font-weight:700;color:#1e293b;">Defect Summary</h3>
                    </div>
                    <hr style="border:none;border-top:2px solid #667eea;margin-bottom:14px;">
                    <p style="margin:0 0 12px 0;font-size:12px;color:#334155;">
                        Defect Query:
                        <a href="{_defect_query_url}" target="_blank"
                           style="color:#667eea;font-weight:600;text-decoration:underline;">
                            Feb 21st Prod Sanity Defects - Boards
                        </a>
                    </p>
                    <!-- Summary Cards -->
                    <div style="display:flex;flex-wrap:wrap;gap:12px;margin-bottom:22px;">
                        <!-- Total Defects card -->
                        <div style="background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);color:white;
                                    border-radius:10px;padding:18px 24px;min-width:140px;text-align:center;
                                    box-shadow:0 2px 8px rgba(0,0,0,0.15);">
                            <div style="font-size:10px;font-weight:600;text-transform:uppercase;
                                        letter-spacing:0.5px;margin-bottom:8px;opacity:0.9;">TOTAL DEFECTS</div>
                            <div style="font-size:36px;font-weight:800;line-height:1;">{_total_count}</div>
                        </div>"""

        for _sk in _sev_order:
            _cnt = _sev_totals[_sk]
            _lbl = _sev_labels[_sk]
            _bg  = _sev_colors[_sk]
            _tx  = _sev_text[_sk]
            html += f"""
                        <!-- {_lbl} card -->
                        <div style="background:{_bg};color:{_tx};border-radius:10px;padding:18px 24px;
                                    min-width:120px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,0.15);">
                            <div style="font-size:36px;font-weight:800;line-height:1;margin-bottom:6px;">{_cnt}</div>
                            <div style="font-size:11px;font-weight:600;">{_lbl}</div>
                        </div>"""

        html += """
                    </div>

                    <!-- Total Defect Breakdown table -->
                    <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
                        <span style="font-size:18px;">&#128202;</span>
                        <h3 style="margin:0;font-size:15px;font-weight:700;color:#1e293b;">Total Defect Breakdown</h3>
                    </div>
                    <hr style="border:none;border-top:2px solid #667eea;margin-bottom:12px;">
                    <div style="overflow-x:auto;margin-bottom:22px;">
                        <table style="border-collapse:collapse;min-width:600px;font-size:12px;">
                            <thead>
                                <tr>
                                    <th style="background:#1e293b;color:white;padding:8px 14px;text-align:left;
                                               border:1px solid #334155;min-width:200px;"></th>
                                    <th style="background:#1e293b;color:white;padding:8px 14px;text-align:center;
                                               border:1px solid #334155;">Total</th>
                                    <th style="background:#8b0000;color:white;padding:8px 14px;text-align:center;
                                               border:1px solid #334155;">Critical</th>
                                    <th style="background:#e02020;color:white;padding:8px 14px;text-align:center;
                                               border:1px solid #334155;">High</th>
                                    <th style="background:#e6a817;color:#333;padding:8px 14px;text-align:center;
                                               border:1px solid #c49a0a;">Medium</th>
                                    <th style="background:#f5f0c0;color:#333;padding:8px 14px;text-align:center;
                                               border:1px solid #d4c87a;">Low</th>
                                </tr>
                            </thead>
                            <tbody>"""
        # Total bugs raised row
        html += f"""
                                <tr>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;font-weight:700;
                                               background:#f8f9fa;">Total bugs raised</td>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;text-align:center;
                                               background:#f8f9fa;font-weight:600;">{_total_raised}</td>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;text-align:center;
                                               background:#f8f9fa;">{_total_raised_by_sev['1 - Critical']}</td>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;text-align:center;
                                               background:#f8f9fa;">{_total_raised_by_sev['2 - High']}</td>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;text-align:center;
                                               background:#f8f9fa;">{_total_raised_by_sev['3 - Medium']}</td>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;text-align:center;
                                               background:#f8f9fa;">{_total_raised_by_sev['4 - Low']}</td>
                                </tr>"""

        _bd_row_styles = {
            'Active':                   ('color:#dc2626;', 'background:#fff5f5;'),
            'Fixed and Ready to Deploy': ('color:#ca8a04;', 'background:#fffbeb;'),
            'Under Testing':            ('color:#16a34a;', 'background:#f0fdf4;'),
            'Closed':                   ('color:#2563eb;', 'background:#eff6ff;'),
        }
        for _bs in _breakdown_states:
            _lc, _bg = _bd_row_styles.get(_bs, ('color:#333;', 'background:white;'))
            html += f"""
                                <tr>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;
                                               font-weight:600;{_lc}{_bg}">{_bs}</td>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;text-align:center;
                                               font-weight:600;{_lc}{_bg}">{_bd_total[_bs]}</td>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;text-align:center;
                                               {_lc}{_bg}">{_bd[_bs]['1 - Critical']}</td>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;text-align:center;
                                               {_lc}{_bg}">{_bd[_bs]['2 - High']}</td>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;text-align:center;
                                               {_lc}{_bg}">{_bd[_bs]['3 - Medium']}</td>
                                    <td style="padding:7px 14px;border:2px solid #94a3b8;text-align:center;
                                               {_lc}{_bg}">{_bd[_bs]['4 - Low']}</td>
                                </tr>"""

        html += """
                            </tbody>
                        </table>
                    </div>
                </div>
                <!-- End Defect Summary Section -->

                <div class="filter-section">
                    <div class="filter-group">
                        <label for="psdMpocFilter">&#128269; Filter by MPOC:</label>
                        <select id="psdMpocFilter" onchange="applyPSDFilters()">
                            <option value="all">-- All MPOCs --</option>
"""
        psd_mpoc_dict = {}
        psd_has_unassigned = False
        for _d in self.prod_sanity_defects:
            mpoc = _d.get('mpoc', '')
            if mpoc in ['', 'N/A', 'Unassigned']:
                psd_has_unassigned = True
            else:
                psd_mpoc_dict[mpoc.lower()] = mpoc
        if psd_has_unassigned:
            html += '                            <option value="Unassigned">Unassigned</option>\n'
        for mpoc in sorted(psd_mpoc_dict.values()):
            html += f'                            <option value="{mpoc}">{mpoc}</option>\n'

        psd_unique_states = sorted(set(_d['state'] for _d in self.prod_sanity_defects if _d.get('state')))
        html += """                        </select>
                    </div>
                    <div class="filter-group">
                        <label for="psdStateFilter">&#128269; Filter by State:</label>
                        <select id="psdStateFilter" onchange="applyPSDFilters()">
                            <option value="all">-- All States --</option>
"""
        for state in psd_unique_states:
            html += f'                            <option value="{state}">{state}</option>\n'
        html += """                        </select>
                    </div>
                    <button class="reset-btn" onclick="resetPSDFilters()">&#8635; Reset Filters</button>
                    <button class="save-btn" onclick="saveTabData('prodSanityDefectsTab', this)">&#128190; Save</button>
                    <button class="export-btn" onclick="exportCurrentTabToExcel()" title="Export current tab to Excel (.xlsx)">&#128229; Export Tab (.xlsx)</button>
                    <button class="export-all-btn" onclick="exportAllTabsToExcel()" title="Export all 6 tabs to Excel workbook (.xlsx)">&#128194; Export All Tabs (.xlsx)</button>
                    <div class="filter-info"><span id="psdVisibleCount">Showing all defects</span></div>
                </div>
                <div class="table-wrapper">
                    <table class="bug-list-table">
                        <thead>
                            <tr>
                                <th class="sno-col">S.No</th>
                                <th class="bug-mpoc-col">ExternalRef ID</th>
                                <th class="bug-id-col">ID</th>
                                <th class="bug-title-col">Title</th>
                                <th class="bug-state-col">State</th>
                                <th class="bug-defect-col">Defect Record</th>
                                <th class="bug-severity-col">Severity</th>
                                <th class="bug-node-col">Node Name</th>
                                <th class="bug-stage-col">Stage Found</th>
                            </tr>
                        </thead>
                        <tbody id="psdTableBody">
"""
        for sno_d, d in enumerate(self.prod_sanity_defects, 1):
            d_mpoc       = d.get('mpoc', 'Unassigned')
            d_id         = d.get('id', 'N/A')
            d_title      = d.get('title', 'N/A')
            d_state      = d.get('state', 'N/A')
            d_defect_rec = d.get('defect_record', 'N/A')
            d_severity   = d.get('severity', 'N/A')
            d_node       = d.get('node_name', 'N/A')
            d_stage      = d.get('stage_found', 'N/A')
            html += f"""
                            <tr data-mpoc="{d_mpoc}" data-state="{d_state}">
                                <td class="sno-col">{sno_d}</td>
                                <td class="bug-mpoc-col">{d_mpoc}</td>
                                <td class="bug-id-col">{d_id}</td>
                                <td class="bug-title-col">{d_title}</td>
                                <td class="bug-state-col">{d_state}</td>
                                <td class="bug-defect-col">{d_defect_rec}</td>
                                <td class="bug-severity-col">{d_severity}</td>
                                <td class="bug-node-col">{d_node}</td>
                                <td class="bug-stage-col">{d_stage}</td>
                            </tr>
"""
        if not self.prod_sanity_defects:
            html += '                            <tr><td colspan="9" style="text-align:center;color:#6b7280;padding:20px;">No defects found.</td></tr>\n'
        html += """                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- Tab 6 (last): Leadwise Prod Sanity Status -->
        <div id="overallStatusTab" class="tab-content">""" + overall_html + """</div>

        <div class="footer">
            <p><strong>&#169; 2025 Test Execution Report</strong> | """ + f"{ADO_CONFIG['organization']} / {ADO_CONFIG['project']}" + """</p>
        </div>
    </div>

    <script>
        // Store bug data for dynamic filtering
        const bugData = [];
        
        function applyUSFilters() {
            const leadFilter    = document.getElementById('usLeadFilter').value;
            const testerFilter  = document.getElementById('usTesterFilter').value;
            const moduleFilter  = document.getElementById('usModuleFilter').value;
            const outcomeFilter = document.getElementById('usOutcomeFilter').value;
            const statusFilter  = document.getElementById('usStatusFilter').value;
            const rows = document.querySelectorAll('#usSanityTableBody tr');

            let visibleCount = 0;
            let sno = 1;

            rows.forEach(row => {
                const isExtra = row.classList.contains('bug-extra-row');
                const lead    = row.getAttribute('data-lead');
                const tester  = row.getAttribute('data-tester');
                const module  = row.getAttribute('data-module');
                const outcome = row.getAttribute('data-outcome');
                const status  = row.getAttribute('data-status');

                const leadMatch    = leadFilter    === 'all' || lead   === leadFilter;
                const testerMatch  = testerFilter  === 'all' || tester === testerFilter;
                const moduleMatch  = moduleFilter  === 'all' || module === moduleFilter;
                const outcomeMatch = outcomeFilter === 'all' || (outcome || '').toLowerCase() === outcomeFilter.toLowerCase();
                const statusMatch  = statusFilter  === 'all' || status === statusFilter;

                if (leadMatch && testerMatch && moduleMatch && outcomeMatch && statusMatch) {
                    row.classList.remove('hidden');
                    if (!isExtra) {
                        const firstCell = row.querySelector('td:first-child');
                        if (firstCell) firstCell.textContent = sno++;
                        visibleCount++;
                    }
                } else {
                    row.classList.add('hidden');
                }
            });

            const totalRows = Array.from(rows).filter(r => !r.classList.contains('bug-extra-row')).length;
            const el = document.getElementById('usVisibleCount');
            if (el) el.textContent = 'Showing ' + visibleCount + ' of ' + totalRows + ' rows';
        }

        function resetUSFilters() {
            document.getElementById('usLeadFilter').value = 'all';
            document.getElementById('usTesterFilter').value = 'all';
            document.getElementById('usModuleFilter').value = 'all';
            document.getElementById('usOutcomeFilter').value = 'all';
            document.getElementById('usStatusFilter').value = 'all';
            applyUSFilters();
        }

        // Store original module options for each lead
        const leadModuleMap = {};

        // Initialize bug data from table
        function initializeBugData() {
            const rows = document.querySelectorAll('#bugListTableBody tr');
            rows.forEach(row => {
                bugData.push({
                    mpoc: row.getAttribute('data-mpoc'),
                    state: row.getAttribute('data-state')
                });
            });
        }
        
        // Update state checkboxes based on selected MPOC
        function updateStateOptions() {
            const mpocFilter = document.getElementById('bugMpocFilter').value;
            
            // Get unique states for selected MPOC
            let availableStates = new Set();
            
            if (mpocFilter === 'all') {
                // Show all states
                bugData.forEach(bug => availableStates.add(bug.state));
            } else {
                // Show only states for selected MPOC
                bugData.forEach(bug => {
                    if (bug.mpoc === mpocFilter) {
                        availableStates.add(bug.state);
                    }
                });
            }
            
            // Update state checkboxes
            const stateCheckboxes = document.querySelectorAll('.state-checkbox');
            stateCheckboxes.forEach(cb => {
                const stateValue = cb.value;
                const isAvailable = availableStates.has(stateValue);
                
                // Disable/enable checkbox based on availability
                cb.disabled = !isAvailable;
                
                // Uncheck disabled checkboxes
                if (!isAvailable) {
                    cb.checked = false;
                } else {
                    // Check available checkboxes
                    cb.checked = true;
                }
            });
            
            // Update "All States" checkbox
            const allCheckbox = document.querySelector('input[value="all"]');
            const enabledCheckboxes = document.querySelectorAll('.state-checkbox:not(:disabled)');
            allCheckbox.checked = enabledCheckboxes.length > 0;
            
            updateStateFilter();
        }
        
        // Initialize lead-module mapping from table data
        function initializeLeadModuleMap() {
            const rows = document.querySelectorAll('#reportTableBody tr:not(.grand-total-row)');
            rows.forEach(row => {
                const lead = row.getAttribute('data-lead');
                const module = row.getAttribute('data-module');
                
                if (!leadModuleMap[lead]) {
                    leadModuleMap[lead] = new Set();
                }
                leadModuleMap[lead].add(module);
            });
        }
        
        // Update module dropdown based on selected lead
        function updateModuleOptions() {
            const leadFilter = document.getElementById('leadFilter').value;
            const moduleFilter = document.getElementById('moduleFilter');
            const currentModule = moduleFilter.value;
            
            // Clear existing options except "All Modules"
            moduleFilter.innerHTML = '<option value="all">-- All Modules --</option>';
            
            if (leadFilter === 'all') {
                // Show all modules if "All Leads" is selected
                const allModules = new Set();
                Object.values(leadModuleMap).forEach(modules => {
                    modules.forEach(module => allModules.add(module));
                });
                
                Array.from(allModules).sort().forEach(module => {
                    const option = document.createElement('option');
                    option.value = module;
                    option.textContent = module;
                    moduleFilter.appendChild(option);
                });
            } else {
                // Show only modules for selected lead
                const modulesForLead = leadModuleMap[leadFilter] || new Set();
                Array.from(modulesForLead).sort().forEach(module => {
                    const option = document.createElement('option');
                    option.value = module;
                    option.textContent = module;
                    moduleFilter.appendChild(option);
                });
            }
            
            // Try to restore previous module selection if it's still available
            const availableOptions = Array.from(moduleFilter.options).map(opt => opt.value);
            if (availableOptions.includes(currentModule)) {
                moduleFilter.value = currentModule;
            } else {
                moduleFilter.value = 'all';
            }
            
            // Apply filters after updating options
            applyFilters();
        }
        
        function switchTab(tabName) {
            // Hide all tab content panels
            document.querySelectorAll('.tab-content').forEach(function(tab) {
                tab.classList.remove('active');
            });

            // Deactivate all tab buttons
            document.querySelectorAll('.tab').forEach(function(btn) {
                btn.classList.remove('active');
            });

            // Map tabName → content id + button onclick fragment
            const tabMap = {
                'overallProdSanity': 'overallProdSanityTab',
                'prodSanity':        'prodSanityTab',
                'prodUSSanity':      'prodUSSanityTab',
                'insprintStatus':    'insprintStatusTab',
                'readyForProdBug':   'readyForProdBugTab',
                'prodSanityDefects': 'prodSanityDefectsTab',
                'overallStatus':     'overallStatusTab',
            };

            const contentId = tabMap[tabName];
            if (!contentId) return;

            // Show the content panel
            const panel = document.getElementById(contentId);
            if (panel) panel.classList.add('active');

            // Activate the matching button by its onclick attribute
            document.querySelectorAll('.tab').forEach(function(btn) {
                const onclick = btn.getAttribute('onclick') || '';
                if (onclick.indexOf("'" + tabName + "'") !== -1) {
                    btn.classList.add('active');
                }
            });
        }
        function updateGrandTotal() {
            // Get all visible rows (not hidden and not grand total)
            const visibleRows = document.querySelectorAll('#reportTableBody tr:not(.grand-total-row):not(.hidden)');
            
            // Initialize totals
            let totalP1P2 = 0;
            let manualTotal = 0, manualPass = 0, manualFail = 0, manualBlock = 0, manualNA = 0, manualNotRun = 0;
            let autoTotal = 0, autoPass = 0, autoFail = 0, autoBlock = 0, autoNA = 0, autoNotRun = 0;
            
            // Sum up values from visible rows
            visibleRows.forEach(row => {
                const cells = row.querySelectorAll('td');
                // Column indices: 3=TotalP1P2, 4=ManTotal, 5=ManPass, 6=ManFail, 7=ManBlock, 8=ManNA, 9=ManNotRun
                // 12=AutoTotal, 13=AutoPass, 14=AutoFail, 15=AutoBlock, 16=AutoNA, 17=AutoNotRun
                totalP1P2 += parseInt(cells[3].textContent) || 0;
                manualTotal += parseInt(cells[4].textContent) || 0;
                manualPass += parseInt(cells[5].textContent) || 0;
                manualFail += parseInt(cells[6].textContent) || 0;
                manualBlock += parseInt(cells[7].textContent) || 0;
                manualNA += parseInt(cells[8].textContent) || 0;
                manualNotRun += parseInt(cells[9].textContent) || 0;
                autoTotal += parseInt(cells[12].textContent) || 0;
                autoPass += parseInt(cells[13].textContent) || 0;
                autoFail += parseInt(cells[14].textContent) || 0;
                autoBlock += parseInt(cells[15].textContent) || 0;
                autoNA += parseInt(cells[16].textContent) || 0;
                autoNotRun += parseInt(cells[17].textContent) || 0;
            });
            
            // Calculate percentages for Manual
            const manualDenomPass = manualPass + manualFail + manualBlock;
            const manualPassPct = manualDenomPass > 0 ? (manualPass / manualDenomPass * 100) : 0;
            const manualDenomExec = manualTotal - manualNA;
            const manualExecPct = manualDenomExec > 0 ? ((manualPass + manualFail + manualBlock) / manualDenomExec * 100) : 0;
            
            // Calculate percentages for Automation
            const autoDenomPass = autoPass + autoFail + autoBlock;
            const autoPassPct = autoDenomPass > 0 ? (autoPass / autoDenomPass * 100) : 0;
            const autoDenomExec = autoTotal - autoNA;
            const autoExecPct = autoDenomExec > 0 ? ((autoPass + autoFail + autoBlock) / autoDenomExec * 100) : 0;
            
            // Update grand total row
            const grandTotalRow = document.getElementById('grandTotalRow');
            if (grandTotalRow) {
                const cells = grandTotalRow.querySelectorAll('td');
                cells[1].textContent = totalP1P2;
                cells[2].textContent = manualTotal;
                cells[3].textContent = manualPass;
                cells[4].textContent = manualFail;
                cells[5].textContent = manualBlock;
                cells[6].textContent = manualNA;
                cells[7].textContent = manualNotRun;
                cells[8].textContent = manualExecPct.toFixed(2) + '%';
                cells[9].textContent = manualPassPct.toFixed(2) + '%';
                cells[10].textContent = autoTotal;
                cells[11].textContent = autoPass;
                cells[12].textContent = autoFail;
                cells[13].textContent = autoBlock;
                cells[14].textContent = autoNA;
                cells[15].textContent = autoNotRun;
                cells[16].textContent = autoExecPct.toFixed(2) + '%';
                cells[17].textContent = autoPassPct.toFixed(2) + '%';
            }
        }
        
        function applyFilters() {
            const leadFilter = document.getElementById('leadFilter').value;
            const moduleFilter = document.getElementById('moduleFilter').value;
            const rows = document.querySelectorAll('#reportTableBody tr:not(.grand-total-row)');
            
            let visibleCount = 0;
            
            rows.forEach(row => {
                const lead = row.getAttribute('data-lead');
                const module = row.getAttribute('data-module');
                
                const leadMatch = leadFilter === 'all' || lead === leadFilter;
                const moduleMatch = moduleFilter === 'all' || module === moduleFilter;
                
                if (leadMatch && moduleMatch) {
                    row.classList.remove('hidden');
                    visibleCount++;
                } else {
                    row.classList.add('hidden');
                }
            });
            
            // Update visible count
            const totalRows = rows.length;
            document.getElementById('visibleCount').textContent = 
                `Showing ${visibleCount} of ${totalRows} rows`;
            
            // Update grand total based on visible rows
            updateGrandTotal();
        }
        
        function resetFilters() {
            document.getElementById('leadFilter').value = 'all';
            document.getElementById('moduleFilter').value = 'all';
            updateModuleOptions();
        }

        // ── Prod Sanity Scenarios Tab Filter Functions ──────────────────────
        function applyPSFilters() {
            const leadFilter    = document.getElementById('psLeadFilter').value;
            const moduleFilter  = document.getElementById('psModuleFilter').value;
            const statusFilter  = document.getElementById('psStatusFilter').value;
            const testerFilter  = document.getElementById('psTesterFilter').value;
            const outcomeFilter = document.getElementById('psOutcomeFilter').value;
            const rows = document.querySelectorAll('#psScenariosBody tr');

            let visible = 0;
            rows.forEach(row => {
                const isExtra = row.classList.contains('bug-extra-row');
                const leadMatch    = leadFilter    === 'all' || row.getAttribute('data-lead')    === leadFilter;
                const moduleMatch  = moduleFilter  === 'all' || row.getAttribute('data-module')  === moduleFilter;
                const statusMatch  = statusFilter  === 'all' || row.getAttribute('data-status')  === statusFilter;
                const testerMatch  = testerFilter  === 'all' || row.getAttribute('data-tester')  === testerFilter;
                const outcomeMatch = outcomeFilter === 'all' || row.getAttribute('data-outcome') === outcomeFilter;
                if (leadMatch && moduleMatch && statusMatch && testerMatch && outcomeMatch) {
                    row.classList.remove('hidden');
                    if (!isExtra) visible++;
                } else {
                    row.classList.add('hidden');
                }
            });

            const total = Array.from(rows).filter(r => !r.classList.contains('bug-extra-row')).length;
            const el = document.getElementById('psVisibleCount');
            if (el) el.textContent = 'Showing ' + visible + ' of ' + total + ' rows';

            // Re-number visible rows
            let sno = 1;
            rows.forEach(row => {
                if (row.classList.contains('bug-extra-row')) return;
                if (!row.classList.contains('hidden')) {
                    const firstCell = row.querySelector('td:first-child');
                    if (firstCell) firstCell.textContent = sno++;
                }
            });
        }

        function resetPSFilters() {
            document.getElementById('psLeadFilter').value    = 'all';
            document.getElementById('psModuleFilter').value  = 'all';
            document.getElementById('psStatusFilter').value  = 'all';
            document.getElementById('psTesterFilter').value  = 'all';
            document.getElementById('psOutcomeFilter').value = 'all';
            applyPSFilters();
        }
        // ────────────────────────────────────────────────────────────────────

        // Insprint Tab Filter Functions
        function applyFiltersInsprint() {
            const leadFilter    = document.getElementById('leadFilterInsprint').value;
            const testerFilter  = document.getElementById('insprintTesterFilter').value;
            const moduleFilter  = document.getElementById('insprintModuleFilter').value;
            const outcomeFilter = document.getElementById('insprintOutcomeFilter').value;
            const statusFilter  = document.getElementById('insprintStatusFilter').value;
            const rows = document.querySelectorAll('#reportTableBodyInsprint tr');

            let visible = 0;
            rows.forEach(row => {
                const isExtra = row.classList.contains('bug-extra-row');
                const leadMatch    = leadFilter    === 'all' || row.getAttribute('data-lead')    === leadFilter;
                const testerMatch  = testerFilter  === 'all' || row.getAttribute('data-tester')  === testerFilter;
                const moduleMatch  = moduleFilter  === 'all' || row.getAttribute('data-module')  === moduleFilter;
                const outcomeMatch = outcomeFilter === 'all' || (row.getAttribute('data-outcome') || '').toLowerCase() === outcomeFilter.toLowerCase();
                const statusMatch  = statusFilter  === 'all' || row.getAttribute('data-status')  === statusFilter;
                if (leadMatch && testerMatch && moduleMatch && outcomeMatch && statusMatch) {
                    row.classList.remove('hidden');
                    if (!isExtra) visible++;
                } else {
                    row.classList.add('hidden');
                }
            });

            const total = Array.from(rows).filter(r => !r.classList.contains('bug-extra-row')).length;
            const el = document.getElementById('visibleCountInsprint');
            if (el) el.textContent = 'Showing ' + visible + ' of ' + total + ' rows';

            let sno = 1;
            rows.forEach(row => {
                if (row.classList.contains('bug-extra-row')) return;
                if (!row.classList.contains('hidden')) {
                    const firstCell = row.querySelector('td:first-child');
                    if (firstCell) firstCell.textContent = sno++;
                }
            });
        }

        function resetFiltersInsprint() {
            document.getElementById('leadFilterInsprint').value       = 'all';
            document.getElementById('insprintTesterFilter').value     = 'all';
            document.getElementById('insprintModuleFilter').value     = 'all';
            document.getElementById('insprintOutcomeFilter').value    = 'all';
            document.getElementById('insprintStatusFilter').value     = 'all';
            applyFiltersInsprint();
        }
        
        // Bug List Filter Functions
        function toggleStateDropdown() {
            const menu = document.getElementById('stateDropdownMenu');
            const button = document.getElementById('stateDropdownToggle');
            const isShowing = menu.classList.contains('show');
            
            if (isShowing) {
                menu.classList.remove('show');
            } else {
                // Reset to default position first
                menu.style.top = '100%';
                menu.style.bottom = 'auto';
                menu.style.marginTop = '2px';
                menu.style.marginBottom = '0';
                
                menu.classList.add('show');
                
                // Adjust dropdown position if it goes off-screen
                setTimeout(() => {
                    const buttonRect = button.getBoundingClientRect();
                    const menuRect = menu.getBoundingClientRect();
                    const viewportHeight = window.innerHeight;
                    
                    // Calculate space below and above
                    const spaceBelow = viewportHeight - buttonRect.bottom;
                    const spaceAbove = buttonRect.top;
                    const menuHeight = menuRect.height;
                    
                    // If not enough space below but enough space above, flip upward
                    if (spaceBelow < menuHeight + 20 && spaceAbove > menuHeight + 20) {
                        menu.style.top = 'auto';
                        menu.style.bottom = '100%';
                        menu.style.marginBottom = '2px';
                        menu.style.marginTop = '0';
                    }
                }, 10);
            }
            
            // Close dropdown when clicking outside
            document.addEventListener('click', function closeDropdown(e) {
                if (!e.target.closest('.custom-dropdown')) {
                    menu.classList.remove('show');
                    // Reset position
                    menu.style.top = '100%';
                    menu.style.bottom = 'auto';
                    menu.style.marginTop = '2px';
                    menu.style.marginBottom = '0';
                    document.removeEventListener('click', closeDropdown);
                }
            });
        }
        
        function toggleAllStates(checkbox) {
            const stateCheckboxes = document.querySelectorAll('.state-checkbox');
            stateCheckboxes.forEach(cb => {
                cb.checked = checkbox.checked;
            });
            updateStateFilter();
        }
        
        function updateStateFilter() {
            const allCheckbox = document.querySelector('input[value="all"]');
            const stateCheckboxes = document.querySelectorAll('.state-checkbox');
            const checkedCount = Array.from(stateCheckboxes).filter(cb => cb.checked).length;
            
            // Update "All States" checkbox
            allCheckbox.checked = checkedCount === stateCheckboxes.length;
            
            // Update dropdown button text
            const dropdownToggle = document.getElementById('stateDropdownToggle');
            if (checkedCount === 0) {
                dropdownToggle.textContent = 'No States Selected ▼';
            } else if (checkedCount === stateCheckboxes.length) {
                dropdownToggle.textContent = 'All States ▼';
            } else {
                dropdownToggle.textContent = `${checkedCount} State(s) Selected ▼`;
            }
            
            applyBugFilters();
        }
        
        function applyBugFilters() {
            const mpocFilter   = document.getElementById('bugMpocFilter').value;
            const leadFilter   = document.getElementById('bugLeadFilter').value;
            const psFilter     = document.getElementById('bugPsStatusFilter').value;
            const stateCheckboxes = document.querySelectorAll('.state-checkbox:checked');
            const selectedStates  = Array.from(stateCheckboxes).map(cb => cb.value);
            const rows = document.querySelectorAll('#bugListTableBody tr');

            let visibleCount = 0;

            rows.forEach(row => {
                const mpoc     = row.getAttribute('data-mpoc');
                const state    = row.getAttribute('data-state');
                const lead     = row.getAttribute('data-lead');
                // Read current dropdown value for prod sanity status
                const psSelect = row.querySelector('td:last-child select');
                const psStatus = psSelect ? psSelect.value : '';

                // Keep data-ps-status in sync for future filter reads
                row.setAttribute('data-ps-status', psStatus);

                const mpocMatch  = mpocFilter === 'all' || mpoc === mpocFilter;
                const leadMatch  = leadFilter === 'all' || lead === leadFilter;
                const stateMatch = selectedStates.length === 0 || selectedStates.includes(state);
                const psMatch    = psFilter === 'all' || psStatus === psFilter;

                if (mpocMatch && leadMatch && stateMatch && psMatch) {
                    row.classList.remove('hidden');
                    visibleCount++;
                } else {
                    row.classList.add('hidden');
                }
            });

            const totalRows = rows.length;
            document.getElementById('bugVisibleCount').textContent =
                `Showing ${visibleCount} of ${totalRows} bugs`;
        }

        function resetBugFilters() {
            document.getElementById('bugMpocFilter').value   = 'all';
            document.getElementById('bugLeadFilter').value   = 'all';
            document.getElementById('bugPsStatusFilter').value = 'all';

            // Re-enable and check all state checkboxes
            const allCheckbox = document.querySelector('#stateDropdownMenu input[value="all"]');
            const stateCheckboxes = document.querySelectorAll('.state-checkbox');
            if (allCheckbox) { allCheckbox.checked = true; allCheckbox.disabled = false; }
            stateCheckboxes.forEach(cb => {
                cb.checked = true;
                cb.disabled = false;
                cb.parentElement.style.opacity = '1';
            });

            updateStateFilter();
        }

        // ── Prod Sanity Defects tab filters ──────────────────────────────────
        function applyPSDFilters() {
            const mpocFilter  = document.getElementById('psdMpocFilter').value;
            const stateFilter = document.getElementById('psdStateFilter').value;
            const rows = document.querySelectorAll('#psdTableBody tr');
            let visible = 0;
            rows.forEach(row => {
                const mpocMatch  = mpocFilter  === 'all' || row.getAttribute('data-mpoc')  === mpocFilter;
                const stateMatch = stateFilter === 'all' || row.getAttribute('data-state') === stateFilter;
                if (mpocMatch && stateMatch) {
                    row.classList.remove('hidden');
                    visible++;
                } else {
                    row.classList.add('hidden');
                }
            });
            const el = document.getElementById('psdVisibleCount');
            if (el) el.textContent = 'Showing ' + visible + ' of ' + rows.length + ' defects';
            let sno = 1;
            rows.forEach(row => {
                if (!row.classList.contains('hidden')) {
                    const fc = row.querySelector('td:first-child');
                    if (fc) fc.textContent = sno++;
                }
            });
        }

        function resetPSDFilters() {
            document.getElementById('psdMpocFilter').value  = 'all';
            document.getElementById('psdStateFilter').value = 'all';
            applyPSDFilters();
        }

        // ============================================================
        // SAVE / LOAD tab data using localStorage
        // Persists all <select> and <input type="text"> values inside
        // each tab content div, keyed by element index within that tab.
        // ============================================================
        function saveTabData(tabId, btn) {
            const tab = document.getElementById(tabId);
            if (!tab) return;
            const data = {};
            tab.querySelectorAll('select, input[type="text"]').forEach(function(el, i) {
                data[i] = el.value;
            });
            try {
                localStorage.setItem('tabSave_' + tabId, JSON.stringify(data));
                if (btn) {
                    const orig = btn.innerHTML;
                    btn.innerHTML = '&#10003; Saved';
                    btn.classList.add('saved');
                    setTimeout(function() {
                        btn.innerHTML = orig;
                        btn.classList.remove('saved');
                    }, 1500);
                }
            } catch(e) {
                alert('Save failed: ' + e.message);
            }
        }

        function loadTabData(tabId) {
            const tab = document.getElementById(tabId);
            if (!tab) return;
            const raw = localStorage.getItem('tabSave_' + tabId);
            if (!raw) return;
            try {
                const data = JSON.parse(raw);
                tab.querySelectorAll('select, input[type="text"]').forEach(function(el, i) {
                    if (data[i] !== undefined) el.value = data[i];
                });
            } catch(e) {}
        }

        // ============================================================
        // SAVE / LOAD tab data using localStorage
        // Persists all <select> and <input type="text"> values inside
        // each tab content div, keyed by element index within that tab.
        // ============================================================
        function saveTabData(tabId, btn) {
            const tab = document.getElementById(tabId);
            if (!tab) return;
            const data = {};
            tab.querySelectorAll('select, input[type="text"]').forEach(function(el, i) {
                data[i] = el.value;
            });
            try {
                localStorage.setItem('tabSave_' + tabId, JSON.stringify(data));
                if (btn) {
                    const orig = btn.innerHTML;
                    btn.innerHTML = '&#10003; Saved';
                    btn.classList.add('saved');
                    setTimeout(function() {
                        btn.innerHTML = orig;
                        btn.classList.remove('saved');
                    }, 1500);
                }
            } catch(e) {
                alert('Save failed: ' + e.message);
            }
        }

        function loadTabData(tabId) {
            const tab = document.getElementById(tabId);
            if (!tab) return;
            const raw = localStorage.getItem('tabSave_' + tabId);
            if (!raw) return;
            try {
                const data = JSON.parse(raw);
                tab.querySelectorAll('select, input[type="text"]').forEach(function(el, i) {
                    if (data[i] !== undefined) el.value = data[i];
                });
            } catch(e) {}
        }

        // ============================================================
        // REFRESH — Overall Prod Sanity Status pivot tables
        // Reads current select values from all three data-tab bodies and
        // rebuilds the three pivot tables in Tab 0.
        // ============================================================
        const PIVOT_STATUSES = [
            'Working fine', 'Working fine with active bug', 'In Progress', 'Failed',
            'Blocked', 'Cannot be validated', 'Can be validated on Monday',
            'Taken care by BPMS, IMS test team',
            'Story not deployed to prod as part of this release',
            'Toggle off/Code reversal validated', 'Yet to start'
        ];

        function _buildPivotFromBody(bodySelector) {
            const pivot = {};
            document.querySelectorAll(bodySelector + ' tr').forEach(function(row) {
                const lead = row.getAttribute('data-lead') || 'Unassigned';
                // First <select> in the row holds Prod Sanity Status
                const sel = row.querySelector('select');
                const status = (sel ? sel.value : '') || 'Yet to start';
                if (!pivot[lead]) pivot[lead] = {};
                pivot[lead][status] = (pivot[lead][status] || 0) + 1;
            });
            return pivot;
        }

        function _renderPivotInto(containerId, title, pivot, statusCols, accentColor) {
            const container = document.getElementById(containerId);
            if (!container) return;
            const accent = accentColor || '#667eea';
            const allCols = (statusCols && statusCols.length) ? statusCols : PIVOT_STATUSES;
            const leads = Object.keys(pivot).sort();
            if (leads.length === 0) return;
            const used = allCols.filter(function(s) {
                return leads.some(function(l) { return (pivot[l][s] || 0) > 0; });
            });
            const cols = used.length > 0 ? used : allCols;
            const grand = {};
            cols.forEach(function(s) {
                grand[s] = leads.reduce(function(acc, l) { return acc + (pivot[l][s] || 0); }, 0);
            });
            const grandTotal = Object.values(grand).reduce(function(a, b) { return a + b; }, 0);

            let html = '<div style="background:white;border-radius:10px;box-shadow:0 2px 12px rgba(0,0,0,0.09);overflow:hidden;">';
            html += '<div style="background:linear-gradient(135deg,' + accent + ' 0%,' + accent + 'cc 100%);padding:10px 16px;display:flex;align-items:center;gap:8px;">';
            html += '<span style="color:white;font-size:12px;font-weight:700;">' + title + '</span>';
            html += '<span style="margin-left:auto;background:rgba(255,255,255,0.25);color:white;font-size:10px;font-weight:600;padding:2px 8px;border-radius:10px;">Total: ' + grandTotal + '</span>';
            html += '</div><div class="table-wrapper" style="overflow-x:auto;padding:12px;">';
            html += '<table style="border-collapse:collapse;font-size:10px;width:auto;min-width:400px;"><thead><tr>';
            html += '<th style="background:#1e293b;color:white;padding:7px 12px;border:1px solid #334155;text-align:left;white-space:nowrap;">Test Lead</th>';
            cols.forEach(function(s) {
                html += '<th style="background:#1e293b;color:white;padding:7px 8px;border:1px solid #334155;text-align:center;white-space:nowrap;font-size:9px;">' + s + '</th>';
            });
            html += '<th style="background:#1e293b;color:white;padding:7px 12px;border:1px solid #334155;text-align:center;white-space:nowrap;">Total</th></tr></thead><tbody>';
            leads.forEach(function(lead, idx) {
                const rowTotal = cols.reduce(function(acc, s) { return acc + ((pivot[lead] || {})[s] || 0); }, 0);
                const bg = idx % 2 === 0 ? 'white' : '#f8faff';
                html += '<tr style="background:' + bg + ';"><td style="padding:6px 12px;border:2px solid #64748b;font-weight:600;text-align:left;white-space:nowrap;background:#f0f4ff;color:#3730a3;">' + lead + '</td>';
                cols.forEach(function(s) {
                    const cnt = (pivot[lead] || {})[s] || 0;
                    const st = cnt > 0 ? 'color:#16a34a;font-weight:600;' : 'color:#94a3b8;';
                    html += '<td style="padding:6px 8px;border:2px solid #64748b;text-align:center;' + st + '">' + (cnt > 0 ? cnt : '\u2014') + '</td>';
                });
                html += '<td style="padding:6px 12px;border:2px solid #64748b;text-align:center;font-weight:700;background:#eff6ff;color:#1d4ed8;">' + rowTotal + '</td></tr>';
            });
            html += '<tr style="background:linear-gradient(90deg,' + accent + '33 0%,' + accent + '11 100%);">';
            html += '<td style="padding:6px 12px;border:2px solid #64748b;font-weight:700;color:' + accent + ';text-align:left;">Grand Total</td>';
            cols.forEach(function(s) {
                html += '<td style="padding:6px 8px;border:2px solid #64748b;font-weight:700;color:' + accent + ';text-align:center;">' + (grand[s] > 0 ? grand[s] : '\u2014') + '</td>';
            });
            html += '<td style="padding:6px 12px;border:2px solid #64748b;font-weight:700;color:' + accent + ';text-align:center;">' + grandTotal + '</td></tr>';
            html += '</tbody></table></div></div>';
            container.innerHTML = html;
        }

        function refreshOverallStatus() {
            _renderPivotInto('pivotProdSanity',      'Prod Sanity Scenarios \u2014 Status by Test Lead',  _buildPivotFromBody('#psScenariosBody'),         PIVOT_STATUSES, '#16a34a');
            _renderPivotInto('pivotProdUSSanity',    'Prod US Sanity \u2014 Status by Test Lead',          _buildPivotFromBody('#usSanityTableBody'),       PIVOT_STATUSES, '#1d4ed8');
            _renderPivotInto('pivotInsprintUS',      'Insprint US Prod Status \u2014 Status by Test Lead', _buildPivotFromBody('#reportTableBodyInsprint'), PIVOT_STATUSES, '#7e22ce');
            _renderPivotInto('pivotReadyForProdBug', 'Ready for Prod Bug \u2014 Severity by PT Lead',      _buildBugPivotFromBody(),                        BUG_SEV_COLS,   '#dc2626');
            const btn = document.querySelector('#overallStatusTab .refresh-btn');
            if (btn) {
                const orig = btn.innerHTML;
                btn.innerHTML = '&#10003; Refreshed';
                setTimeout(function() { btn.innerHTML = orig; }, 1500);
            }
        }

        // ============================================================
        // EXCEL EXPORT — Styled with ExcelJS for leadership sharing
        // Professional formatting: headers, traffic-light status colors,
        // borders, auto-column widths, frozen header rows.
        // ============================================================

        // ── Color palette (ARGB format for ExcelJS) ──────────────────
        const XL = {
            headerBg:      'FF3730A3',  // Deep indigo   (matches report gradient)
            headerFg:      'FFFFFFFF',  // White
            bugHeaderBg:   'FFC1121F',  // Red  (Bug Details span)
            bugHeaderFg:   'FFFFFFFF',
            passedBg:      'FFD6EED8',  passedFg:      'FF14532D',
            failedBg:      'FFF4C6C6',  failedFg:      'FF7F1D1D',
            blockedBg:     'FFFFE4CC',  blockedFg:     'FF7D3C00',
            notRunBg:      'FFE5E7EB',  notRunFg:      'FF374151',
            naBg:          'FFF0F0F0',  naFg:          'FF6B7280',
            inProgressBg:  'FFFEF3C7',  inProgressFg:  'FF78350F',
            grandTotalBg:  'FF1E293B',  grandTotalFg:  'FFFFFFFF',
            totalRowBg:    'FFE0E7FF',  totalRowFg:    'FF1E40AF',
            rowNormal:     'FFFFFFFF',
            rowAlt:        'FFF5F3FF',  // Very light lavender
            border:        'FFCBD5E1',
            bugIdFg:       'FFDC2626',
        };

        function _xlFill(argb) {
            return { type: 'pattern', pattern: 'solid', fgColor: { argb: argb } };
        }
        function _xlFont(bold, argb, size) {
            return { name: 'Calibri', size: size || 10, bold: !!bold, color: { argb: argb || 'FF000000' } };
        }
        function _xlBorder() {
            const c = { argb: 'FF475569' };
            return { top:{style:'thin',color:c}, left:{style:'thin',color:c},
                     bottom:{style:'thin',color:c}, right:{style:'thin',color:c} };
        }
        function _xlStatusStyle(text) {
            const t = (text || '').trim().toLowerCase();
            if (t === 'passed' || t === 'pass')
                return { fill: XL.passedBg, font: XL.passedFg };
            if (t === 'failed' || t === 'fail')
                return { fill: XL.failedBg, font: XL.failedFg };
            if (t === 'blocked' || t === 'block')
                return { fill: XL.blockedBg, font: XL.blockedFg };
            if (t === 'not run' || t === 'notrun')
                return { fill: XL.notRunBg, font: XL.notRunFg };
            if (t === 'not applicable' || t === 'na' || t === 'n/a')
                return { fill: XL.naBg, font: XL.naFg };
            if (t.includes('progress'))
                return { fill: XL.inProgressBg, font: XL.inProgressFg };
            return null;
        }

        /**
         * Parse an HTML table (already cloned/cleaned) into a flat list of cell
         * descriptors, correctly handling rowspan and colspan via a grid map.
         */
        function _parseTable(table) {
            const rows = Array.from(table.querySelectorAll('tr'));
            const occupied = {};        // 'r,c' -> true
            const result   = [];

            function isOcc(r, c) { return !!occupied[r + ',' + c]; }
            function mark(r, c)  { occupied[r + ',' + c] = true; }
            function nextFree(r, startC) {
                let c = startC;
                while (isOcc(r, c)) c++;
                return c;
            }

            rows.forEach(function(tr, ri) {
                const cellEls = Array.from(tr.querySelectorAll('th,td'));
                let col = 1;
                cellEls.forEach(function(el) {
                    col = nextFree(ri + 1, col);
                    const rs = parseInt(el.getAttribute('rowspan') || '1');
                    const cs = parseInt(el.getAttribute('colspan') || '1');
                    const text = (el.textContent || '').trim().replace(/\s+/g, ' ');
                    result.push({ row: ri + 1, col: col, rs: rs, cs: cs,
                                  isTh: el.tagName === 'TH', text: text });
                    for (let dr = 0; dr < rs; dr++)
                        for (let dc = 0; dc < cs; dc++)
                            mark(ri + 1 + dr, col + dc);
                    col += cs;
                });
            });
            return result;
        }

        /** Count consecutive header rows (rows that contain at least one <th>). */
        function _headerRowCount(cells) {
            const thRows = new Set(cells.filter(function(c) { return c.isTh; }).map(function(c) { return c.row; }));
            let n = 0;
            for (let r = 1; thRows.has(r); r++) n++;
            return n;
        }

        /** Clean and clone a table, replacing form controls with their values. */
        function _cleanTable(table) {
            const clone = table.cloneNode(true);
            clone.querySelectorAll('tr.hidden').forEach(function(r) { r.remove(); });
            clone.querySelectorAll('select').forEach(function(sel) {
                const td = sel.closest('td');
                if (td) td.textContent = sel.options[sel.selectedIndex]
                    ? sel.options[sel.selectedIndex].text : (sel.value || '');
            });
            clone.querySelectorAll('input[type="text"]').forEach(function(inp) {
                const td = inp.closest('td');
                if (td) td.textContent = inp.value || '';
            });
            return clone;
        }

        /** Build one styled worksheet from an HTML table element. */
        async function _tableToSheet(wb, sheetName, table) {
            const clean = _cleanTable(table);
            const cells = _parseTable(clean);
            if (cells.length === 0) return;

            const hCount = _headerRowCount(cells);

            // Identify "Title"/"Scenarios" columns — single-line display (no wrap)
            const titleCols = new Set();
            cells.forEach(function(c) {
                if (c.isTh) {
                    const t = (c.text || '').trim().toLowerCase();
                    if (t === 'title' || t === 'scenarios') titleCols.add(c.col);
                }
            });

            const ws = wb.addWorksheet(sheetName, {
                views: [{ state: 'frozen', ySplit: hCount }],
                properties: { tabColor: { argb: XL.headerBg } },
            });

            // Track data-row parity for alternating row colors
            const rowParity = {};
            let dataRowN = 0;

            // Max dimensions for column width pass
            const colMaxLen = {};

            cells.forEach(function(cell) {
                const el = ws.getCell(cell.row, cell.col);
                el.value = cell.text || null;

                const textLow = (cell.text || '').toLowerCase();
                let fillArgb, fontArgb, bold = false, hAlign = 'left';

                if (cell.isTh) {
                    // Header cell
                    if (textLow === 'bug details') {
                        fillArgb = XL.bugHeaderBg; fontArgb = XL.bugHeaderFg;
                    } else if (textLow.includes('grand total')) {
                        fillArgb = XL.grandTotalBg; fontArgb = XL.grandTotalFg;
                    } else {
                        fillArgb = XL.headerBg; fontArgb = XL.headerFg;
                    }
                    bold   = true;
                    hAlign = 'center';
                } else {
                    // Data cell — parity for alternating rows
                    if (rowParity[cell.row] === undefined) {
                        dataRowN++;
                        rowParity[cell.row] = dataRowN;
                    }
                    const isAlt = rowParity[cell.row] % 2 === 0;

                    if (textLow.includes('grand total') || textLow === 'grand total') {
                        fillArgb = XL.grandTotalBg; fontArgb = XL.grandTotalFg;
                        bold = true; hAlign = 'center';
                    } else {
                        const ss = _xlStatusStyle(cell.text);
                        if (ss) {
                            fillArgb = ss.fill; fontArgb = ss.font;
                            bold = true; hAlign = 'center';
                        } else {
                            fillArgb = isAlt ? XL.rowAlt : XL.rowNormal;
                            fontArgb = 'FF111827';
                            // Right-align numbers
                            if (cell.text && !isNaN(cell.text.replace('%', '').trim()) && cell.text.trim() !== '') {
                                hAlign = 'center';
                            }
                        }
                    }
                }

                el.font      = _xlFont(bold, fontArgb || 'FF000000', 10);
                el.fill      = _xlFill(fillArgb || XL.rowNormal);
                el.border    = _xlBorder();
                el.alignment = { vertical: 'middle', horizontal: hAlign, wrapText: !titleCols.has(cell.col) };

                // Merge
                if (cell.rs > 1 || cell.cs > 1) {
                    ws.mergeCells(cell.row, cell.col,
                                  cell.row + cell.rs - 1, cell.col + cell.cs - 1);
                }

                // Track max length for column width
                if (cell.cs === 1) {
                    const len = (cell.text || '').length;
                    const capW = titleCols.has(cell.col) ? 45 : 55;
                    const minW = titleCols.has(cell.col) ? 30 : 8;
                    colMaxLen[cell.col] = Math.max(colMaxLen[cell.col] || minW,
                                                   Math.min(len + 4, capW));
                }
            });

            // Column widths
            const maxCol = cells.reduce(function(m, c) {
                return Math.max(m, c.col + c.cs - 1);
            }, 0);
            for (let c = 1; c <= maxCol; c++) {
                ws.getColumn(c).width = colMaxLen[c] || 14;
            }

            // Row heights
            const maxRow = cells.reduce(function(m, c) {
                return Math.max(m, c.row + c.rs - 1);
            }, 0);
            for (let r = 1; r <= maxRow; r++) {
                ws.getRow(r).height = r <= hCount ? 22 : 18;
            }
        }

        /**
         * Write multiple tables into ONE worksheet, each preceded by a bold section heading.
         * Used for Overall Production Sanity Status and Leadwise Prod Sanity Status tabs
         * so they each export as a single sheet rather than multiple sheets.
         */
        async function _combinedTableSheet(wb, sheetName, tables, sectionNames) {
            const ws = wb.addWorksheet(sheetName, {
                properties: { tabColor: { argb: XL.headerBg } }
            });
            let currentRow = 1;
            const colMaxLen = {};

            for (let i = 0; i < tables.length; i++) {
                const table   = tables[i];
                const heading = sectionNames[i] || ('Section ' + (i + 1));

                // ── Section heading row ──────────────────────────────────────
                const hCell = ws.getCell(currentRow, 1);
                hCell.value = heading;
                hCell.font  = { name: 'Calibri', size: 12, bold: true, color: { argb: 'FFFFFFFF' } };
                hCell.fill  = _xlFill('FF1E3A8A');
                hCell.border = _xlBorder();
                hCell.alignment = { vertical: 'middle', horizontal: 'left' };
                ws.getRow(currentRow).height = 26;
                ws.mergeCells(currentRow, 1, currentRow, 10);
                currentRow++;

                // ── Parse & style the table ──────────────────────────────────
                const clean = _cleanTable(table);
                const cells = _parseTable(clean);
                if (cells.length === 0) { currentRow++; continue; }

                const hCount = _headerRowCount(cells);
                const titleCols = new Set();
                cells.forEach(function(c) {
                    if (c.isTh) {
                        const t = (c.text || '').trim().toLowerCase();
                        if (t === 'title' || t === 'scenarios') titleCols.add(c.col);
                    }
                });

                const rowParity = {};
                let dataRowN = 0;
                const maxRow = cells.reduce(function(m, c) { return Math.max(m, c.row + c.rs - 1); }, 0);

                cells.forEach(function(cell) {
                    const absRow = cell.row + currentRow - 1;
                    const el = ws.getCell(absRow, cell.col);
                    el.value = cell.text || null;

                    const textLow = (cell.text || '').toLowerCase();
                    let fillArgb, fontArgb, bold = false, hAlign = 'left';

                    if (cell.isTh) {
                        if (textLow === 'bug details') {
                            fillArgb = XL.bugHeaderBg; fontArgb = XL.bugHeaderFg;
                        } else if (textLow.includes('grand total')) {
                            fillArgb = XL.grandTotalBg; fontArgb = XL.grandTotalFg;
                        } else {
                            fillArgb = XL.headerBg; fontArgb = XL.headerFg;
                        }
                        bold = true; hAlign = 'center';
                    } else {
                        if (rowParity[cell.row] === undefined) { dataRowN++; rowParity[cell.row] = dataRowN; }
                        const isAlt = rowParity[cell.row] % 2 === 0;
                        if (textLow.includes('grand total') || textLow === 'grand total') {
                            fillArgb = XL.grandTotalBg; fontArgb = XL.grandTotalFg;
                            bold = true; hAlign = 'center';
                        } else {
                            const ss = _xlStatusStyle(cell.text);
                            if (ss) {
                                fillArgb = ss.fill; fontArgb = ss.font;
                                bold = true; hAlign = 'center';
                            } else {
                                fillArgb = isAlt ? XL.rowAlt : XL.rowNormal;
                                fontArgb = 'FF111827';
                                if (cell.text && !isNaN(cell.text.replace('%', '').trim()) && cell.text.trim() !== '') {
                                    hAlign = 'center';
                                }
                            }
                        }
                    }

                    el.font      = _xlFont(bold, fontArgb || 'FF000000', 10);
                    el.fill      = _xlFill(fillArgb || XL.rowNormal);
                    el.border    = _xlBorder();
                    el.alignment = { vertical: 'middle', horizontal: hAlign, wrapText: !titleCols.has(cell.col) };

                    if (cell.rs > 1 || cell.cs > 1) {
                        ws.mergeCells(absRow, cell.col, absRow + cell.rs - 1, cell.col + cell.cs - 1);
                    }
                    if (cell.cs === 1) {
                        const len = (cell.text || '').length;
                        const capW = titleCols.has(cell.col) ? 45 : 55;
                        const minW = titleCols.has(cell.col) ? 30 : 8;
                        colMaxLen[cell.col] = Math.max(colMaxLen[cell.col] || minW, Math.min(len + 4, capW));
                    }
                });

                for (let r = 1; r <= maxRow; r++) {
                    ws.getRow(r + currentRow - 1).height = r <= hCount ? 22 : 18;
                }
                currentRow += maxRow + 2;  // gap between sections
            }

            // Apply accumulated column widths
            const maxCol = Object.keys(colMaxLen).reduce(function(m, k) { return Math.max(m, parseInt(k)); }, 10);
            for (let c = 1; c <= maxCol; c++) {
                ws.getColumn(c).width = colMaxLen[c] || 14;
            }
        }

        // Section names used when exporting the two summary tabs as single sheets
        const COMBINED_SECTIONS = [
            'Prod Sanity Scenarios',
            'Prod US Sanity',
            'Insprint US Prod Status',
            'Ready for Prod Bugs'
        ];

        async function _buildWorkbook(tabPanels) {
            const wb = new ExcelJS.Workbook();
            wb.creator  = 'Prod Sanity Report';
            wb.created  = new Date();
            wb.modified = new Date();

            const used = new Set();
            for (const tabInfo of tabPanels) {
                const tabEl = document.getElementById(tabInfo.id);
                if (!tabEl) continue;
                const tables = Array.from(tabEl.querySelectorAll('table'));
                if (tables.length === 0) continue;

                // Overall Production Sanity Status and Leadwise Prod Sanity Status:
                // export all tables as one sheet with section headings
                if (tabInfo.id === 'overallProdSanityTab' || tabInfo.id === 'overallStatusTab') {
                    let name = tabInfo.name.length > 31 ? tabInfo.name.substring(0, 31) : tabInfo.name;
                    let unique = name; let n = 2;
                    while (used.has(unique)) { unique = (name.substring(0, 28) + ' ' + n++).substring(0, 31); }
                    used.add(unique);
                    const sections = COMBINED_SECTIONS.slice();
                    while (sections.length < tables.length) sections.push('Section ' + (sections.length + 1));
                    await _combinedTableSheet(wb, unique, tables, sections);
                    continue;
                }

                let idx = 0;
                for (const table of tables) {
                    let name = tables.length === 1
                        ? tabInfo.name
                        : tabInfo.name.substring(0, 27) + ' ' + (idx + 1);
                    name = name.length > 31 ? name.substring(0, 31) : name;
                    let unique = name; let n = 2;
                    while (used.has(unique)) {
                        unique = (name.substring(0, 28) + ' ' + n++).substring(0, 31);
                    }
                    used.add(unique);
                    await _tableToSheet(wb, unique, table);
                    idx++;
                }
            }
            return wb;
        }

        async function _downloadWorkbook(wb, filename) {
            const buf  = await wb.xlsx.writeBuffer();
            const blob = new Blob([buf], {
                type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            });
            const url = URL.createObjectURL(blob);
            const a   = document.createElement('a');
            a.href = url; a.download = filename;
            document.body.appendChild(a); a.click();
            document.body.removeChild(a); URL.revokeObjectURL(url);
        }

        function _truncateSheetName(name) {
            return name.length > 31 ? name.substring(0, 31) : name;
        }

        const TAB_PANELS = [
            { id: 'overallProdSanityTab', name: 'Overall Production Sanity Status' },
            { id: 'prodSanityTab',       name: 'Prod Sanity Scenarios' },
            { id: 'prodUSSanityTab',     name: 'Prod US Sanity' },
            { id: 'insprintStatusTab',   name: 'Insprint US Prod Status' },
            { id: 'readyForProdBugTab',  name: 'Ready for Prod Bug' },
            { id: 'prodSanityDefectsTab', name: 'Prod Sanity Defects' },
            { id: 'overallStatusTab',    name: 'Leadwise Prod Sanity Status' }
        ];

        async function exportCurrentTabToExcel() {
            if (typeof ExcelJS === 'undefined') { alert('Excel export library not loaded. Please check internet connectivity.'); return; }
            const active = TAB_PANELS.find(function(t) {
                const el = document.getElementById(t.id);
                return el && el.classList.contains('active');
            });
            if (!active) return;
            const wb = await _buildWorkbook([active]);
            if (wb.worksheets.length === 0) { alert('No data to export.'); return; }
            const date = new Date().toLocaleDateString('en-GB').split('/').join('-');
            await _downloadWorkbook(wb, active.name + ' - ' + date + '.xlsx');
        }

        async function exportAllTabsToExcel() {
            if (typeof ExcelJS === 'undefined') { alert('Excel export library not loaded. Please check internet connectivity.'); return; }
            const wb = await _buildWorkbook(TAB_PANELS);
            if (wb.worksheets.length === 0) { alert('No data to export.'); return; }
            const date = new Date().toLocaleDateString('en-GB').split('/').join('-');
            await _downloadWorkbook(wb, 'Prod Sanity Report - ' + date + '.xlsx');
        }

        // Initialize on page load
        document.addEventListener('DOMContentLoaded', function() {
            initializeBugData();
            updateStateOptions(); // Initialize state options on load
        });
    </script>
</body>
</html>
"""
        return html
    
    def calculate_leads_summary(self, organized_data):
        """Calculate summary by lead (manual tests only)"""
        leads_summary = defaultdict(lambda: {
            'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0
        })
        
        for lead in organized_data:
            for module in organized_data[lead]:
                manual = organized_data[lead][module]['manual']
                for key in leads_summary[lead]:
                    leads_summary[lead][key] += manual[key]
        
        return leads_summary
    
    def calculate_automation_leads_summary(self, organized_data):
        """Calculate summary by lead (automation tests only)"""
        leads_summary = defaultdict(lambda: {
            'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0
        })
        
        for lead in organized_data:
            for module in organized_data[lead]:
                automation = organized_data[lead][module]['automation']
                for key in leads_summary[lead]:
                    leads_summary[lead][key] += automation[key]
        
        return leads_summary
    
    def calculate_insprint_leads_summary(self, organized_data):
        """Calculate summary by lead (insprint tests only)"""
        leads_summary = defaultdict(lambda: {
            'total': 0, 'passed': 0, 'failed': 0, 'blocked': 0, 'na': 0, 'not_run': 0
        })
        
        for lead in organized_data:
            for module in organized_data[lead]:
                manual = organized_data[lead][module]['manual']
                for key in leads_summary[lead]:
                    leads_summary[lead][key] += manual[key]
        
        return leads_summary
    
    def generate_html_file(self, filename=None, dashboard_style=True):
        """Generate and save HTML report to file
        
        Args:
            filename: Output filename (auto-generated if not provided)
            dashboard_style: If True, generates modern dashboard. If False, generates traditional tabbed report.
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Production_execution_report_{timestamp}.html"
        
        if dashboard_style:
            html_content = self.generate_dashboard_html()
        else:
            html_content = self.generate_html()
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"\n✅ Report generated: {filename}")
        return filename
    
    def generate_dashboard_html(self):
        """Generate modern dashboard-style HTML report with embedded data"""
        # Calculate statistics
        organized_data = self.organize_data_by_lead_module()
        grand_totals = self.calculate_grand_totals(organized_data)
        
        # Calculate percentages
        manual_pass_pct, manual_exec_pct = self.calculate_grand_total_percentages(grand_totals['manual'])
        auto_pass_pct, auto_exec_pct = self.calculate_grand_total_percentages(grand_totals['automation'])
        
        manual_count = sum(1 for t in self.test_data if t['type'].lower() == 'manual')
        auto_count = sum(1 for t in self.test_data if t['type'].lower() == 'automation')
        
        # Outcome counts
        outcomes = {}
        for test in self.test_data:
            outcome = test['outcome']
            outcomes[outcome] = outcomes.get(outcome, 0) + 1
        
        # Bug stats
        allowed_states = {
            'new', 'active', 'blocked', 'ready to deploy', 'resolved', 
            'ba clarification', 're-open', 'blocked in pt', 'blocked in uat', 'deferred'
        }
        filtered_bugs = [bug for bug in self.bug_data if bug['state'].lower() in allowed_states]
        
        # Build HTML
        html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Production Execution Report - {self.suite_name}</title>
    <style>
{self._dashboard_styles()}
    </style>
</head>
<body>
    <div class="container">
        <!-- Header -->
        <div class="header">
            <h1>
                🚀 Production Execution Report
                <span class="static-indicator">SNAPSHOT</span>
            </h1>
            <div class="subtitle">{self.suite_name}</div>
            <div class="last-update">
                Generated: <strong>{self.timestamp}</strong>
            </div>
        </div>

        <!-- Statistics Cards -->
        <div class="stats-grid">
            <div class="stat-card">
                <div class="stat-icon">📊</div>
                <div class="stat-label">Total Tests</div>
                <div class="stat-value">{len(self.test_data):,}</div>
            </div>

            <div class="stat-card">
                <div class="stat-icon">✏️</div>
                <div class="stat-label">Manual Tests</div>
                <div class="stat-value">{manual_count:,}</div>
                <div class="progress-bar">
                    <div class="progress-fill" style="width: {manual_pass_pct:.1f}%"></div>
                </div>
                <div class="subtitle" style="margin-top: 8px;">
                    Pass: <strong>{manual_pass_pct:.1f}%</strong> | 
                    Exec: <strong>{manual_exec_pct:.1f}%</strong>
                </div>
            </div>

            <div class="stat-card">
                <div class="stat-icon">🤖</div>
                <div class="stat-label">Automation Tests</div>
                <div class="stat-value">{auto_count:,}</div>
                <div class="progress-bar">
                    <div class="progress-fill" style="width: {auto_pass_pct:.1f}%"></div>
                </div>
                <div class="subtitle" style="margin-top: 8px;">
                    Pass: <strong>{auto_pass_pct:.1f}%</strong> | 
                    Exec: <strong>{auto_exec_pct:.1f}%</strong>
                </div>
            </div>

            <div class="stat-card">
                <div class="stat-icon">🐛</div>
                <div class="stat-label">Active Bugs</div>
                <div class="stat-value">{len(filtered_bugs):,}</div>
                <div class="subtitle" style="margin-top: 8px;">
                    From query: <strong>{len(self.bug_data)}</strong>
                </div>
            </div>
        </div>

        <!-- Test Outcomes Table -->
        <div class="table-container">
            <h2>📈 Test Outcomes Summary</h2>
            <table>
                <thead>
                    <tr>
                        <th>Outcome</th>
                        <th>Count</th>
                        <th>Percentage</th>
                    </tr>
                </thead>
                <tbody>
{self._generate_outcomes_rows(outcomes, len(self.test_data))}
                </tbody>
            </table>
        </div>

        <!-- Lead/Module Breakdown -->
        <div class="table-container">
            <h2>👥 Test Execution by Lead & Module</h2>
            <table>
                <thead>
                    <tr>
                        <th>Lead</th>
                        <th>Module</th>
                        <th>Type</th>
                        <th>Total</th>
                        <th>Passed</th>
                        <th>Failed</th>
                        <th>Blocked</th>
                        <th>Pass %</th>
                        <th>Exec %</th>
                    </tr>
                </thead>
                <tbody>
{self._generate_lead_module_rows(organized_data)}
                </tbody>
            </table>
        </div>

        <!-- Bugs Table -->
        {self._generate_bugs_section(filtered_bugs)}
    </div>
</body>
</html>'''
        return html
    
    def _dashboard_styles(self):
        """Return CSS styles for dashboard"""
        return '''* {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }

        .container {
            max-width: 1400px;
            margin: 0 auto;
        }

        .header {
            background: white;
            border-radius: 12px;
            padding: 30px;
            margin-bottom: 20px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.1);
        }

        .header h1 {
            color: #2d3748;
            font-size: 28px;
            margin-bottom: 10px;
            display: flex;
            align-items: center;
            gap: 10px;
        }

        .static-indicator {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            background: #3b82f6;
            color: white;
            padding: 6px 14px;
            border-radius: 20px;
            font-size: 14px;
            font-weight: 600;
        }

        .subtitle {
            color: #718096;
            font-size: 14px;
            margin-top: 8px;
        }

        .last-update {
            color: #4a5568;
            font-size: 14px;
            margin-top: 12px;
            font-weight: 500;
        }

        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 20px;
        }

        .stat-card {
            background: white;
            border-radius: 12px;
            padding: 24px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.1);
            transition: transform 0.3s ease;
        }

        .stat-card:hover {
            transform: translateY(-5px);
        }

        .stat-value {
            font-size: 36px;
            font-weight: bold;
            color: #2d3748;
            margin: 8px 0;
        }

        .stat-label {
            color: #718096;
            font-size: 14px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        .stat-icon {
            font-size: 24px;
            margin-bottom: 8px;
        }

        .progress-bar {
            height: 8px;
            background: #e2e8f0;
            border-radius: 4px;
            overflow: hidden;
            margin-top: 12px;
        }

        .progress-fill {
            height: 100%;
            background: linear-gradient(90deg, #10b981, #059669);
        }

        .table-container {
            background: white;
            border-radius: 12px;
            padding: 30px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.1);
            overflow-x: auto;
            margin-bottom: 20px;
        }

        .table-container h2 {
            color: #2d3748;
            margin-bottom: 20px;
            font-size: 22px;
        }

        table {
            width: 100%;
            border-collapse: collapse;
        }

        thead th {
            background: #f7fafc;
            color: #2d3748;
            font-weight: 600;
            text-align: left;
            padding: 12px 16px;
            border-bottom: 2px solid #e2e8f0;
            font-size: 14px;
        }

        tbody td {
            padding: 12px 16px;
            border-bottom: 1px solid #e2e8f0;
            color: #4a5568;
            font-size: 14px;
        }

        tbody tr:hover {
            background: #f7fafc;
        }

        .badge {
            display: inline-block;
            padding: 4px 10px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 600;
        }

        .badge-passed { background: #d1fae5; color: #065f46; }
        .badge-failed { background: #fee2e2; color: #991b1b; }
        .badge-blocked { background: #fef3c7; color: #92400e; }
        .badge-na { background: #e0e7ff; color: #3730a3; }
        .badge-notrun { background: #f3f4f6; color: #374151; }
        
        .text-center { text-align: center; }
        .text-right { text-align: right; }
        .font-bold { font-weight: 700; }'''
    
    def _generate_outcomes_rows(self, outcomes, total):
        """Generate HTML rows for outcomes table"""
        sorted_outcomes = sorted(outcomes.items(), key=lambda x: x[1], reverse=True)
        rows = []
        for outcome, count in sorted_outcomes:
            percentage = (count / total * 100) if total > 0 else 0
            badge_class = self._get_badge_class(outcome)
            rows.append(f'''                    <tr>
                        <td><span class="badge {badge_class}">{outcome}</span></td>
                        <td class="font-bold">{count:,}</td>
                        <td>{percentage:.1f}%</td>
                    </tr>''')
        return '\n'.join(rows)
    
    def _get_badge_class(self, outcome):
        """Get CSS class for outcome badge"""
        outcome_lower = outcome.lower()
        if 'pass' in outcome_lower:
            return 'badge-passed'
        elif 'fail' in outcome_lower:
            return 'badge-failed'
        elif 'block' in outcome_lower:
            return 'badge-blocked'
        elif 'not applicable' in outcome_lower or outcome_lower == 'na':
            return 'badge-na'
        else:
            return 'badge-notrun'
    
    def _generate_lead_module_rows(self, organized_data):
        """Generate HTML rows for lead/module breakdown"""
        rows = []
        for lead, modules in sorted(organized_data.items()):
            for module, types in sorted(modules.items()):
                for test_type in ['manual', 'automation']:
                    data = types[test_type]
                    if data['total'] > 0:
                        pass_pct, exec_pct = self.calculate_percentages(data)
                        rows.append(f'''                    <tr>
                        <td>{lead}</td>
                        <td>{module}</td>
                        <td style="text-transform: capitalize;">{test_type}</td>
                        <td class="text-center font-bold">{data['total']}</td>
                        <td class="text-center">{data['passed']}</td>
                        <td class="text-center">{data['failed']}</td>
                        <td class="text-center">{data['blocked']}</td>
                        <td class="text-center font-bold">{pass_pct:.1f}%</td>
                        <td class="text-center font-bold">{exec_pct:.1f}%</td>
                    </tr>''')
        return '\n'.join(rows) if rows else '<tr><td colspan="9" class="text-center">No data available</td></tr>'
    
    def _generate_bugs_section(self, bugs):
        """Generate bugs table section"""
        if not bugs:
            return ''
        
        rows = []
        for bug in bugs[:50]:  # Limit to 50 bugs
            rows.append(f'''                    <tr>
                        <td><a href="https://dev.azure.com/accenturecio08/AutomationProcess_29697/_workitems/edit/{bug['id']}" target="_blank">#{bug['id']}</a></td>
                        <td>{bug['title']}</td>
                        <td><span class="badge badge-na">{bug['state']}</span></td>
                        <td><span class="badge badge-{'failed' if '1 -' in bug['severity'] else 'blocked'}">{bug['severity']}</span></td>
                        <td>{bug['assigned_to']}</td>
                    </tr>''')
        
        return f'''        <div class="table-container">
            <h2>🐛 Active Bugs ({len(bugs)} total)</h2>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Title</th>
                        <th>State</th>
                        <th>Severity</th>
                        <th>Assigned To</th>
                    </tr>
                </thead>
                <tbody>
{chr(10).join(rows)}
                </tbody>
            </table>
        </div>'''
    
    def export_to_json(self, filename="latest_report.json"):
        """Export report data to JSON format for GitHub-hosted live dashboard"""
        print(f"\n📊 Exporting data to JSON...")
        
        # Organize data
        organized_data = self.organize_data_by_lead_module()
        grand_totals = self.calculate_grand_totals(organized_data)
        
        # Calculate statistics
        manual_count = sum(1 for t in self.test_data if t['type'].lower() == 'manual')
        auto_count = sum(1 for t in self.test_data if t['type'].lower() == 'automation')
        
        # Outcome counts
        outcomes = {}
        for test in self.test_data:
            outcome = test['outcome']
            outcomes[outcome] = outcomes.get(outcome, 0) + 1
        
        # Filter bugs (same logic as report)
        allowed_states = {
            'new', 'active', 'blocked', 'ready to deploy', 'resolved', 
            'ba clarification', 're-open', 'blocked in pt', 'blocked in uat', 'deferred'
        }
        filtered_bugs = [bug for bug in self.bug_data if bug['state'].lower() in allowed_states]
        
        # Build JSON structure
        export_data = {
            "generated_at": datetime.now().isoformat(),
            "timestamp_display": self.timestamp,
            "suite_name": self.suite_name,
            "statistics": {
                "total_tests": len(self.test_data),
                "manual_tests": manual_count,
                "automation_tests": auto_count,
                "outcomes": outcomes,
                "grand_totals": {
                    "manual": {
                        "total": grand_totals['manual']['total'],
                        "passed": grand_totals['manual']['passed'],
                        "failed": grand_totals['manual']['failed'],
                        "blocked": grand_totals['manual']['blocked'],
                        "na": grand_totals['manual']['na'],
                        "not_run": grand_totals['manual']['not_run'],
                        "pass_percentage": round(self.calculate_grand_total_percentages(grand_totals['manual'])[0], 2),
                        "execution_percentage": round(self.calculate_grand_total_percentages(grand_totals['manual'])[1], 2)
                    },
                    "automation": {
                        "total": grand_totals['automation']['total'],
                        "passed": grand_totals['automation']['passed'],
                        "failed": grand_totals['automation']['failed'],
                        "blocked": grand_totals['automation']['blocked'],
                        "na": grand_totals['automation']['na'],
                        "not_run": grand_totals['automation']['not_run'],
                        "pass_percentage": round(self.calculate_grand_total_percentages(grand_totals['automation'])[0], 2),
                        "execution_percentage": round(self.calculate_grand_total_percentages(grand_totals['automation'])[1], 2)
                    }
                }
            },
            "bugs": {
                "total_from_query": len(self.bug_data),
                "filtered_count": len(filtered_bugs),
                "bugs_list": filtered_bugs
            },
            "tests_by_lead_module": {},
            "test_details": self.test_data[:100]  # First 100 tests for dashboard preview
        }
        
        # Add organized data by lead/module
        for lead, modules in organized_data.items():
            export_data["tests_by_lead_module"][lead] = {}
            for module, types in modules.items():
                export_data["tests_by_lead_module"][lead][module] = {
                    "manual": {
                        **types['manual'],
                        "percentages": {
                            "pass": round(self.calculate_percentages(types['manual'])[0], 2),
                            "execution": round(self.calculate_percentages(types['manual'])[1], 2)
                        }
                    },
                    "automation": {
                        **types['automation'],
                        "percentages": {
                            "pass": round(self.calculate_percentages(types['automation'])[0], 2),
                            "execution": round(self.calculate_percentages(types['automation'])[1], 2)
                        }
                    }
                }
        
        # Save to file
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        print(f"   ✅ JSON exported: {filename}")
        print(f"   📏 File size: {os.path.getsize(filename) / 1024:.2f} KB")
        return filename

    @staticmethod
    def save_to_onedrive_sync(local_file, sharepoint_sync_folder):
        """Copy the HTML file to the OneDrive-synced SharePoint folder."""
        if not os.path.exists(sharepoint_sync_folder):
            print(f"\n❌ SharePoint sync folder does not exist: {sharepoint_sync_folder}")
            print("   Please sync the folder using OneDrive first.")
            return None
        dest_file = os.path.join(sharepoint_sync_folder, os.path.basename(local_file))
        shutil.copy2(local_file, dest_file)
        print(f"\n✅ Report also copied to SharePoint sync folder:\n   {dest_file}")
        print("   OneDrive will sync this file to SharePoint automatically.")
        return dest_file

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution flow"""
    print("=" * 80)
    print("🚀 AZURE DEVOPS TEST EXECUTION REPORT GENERATOR")
    print("=" * 80)
    
    # Initialize client
    client = AzureDevOpsClient(ADO_CONFIG)
    
    # Test database connection (if available)
    if DB_AVAILABLE:
        test_database_connection()
    
    # Test connection
    if not client.test_connection():
        print("\n❌ Connection failed. Please check your configuration.")
        return
    
    # Get test plan info
    plan_info = client.get_test_plan()
    if not plan_info:
        print("\n❌ Could not fetch test plan information.")
        return
    
    # Verify target suite exists
    suite_info = client.verify_suite_exists(ADO_CONFIG['suite_id'])
    if not suite_info:
        print(f"\n⚠️  Warning: Suite {ADO_CONFIG['suite_id']} not directly accessible")
        print("   Attempting to find suite in plan hierarchy...")
        
        all_suites = client.get_all_suites_in_plan()
        target_suite = None
        
        for suite in all_suites:
            if str(suite.get('id')) == str(ADO_CONFIG['suite_id']):
                target_suite = suite
                break
            elif suite.get('name', '').lower() == ADO_CONFIG['target_suite_name'].lower():
                target_suite = suite
                ADO_CONFIG['suite_id'] = suite.get('id')
                break
        
        if target_suite:
            suite_info = target_suite
            print(f"   ✅ Found suite: {suite_info.get('name')}")
        else:
            print(f"\n❌ Could not find suite '{ADO_CONFIG['target_suite_name']}'")
            print("\n📋 Available suites in this plan:")
            for suite in all_suites[:20]:
                print(f"   - {suite.get('name')} (ID: {suite.get('id')})")
            return
    
    suite_name = suite_info.get('name', ADO_CONFIG['target_suite_name'])
    suite_id = suite_info.get('id', ADO_CONFIG['suite_id'])
    
    # Collect test data
    test_data = client.get_all_test_data_from_suite(suite_id, suite_name)
    
    if not test_data:
        print("\n⚠️  No test data found in the suite.")
        return
    
    # Collect Insprint test data
    insprint_suite_id = ADO_CONFIG.get('insprint_suite_id')
    insprint_suite_name = ADO_CONFIG.get('insprint_suite_name', 'Insprint Execution')
    insprint_data = []
    
    if insprint_suite_id:
        print(f"\n📊 Collecting Insprint Test Data from Suite {insprint_suite_id}...")
        insprint_suite_info = client.verify_suite_exists(insprint_suite_id)
        
        if insprint_suite_info:
            insprint_data = client.get_all_test_data_from_suite(insprint_suite_id, insprint_suite_name)
            print(f"   ✅ Found {len(insprint_data)} Insprint test items")
        else:
            print(f"   ⚠️  Could not access Insprint suite {insprint_suite_id}")
    
    # Fetch Prod Sanity Scenarios data
    prod_sanity_suite_id = ADO_CONFIG.get('prod_sanity_suite_id')
    prod_sanity_suite_name = ADO_CONFIG.get('prod_sanity_suite_name', 'Prod Sanity Scenarios')
    prod_sanity_data = []
    if prod_sanity_suite_id:
        print(f"\n📊 Collecting Prod Sanity Scenarios data from Suite {prod_sanity_suite_id}...")
        ps_suite_info = client.verify_suite_exists(prod_sanity_suite_id)
        if ps_suite_info:
            prod_sanity_data = client.get_all_test_data_from_suite(prod_sanity_suite_id, prod_sanity_suite_name)
            print(f"   ✅ Found {len(prod_sanity_data)} Prod Sanity test items")
        else:
            print(f"   ⚠️  Could not access Prod Sanity suite {prod_sanity_suite_id}")

    # Fetch bug data from query
    bug_query_id = '3e2de1af-6804-4b73-98e8-d3f51beab824'
    bug_data = client.get_bugs_from_query(bug_query_id)

    # Fetch Prod Sanity Defects from dedicated query
    prod_sanity_defects_query_id = '81998d66-995b-4cec-82bc-f4ed86737e0f'
    prod_sanity_defects = client.get_bugs_from_query(prod_sanity_defects_query_id)

    # Fetch US-linked bugs for the Prod US Sanity tab (Bug ID / State / Severity)
    us_bug_query_id = '7afa36ec-1e4e-4a60-89f1-21fe206ce4c0'
    us_bug_map = client.get_us_bug_map_from_query(us_bug_query_id)

    # Fetch defects with Insprint_Regression and Automation Regression tags created after Feb 12, 2026
    print(f"\n📊 Fetching Insprint_Regression and Automation Regression defects...")
    insprint_defects = client.get_defects_by_tag_and_date(
        tags=['Insprint_Regression', 'Automation Regression'],
        created_after_date='2026-02-12'
    )
    
    # Note: Keep insprint_defects separate for the dedicated tab
    # Don't merge with bug_data
    
    # Save to database (if available)
    if DB_AVAILABLE:
        try:
            print(f"\n💾 Saving data to PostgreSQL database...")
            save_test_data_to_db(test_data, suite_name)
            print(f"   ✅ Data saved successfully to database")
        except Exception as e:
            print(f"   ⚠️  Database save failed: {e}")
            print(f"   ℹ️  Continuing with report generation...")
    
    # Generate HTML report
    print(f"\n📝 Generating HTML Reports...")
    report_gen = CustomHTMLReportGenerator(
        test_data,
        plan_info,
        suite_name,
        bug_data,
        insprint_data,
        insprint_defects,
        prod_sanity_data,
        prod_sanity_defects,
        us_bug_map=us_bug_map,
    )
    
    # Export data to JSON for GitHub-hosted live dashboard
    json_file = report_gen.export_to_json()
    
    # Generate dashboard-style HTML report (timestamped)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = f"Production_execution_report_{timestamp}.html"
    print(f"   → Generating {report_file} (dashboard style)...")
    report_gen.generate_html_file(filename=report_file, dashboard_style=True)
    
    # Generate standalone report with embedded JSON for local viewing
    print(f"   → Generating standalone_report.html (embedded data for local viewing)...")
    try:
        import create_standalone_report
        if create_standalone_report.create_standalone_report('latest_report.json', 'standalone_report.html'):
            print(f"   ✓ Created: standalone_report.html")
    except Exception as e:
        print(f"   ⚠️  Could not generate standalone report: {e}")
        print(f"   ℹ️  Run 'python create_standalone_report.py' manually if needed")
    
    # Note: live_report.html is a template that fetches latest_report.json
    #       It should NOT be overwritten - it's for GitHub Pages deployment
    
    # --- Add this block after report_file is generated ---
    # Set your local OneDrive sync folder path here:
    sharepoint_sync_folder = r"C:\Users\hajara.ayyubkhan\OneDrive - Accenture\Prod\Prod Execution Report"
    report_gen.save_to_onedrive_sync(report_file, sharepoint_sync_folder)
    # -----------------------------------------------------
    
    print("\n" + "=" * 80)
    print("✅ REPORT GENERATION COMPLETED")
    print("=" * 80)
    print(f"\n📄 Generated Files:")
    print(f"   ✅ {report_file} - Dashboard snapshot (opens directly)")
    print(f"   ✅ standalone_report.html - Works with file:// protocol")
    print(f"   ✅ latest_report.json - Data file (124 KB)")
    print(f"\n🌐 GitHub Pages Live Dashboard:")
    print(f"   📊 live_report.html - Auto-refreshing template")
    print(f"   🔗 https://vishnuramalingam07.github.io/Myisp_Tools/live_report.html")
    print(f"\n📊 Total Test Cases: {len(test_data)}")
    
    # Summary statistics
    manual_count = sum(1 for t in test_data if t['type'].lower() == 'manual')
    auto_count = sum(1 for t in test_data if t['type'].lower() == 'automation')
    
    print(f"   - Manual: {manual_count}")
    print(f"   - Automation: {auto_count}")
    
    # Outcome summary
    outcomes = {}
    for test in test_data:
        outcome = test['outcome']
        outcomes[outcome] = outcomes.get(outcome, 0) + 1
    
    print(f"\n📈 Outcome Summary:")
    for outcome, count in sorted(outcomes.items()):
        print(f"   - {outcome}: {count}")
    
    # Bug summary
    if bug_data:
        allowed_states = {
            'new', 'active', 'blocked', 'ready to deploy', 'resolved', 
            'ba clarification', 're-open', 'blocked in pt', 'blocked in uat', 'deferred'
        }
        filtered_bugs = [bug for bug in bug_data if bug['state'].lower() in allowed_states]
        
        print(f"\n🐛 Bug Summary:")
        print(f"   - Total Bugs from Query: {len(bug_data)}")
        print(f"   - Bugs in Report (New/Active/Blocked/etc.): {len(filtered_bugs)}")
        
        severity_counts = {}
        for bug in filtered_bugs:
            severity = bug['severity']
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        for severity, count in sorted(severity_counts.items()):
            print(f"   - {severity}: {count}")
    
    # Insprint Defects summary
    if insprint_defects:
        print(f"\n🏷️  Regression Defects Summary (Insprint_Regression & Automation Regression):")
        print(f"   - Total Defects (created after Feb 12, 2026): {len(insprint_defects)}")
        
        # State counts
        state_counts = {}
        for defect in insprint_defects:
            state = defect['state']
            state_counts[state] = state_counts.get(state, 0) + 1
        
        print(f"   - By State:")
        for state, count in sorted(state_counts.items()):
            print(f"      • {state}: {count}")
        
        # Severity counts
        severity_counts = {}
        for defect in insprint_defects:
            severity = defect['severity']
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        print(f"   - By Severity:")
        for severity, count in sorted(severity_counts.items()):
            print(f"      • {severity}: {count}")
    
    print("\n" + "=" * 80)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Process interrupted by user")
    except Exception as e:
        print(f"\n\n❌ An error occurred: {e}")
        import traceback
        traceback.print_exc()